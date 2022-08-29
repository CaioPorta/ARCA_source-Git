import os
import sys
import ctypes

import time
from datetime import datetime
from datetime import timedelta
from calendar import monthrange
from shutil import copyfile

import tkinter
from tkinter import filedialog

import xlsxwriter
from openpyxl import load_workbook
import win32com.client as win32

from PIL import Image

import sqlite3
import random

import pandas as pd

from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *

import _thread
import threading
from threading import Lock
from BackgroundProcesses import BackgroundProcess as BP
from BackgroundProcesses import WorkerThread

import HMI_Trades
import HMI_Tributacao
import HMI_ARCA
import HMI_Desempenho
import HMI_Mercado
import DBManager
import YahooFinance
import Calcular
import LoadingWindow

from CalendarWindow import CalendarWindow as Calendario
from RelogioWindow import RelogioWindow as Relogio
import OperacaoRegistradaComSucesso

class IHMmain(QWidget):
    #%% Init
    def __init__(self):
        self.APPVersion = "Versão Beta 0.0.0"
        super().__init__()
        self.setObjectName("IHM")

# Initial parameters definition

# HMI fixed parameters
        user32 = ctypes.windll.user32
        screensize = user32.GetSystemMetrics(0), user32.GetSystemMetrics(1)
        self.screen_height = screensize[1] # FHD
        self.screen_width = screensize[0] # FHD
        self.font8 = QFont('Times New Roman', 8)
        self.font8.setBold(False)
        self.font9 = QFont('Times New Roman', 9)
        self.font9.setBold(False)
        self.font10 = QFont('Times New Roman', 10)
        self.font10.setBold(False)
        self.font12 = QFont('Times New Roman', 12)
        self.font12.setBold(False)
        self.font14 = QFont('Times New Roman', 14)
        self.font16 = QFont('Times New Roman', 16)
        self.font18 = QFont('Times New Roman', 18)
        self.font20 = QFont('Times New Roman', 20, QFont.Bold)
        self.font22 = QFont('Times New Roman', 22, QFont.Bold)
        self.font24 = QFont('Cooper Black', 24, QFont.Bold)
        self.font26 = QFont('Cooper Black', 26, QFont.Bold)
        self.font28 = QFont('Cooper Black', 28, QFont.Bold)
        self.font30 = QFont('Cooper Black', 30, QFont.Bold)

# Layout creation

        # Window creation
        self.setWindowTitle("Finanças")
        self.setWindowIcon(QIcon(".\images\logoARCA.png"))
        self.setAutoFillBackground(True)
        background = self.palette()
        background.setColor(self.backgroundRole(), QColor(str('#%02x%02x%02x' % (5, 5, 15))))
        self.setPalette(background)
        self.setMinimumSize(QSize(int(self.screen_width*0.5), int(self.screen_height*0.6)))
        self.showMaximized()

        self.GLayout = QGridLayout()
        self.setLayout(self.GLayout)

        self.installEventFilter(self)

        # Inicialização programa
        self.F_GLayout = [] # GridLayouts que ficam dentro do GLayout
        self.GCount = 0 # Contador de GridLayouts que tem no GLayout
        self.HBoxCount = 0 # Contador de HBoxLayouts que tem no GLayout
        self.VBoxCount = 0 # Contador de VBoxLayouts que tem no GLayout
        self.PageID = '0'
        self.PageIDAux = ''
        self.OutraJanelaAberta = False
        self.ClockWindowOpen = False
        self.CalendarWindowOpen = False

        # Inicialização usuário
        self.idxCorretora = 0 # Memoriza qual corretora que estava sendo observada antes de trocar de página
        self.idxBanco = 0
        self.idxPapel = 0 # Memoriza qual papel que estava sendo observado antes de trocar de página
        self.idxAtivo = 0
        self.idxTipoDeAtivo = 0
        self.idxSubtipoDeAtivo = 0
        self.idxSetorAtivo = 0
        self.DTeST = "NotDefined"
        self.Patrimonio = 'calculando...'
        self.Rendimento = 'calculando...'
        self.SumContasCorrentes = 0.
        self.AtualizarCotacoes_running = False
        self.ThreadLock = Lock()

        self.LoggedIn = False
        self.ShowValues = True
        self.filedialogIsOpen = False
        self.ParteImportacao = 1

        self.YahooFinance = YahooFinance.YahooFinance(self)
        self.DBManager = DBManager.DBManager(self)
        self.Calcular = Calcular.Calcular(self)

        self.HMI_Trades = HMI_Trades.HMI_Trades(self)
        self.HMI_Tributacao = HMI_Tributacao.HMI_Tributacao(self)
        self.HMI_ARCA = HMI_ARCA.HMI_ARCA(self)
        self.HMI_Desempenho = HMI_Desempenho.HMI_Desempenho(self)
        self.HMI_Mercado = HMI_Mercado.HMI_Mercado(self)

        # self.PageID = '1' # Debug
        # self.PageIDAux = '' # Debug
        # self.User = 'v' # Debug
        # self.Email = 'moreiraporta@gmail.com' # Debug
        # self.DBManager.ConnectTo(self.User, '000a')  # Debug
        # self.LoggedIn = True
        # self.CreatePage('1') # Debug
        # self.DBManager.UpdatePatrimonioRendimento()
        # self.DBManager.AtualizarCotacoes()

        # self.PageID = '1' # Debug
        # self.PageIDAux = '' # Debug
        # self.User = 'c' # Debug
        # self.Email = 'caio.porta12@gmail.com' # Debug
        # self.DBManager.ConnectTo(self.User, '000a')  # Debug
        # self.LoggedIn = True
        # self.CreatePage('1') # Debug
        # self.DBManager.UpdatePatrimonioRendimento()
        # self.DBManager.AtualizarCotacoes()

        self.CreatePage('0') # Página inicial do programa

    #%% Pages Creation
    def CreatePage_Decisao(self):
        if self.PageID == '26':
            if len(self.TextBox_NomeNovoAtivo.text())>0 and len(self.TextBox_NomeNovoAtivo.text())<7 and not self.TextBox_NomeNovoAtivo.text().lower() == "usd":
                self.setCursor(Qt.WaitCursor)
                IsShare, self.Cache_Cotacao = self.YahooFinance.isShare(self.TextBox_NomeNovoAtivo.text())
                self.unsetCursor()
                if IsShare:
                    self.CreatePage('34a')
                else:
                    self.CreatePage('35a')
            else:
                try: self.Label_Erro1.deleteLater()
                except: pass
                self.Label_Erro1 = QLabel()
                self.Label_Erro1.setStyleSheet('color: red')
                self.Label_Erro1.setFont(self.font12)
                self.Label_Erro1.setAlignment(Qt.AlignCenter)
                self.Label_Erro1.setText("Insira um ticker válido com até 6 caracteres.\nPara 'USD' utilize 'USDT' ou equivalente.") # Substitui a string do erro mais provável
                self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 17, 0, 1, 3, Qt.AlignHCenter | Qt.AlignTop)

        elif self.PageID == '27':
            if len(self.TextBox_NovoNomeAtivo.text())>0 and len(self.TextBox_NovoNomeAtivo.text())<7 and not self.TextBox_NovoNomeAtivo.text().lower() == "usd":
                self.setCursor(Qt.WaitCursor)
                IsShare = self.YahooFinance.isShare(self.TextBox_NovoNomeAtivo.text())
                self.unsetCursor()
                if IsShare:
                    self.CreatePage('34b')
                else:
                    self.CreatePage('35b')
            else:
                try: self.Label_Erro1.deleteLater()
                except: pass
                self.Label_Erro1 = QLabel()
                self.Label_Erro1.setStyleSheet('color: red')
                self.Label_Erro1.setFont(self.font12)
                self.Label_Erro1.setAlignment(Qt.AlignCenter)
                self.Label_Erro1.setText("Insira um ticker válido com até 6 caracteres.\nPara 'USD' utilize 'USDT' ou equivalente.") # Substitui a string do erro mais provável
                self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 17, 0, 1, 3, Qt.AlignHCenter | Qt.AlignTop)

    def CreatePage(self, PageID):
        if self.OutraJanelaAberta:
            MessageBox_Msg1 = QMessageBox.about(self,'Aviso','Existe outra janela aberta.\nFeche a janela aberta e tente novamente.')
            if self.CalendarWindowOpen:
                self.Window_Calendario.close()
                self.Window_Calendario.show()
            elif self.ClockWindowOpen:
                self.Window_Relogio.close()
                self.Window_Relogio.show()
        else:
            #%% PageID == '0'
            if PageID == '0':
                if self.PageIDAux == '17b':
                    if self.TextBox_CodigoDeConfirmacao.text() == self.CodigoDeConfirmacaoGerado:
                        if not self.DBManager.FLAG and not self.AtualizarCotacoes_running:
                            self.LoggedIn = self.DBManager.DeleteUserDB()
                            if not self.LoggedIn:
                                if os.path.exists(os.getcwd()+"\\images\\ProfilePhoto_"+self.User+".jpg"):
                                    os.remove(os.getcwd()+"\\images\\ProfilePhoto_"+self.User+".jpg")
                                self.ClearPage()
                                self.PageID = '0'
                                self.PageIDAux = ''
                                self.InsertGridLayout(1, 0, 1, 20)
                                self.CreatePage0()
                                self.PutHeader()
                        else:
                            MessageBox_Msg1 = QMessageBox.about(self,'Aviso','Existe outra operação em andamento.\nPode ser que a tabela de cotações\nesteja sendo atualizada\n\nTente novamente mais tarde.')
                    else:
                        try: self.Label_Erro1.deleteLater()
                        except: pass
                        self.Label_Erro1 = QLabel()
                        self.Label_Erro1.setStyleSheet('color: red')
                        self.Label_Erro1.setFont(self.font12)
                        self.Label_Erro1.setAlignment(Qt.AlignCenter)
                        self.Label_Erro1.setText('O código não confere.')
                        self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 13, 0, 1, 1, Qt.AlignCenter)
                        self.TextBox_CodigoDeConfirmacao.setFocus()
                else:
                    self.LoggedIn = False

                    # Inicialização usuário
                    self.idxCorretora = 0 # Memoriza qual corretora que estava sendo observada antes de trocar de página
                    self.idxBanco = 0
                    self.idxPapel = 0 # Memoriza qual papel que estava sendo observado antes de trocar de página
                    self.idxAtivo = 0
                    self.idxTipoDeAtivo = 0
                    self.idxSubtipoDeAtivo = 0
                    self.idxSetorAtivo = 0
                    self.DTeST = "NotDefined"
                    self.Patrimonio = 'calculando...'
                    self.Rendimento = 'calculando...'
                    self.SumContasCorrentes = 0.
                    self.HMI_ARCA.Flag_RecalcularGraficos = True

                    self.ClearPage()
                    self.PageID = '0'
                    self.PageIDAux = ''
                    self.InsertGridLayout(1, 0, 1, 20)
                    self.CreatePage0()
                    self.PutHeader()
            #%% PageID == '1'
            elif PageID == '1':
                if self.PageID == '7':
                    if (self.TextBox_NovaSenha1.text() == self.TextBox_NovaSenha2.text() and
                        len(self.TextBox_NovaSenha1.text())>3): # Condições para passar pra página 1
                            self.LoggedIn = self.DBManager.ConnectTo(self.User, self.DBManager.GetPasswdByUser(self.User))
                            self.DBManager.ChangePassword(self.TextBox_NovaSenha1.text())
                            self.ClearPage()
                            self.PageID = '1'
                            self.PageIDAux = ''
                            self.InsertGridLayout(1, 0, 1, 20)
                            self.CreatePage1()
                            self.PutHeader()
                            self.DBManager.UpdatePatrimonioRendimento()
                            self.DBManager.AtualizarCotacoes()
                    else:
                        if not self.TextBox_NovaSenha1.text() == self.TextBox_NovaSenha2.text():
                            try: self.Label_Erro1.deleteLater()
                            except: pass
                            self.Label_Erro1 = QLabel()
                            self.Label_Erro1.setStyleSheet('color: red')
                            self.Label_Erro1.setFont(self.font12)
                            self.Label_Erro1.setAlignment(Qt.AlignCenter)
                            self.Label_Erro1.setText('As senhas digitadas não são iguais.')
                            self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 13, 0, 1, 1, Qt.AlignCenter)
                            self.TextBox_NovaSenha1.setFocus()
                        elif not len(self.TextBox_NovaSenha1.text())>3:
                            try: self.Label_Erro1.deleteLater()
                            except: pass
                            self.Label_Erro1 = QLabel()
                            self.Label_Erro1.setStyleSheet('color: red')
                            self.Label_Erro1.setFont(self.font12)
                            self.Label_Erro1.setAlignment(Qt.AlignCenter)
                            self.Label_Erro1.setText('A senha tem que ter pelo menos 4 caracteres.')
                            self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 13, 0, 1, 1, Qt.AlignCenter)
                            self.TextBox_NovaSenha1.setFocus()
                elif self.PageID == '8':
                    if self.TextBox_CodigoDeConfirmacao.text() == self.CodigoDeConfirmacaoGerado: # Condições para passar pra página 1
                        self.LoggedIn = self.DBManager.CriarNovoPerfil()
                        if self.LoggedIn:
                            self.User = self.TextBox_NomeDeUsuario.text()
                            self.Email = self.TextBox_Email.text()
                            if self.DBManager.CheckIfUserExists(self.User) == "Arquivo não encontrado":
                                def UpdateUserPath(UserName, PathNovo):
                                    currdir = os.getcwd()
                                    conn = sqlite3.connect(currdir+'\\Users.db')
                                    cursor = conn.cursor()
                                    cursor.execute("UPDATE Usuários SET Path = (?) Where Usuário = (?)", [(PathNovo), UserName])
                                    conn.commit()
                                    conn.close()
                                self.DBManager.ThreadLock.acquire()
                                UpdateUserPath(self.User, self.TextBox_Path.text())
                                self.DBManager.ThreadLock.release()
                            self.ClearPage()
                            self.PageID = '1'
                            self.PageIDAux = ''
                            self.InsertGridLayout(1, 0, 1, 20)
                            self.CreatePage1()
                            self.PutHeader()
                            self.DBManager.UpdatePatrimonioRendimento()
                            self.DBManager.AtualizarCotacoes()
                        else:
                            try: self.Label_Erro1.deleteLater()
                            except: pass
                            self.Label_Erro1 = QLabel()
                            self.Label_Erro1.setStyleSheet('color: red')
                            self.Label_Erro1.setFont(self.font12)
                            self.Label_Erro1.setAlignment(Qt.AlignCenter)
                            self.Label_Erro1.setText('Ocorreu um erro inesperado. Tente novamente mais tarde.')
                            self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 12, 0, 1, 2, Qt.AlignCenter)
                            self.TextBox_NomeDeUsuario.setFocus()
                    else:
                        try: self.Label_Erro1.deleteLater()
                        except: pass
                        self.Label_Erro1 = QLabel()
                        self.Label_Erro1.setStyleSheet('color: red')
                        self.Label_Erro1.setFont(self.font12)
                        self.Label_Erro1.setAlignment(Qt.AlignCenter)
                        self.Label_Erro1.setText('O código não confere.')
                        self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 12, 0, 1, 2, Qt.AlignCenter)
                        self.TextBox_NomeDeUsuario.setFocus()
                elif self.PageID == '0':
                    Resposta = self.DBManager.ConnectTo(self.TextBox_NomeDeUsuario.text(), self.TextBox_Senha.text())
                    if Resposta == "Arquivo não existe":
                        self.LoggedIn = False
                        try: self.Label_Erro1.deleteLater()
                        except: pass
                        self.Label_Erro1 = QLabel()
                        self.Label_Erro1.setStyleSheet('color: red')
                        self.Label_Erro1.setFont(self.font12)
                        self.Label_Erro1.setAlignment(Qt.AlignCenter)
                        self.Label_Erro1.setText('Usuário não encontrado.\nVerifique se o arquivo está no local declarado.')
                        self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 13, 0, 1, 2, Qt.AlignCenter)
                        self.TextBox_NomeDeUsuario.setFocus()
                    elif Resposta: # Condições para passar pra página 1
                        self.LoggedIn = True
                        self.User = self.DBManager.User
                        self.Email = self.DBManager.Email
                        self.ClearPage()
                        self.PageID = '1'
                        self.PageIDAux = ''
                        self.InsertGridLayout(1, 0, 1, 20)
                        self.CreatePage1()
                        self.PutHeader()
                        self.DBManager.UpdatePatrimonioRendimento()
                        self.DBManager.AtualizarCotacoes()
                    else:
                        self.LoggedIn = False
                        try: self.Label_Erro1.deleteLater()
                        except: pass
                        self.Label_Erro1 = QLabel()
                        self.Label_Erro1.setStyleSheet('color: red')
                        self.Label_Erro1.setFont(self.font12)
                        self.Label_Erro1.setAlignment(Qt.AlignCenter)
                        self.Label_Erro1.setText('Usuário ou senha incorreto.')
                        self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 13, 0, 1, 2, Qt.AlignCenter)
                        self.TextBox_NomeDeUsuario.setFocus()
                else:
                    self.ClearPage()
                    self.PageID = '1'
                    self.PageIDAux = ''
                    self.InsertGridLayout(1, 0, 1, 20)
                    self.CreatePage1()
                    self.PutHeader()
            #%% PageID == '2'
            elif PageID == '2':
                self.ClearPage()
                self.PageID = '2'
                self.PageIDAux = ''
                self.InsertGridLayout(1, 0, 1, 20)
                self.CreatePage2()
                self.PutHeader()
            #%% PageID == '3'
            elif PageID == '3':
                self.ClearPage()
                self.PageID = '3'
                self.PageIDAux = ''
                self.InsertGridLayout(1, 0, 1, 20)
                self.CreatePage3()
                self.PutHeader()
            #%% PageID == '4'
            elif PageID == '4':
                self.ClearPage()
                self.PageID = '4'
                self.PageIDAux = ''
                self.InsertGridLayout(1, 0, 1, 20)
                self.CreatePage4()
                self.PutHeader()
            #%% PageID == '5'
            elif PageID == '5': # Cria a mesma página da 13
                self.ClearPage()
                self.PageID = '5'
                self.PageIDAux = ''
                self.InsertGridLayout(1, 0, 1, 20)
                self.HMI_Mercado.CreatePage5()
                self.PutHeader()
            #%% PageID == '6'
            elif PageID == '6':
                if "@" in self.TextBox_EmailOuUsuario.text(): isEmail = True
                else: isEmail = False

                if isEmail: # Recovering By Email
                    if self.DBManager.CheckIfEmailExists(self.TextBox_EmailOuUsuario.text()):
                        self.Email = self.TextBox_EmailOuUsuario.text()
                        self.User = self.DBManager.GetUserByEmail(self.TextBox_EmailOuUsuario.text())
                        self.ClearPage()
                        self.PageID = '6'
                        self.PageIDAux = ''
                        self.InsertGridLayout(1, 0, 1, 20)
                        self.CreatePage6(self.Email, self.User)
                        self.PutHeader()
                    else:
                        try: self.Label_Erro1.deleteLater()
                        except: pass
                        self.Label_Erro1 = QLabel()
                        self.Label_Erro1.setStyleSheet('color: red')
                        self.Label_Erro1.setFont(self.font12)
                        self.Label_Erro1.setAlignment(Qt.AlignCenter)
                        self.Label_Erro1.setText('Email não encontrado.')
                        self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 11, 0, 1, 1, Qt.AlignCenter)
                        self.TextBox_EmailOuUsuario.setFocus()

                else: # Recovering By User
                    if self.DBManager.CheckIfUserExists(self.TextBox_EmailOuUsuario.text()):
                        self.Email = self.DBManager.GetEmailByUser(self.TextBox_EmailOuUsuario.text())
                        self.User = self.TextBox_EmailOuUsuario.text()
                        self.ClearPage()
                        self.PageID = '6'
                        self.PageIDAux = ''
                        self.InsertGridLayout(1, 0, 1, 20)
                        self.CreatePage6(self.Email, self.User)
                        self.PutHeader()
                    else:
                        try: self.Label_Erro1.deleteLater()
                        except: pass
                        self.Label_Erro1 = QLabel()
                        self.Label_Erro1.setStyleSheet('color: red')
                        self.Label_Erro1.setFont(self.font12)
                        self.Label_Erro1.setAlignment(Qt.AlignCenter)
                        self.Label_Erro1.setText('Usuário não encontrado.')
                        self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 11, 0, 1, 1, Qt.AlignCenter)
                        self.TextBox_EmailOuUsuario.setFocus()
            #%% PageID == '7'
            elif PageID == '7':
                if self.TextBox_CodigoDeConfirmacao.text() == self.CodigoDeConfirmacaoGerado:
                    self.ClearPage()
                    self.PageID = '7'
                    self.PageIDAux = ''
                    self.InsertGridLayout(1, 0, 1, 20)
                    self.CreatePage7()
                    self.PutHeader()
                else:
                    try: self.Label_Erro1.deleteLater()
                    except: pass
                    self.Label_Erro1 = QLabel()
                    self.Label_Erro1.setStyleSheet('color: red')
                    self.Label_Erro1.setFont(self.font12)
                    self.Label_Erro1.setAlignment(Qt.AlignCenter)
                    self.Label_Erro1.setText('Código não confere.')
                    self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 11, 0, 1, 2, Qt.AlignCenter)
                    self.TextBox_CodigoDeConfirmacao.setFocus()
            #%% PageID == '8'
            elif PageID == '8':
                if (len(self.TextBox_NomeDeUsuario.text())>0 and
                    not "@"  in self.TextBox_NomeDeUsuario.text() and
                    not ","  in self.TextBox_NomeDeUsuario.text() and
                    (not self.DBManager.CheckIfUserExists(self.TextBox_NomeDeUsuario.text()) or
                     self.DBManager.CheckIfUserExists(self.TextBox_NomeDeUsuario.text()) == "Arquivo não encontrado") and
                    "@" in self.TextBox_Email.text() and
                    not "'\'" in self.TextBox_Email.text() and
                    not "," in self.TextBox_Email.text() and
                    len(self.TextBox_Email.text())>3 and
                    not self.DBManager.CheckIfEmailExists(self.TextBox_Email.text()) and
                    self.TextBox_Senha1.text() == self.TextBox_Senha2.text() and
                    len(self.TextBox_Senha1.text())>3 and
                    not self.TextBox_Senha1.text().isnumeric() and
                    not self.TextBox_Senha1.text().isalpha()): # Condições para passar pra página 8
                        self.ClearPage()
                        self.PageID = '8'
                        self.PageIDAux = ''
                        self.InsertGridLayout(1, 0, 1, 20)
                        self.CreatePage8()
                        self.PutHeader()
                else:
                        try: self.Label_Erro1.deleteLater()
                        except: pass
                        self.Label_Erro1 = QLabel()
                        self.Label_Erro1.setStyleSheet('color: red')
                        self.Label_Erro1.setFont(self.font12)
                        self.Label_Erro1.setAlignment(Qt.AlignCenter)
                        if not len(self.TextBox_NomeDeUsuario.text()) > 0:
                            self.Label_Erro1.setText('Insira um nome de usuário')
                            self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 6, 1, 1, 1, Qt.AlignLeft)
                            self.TextBox_NomeDeUsuario.setFocus()
                        elif ("@" in self.TextBox_NomeDeUsuario.text() or
                            "," in self.TextBox_NomeDeUsuario.text() or
                            "'\'" in self.TextBox_NomeDeUsuario.text()):
                            self.Label_Erro1.setText("@ '\' , não são permitidos\nno nome de usuário.")
                            self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 6, 1, 1, 1, Qt.AlignLeft)
                            self.TextBox_NomeDeUsuario.setFocus()
                        elif self.DBManager.CheckIfUserExists(self.TextBox_NomeDeUsuario.text()):
                            self.Label_Erro1.setText('Esse nome de usuário já existe.')
                            self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 6, 1, 1, 1, Qt.AlignLeft)
                            self.TextBox_NomeDeUsuario.setFocus()

                        try: self.Label_Erro2.deleteLater()
                        except: pass
                        self.Label_Erro2 = QLabel()
                        self.Label_Erro2.setStyleSheet('color: red')
                        self.Label_Erro2.setFont(self.font12)
                        if (not "@" in self.TextBox_Email.text() or
                            "'\'" in self.TextBox_Email.text() or
                            "," in self.TextBox_Email.text()):
                            self.Label_Erro2.setText("Confira se seu email está correto.\n@ '\' , não são permitidos.")
                            self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro2, 8, 1, 1, 1, Qt.AlignLeft)
                            self.TextBox_Email.setFocus()
                        elif self.DBManager.CheckIfEmailExists(self.TextBox_Email.text()):
                            self.Label_Erro2.setText('Esse endereço de email já possiu uma conta.')
                            self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro2, 8, 1, 1, 1, Qt.AlignLeft)
                            self.TextBox_Email.setFocus()

                        try: self.Label_Erro3.deleteLater()
                        except: pass
                        self.Label_Erro3 = QLabel()
                        self.Label_Erro3.setStyleSheet('color: red')
                        self.Label_Erro3.setFont(self.font12)
                        if not self.TextBox_Senha1.text() == self.TextBox_Senha2.text():
                            self.Label_Erro3.setText('As senhas digitadas não são iguais.')
                            self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro3, 10, 1, 1, 1, Qt.AlignLeft)
                            self.TextBox_Senha1.setFocus()
                        elif not len(self.TextBox_Senha2.text()) > 3:
                            self.Label_Erro3.setText('A senha tem que ter pelo menos 4 caracteres.')
                            self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro3, 10, 1, 1, 1, Qt.AlignLeft)
                            self.TextBox_Senha1.setFocus()
                        elif not self.TextBox_Senha1.text().isnumeric() or not self.TextBox_Senha1.text().isalpha():
                            self.Label_Erro3.setText('A senha precisa ter pelo menos um número e uma letra ou símbolo.')
                            self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro3, 10, 1, 1, 1, Qt.AlignLeft)
                            self.TextBox_NomeDeUsuario.setFocus()
            #%% PageID == '9'
            elif PageID in ['9', '9_']: # Perfil, Tem ID especial
                if PageID == '9_':
                    if self.PageIDAux == '15':
                        if not self.DBManager.FLAG:
                            if self.TextBox_SenhaAtual.text() == self.DBManager.GetPasswdByUser(self.User):
                                if self.TextBox_Senha1.text() == self.TextBox_Senha2.text() and len(self.TextBox_Senha1.text())>3 and not self.TextBox_Senha1.text().isnumeric() and not self.TextBox_Senha1.text().isalpha():
                                    self.DBManager.ChangePassword(self.TextBox_Senha1.text())
                                    self.ClearPage()
                                    self.PageIDAux = '9'
                                    self.InsertGridLayout(1, 0, 1, 3)
                                    self.CreatePage9()
                                    self.PutHeader()
                                else:
                                    try: self.Label_Erro1.deleteLater()
                                    except: pass
                                    self.Label_Erro1 = QLabel()
                                    self.Label_Erro1.setStyleSheet('color: red')
                                    self.Label_Erro1.setFont(self.font12)
                                    self.Label_Erro1.setAlignment(Qt.AlignCenter)
                                    self.Label_Erro1.setText('As senhas digitadas não são iguais.' if not self.TextBox_Senha1.text() == self.TextBox_Senha2.text() else "A senha tem que ter pelo menos 4 caracteres,\nter pelo menos um número e uma letra ou símbolo.")
                                    self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 13, 0, 1, 1, Qt.AlignCenter)
                                    self.TextBox_Senha2.setFocus()
                            else:
                                try: self.Label_Erro1.deleteLater()
                                except: pass
                                self.Label_Erro1 = QLabel()
                                self.Label_Erro1.setStyleSheet('color: red')
                                self.Label_Erro1.setFont(self.font12)
                                self.Label_Erro1.setAlignment(Qt.AlignCenter)
                                self.Label_Erro1.setText('Senha atual não confere.')
                                self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 13, 0, 1, 1, Qt.AlignCenter)
                                self.TextBox_SenhaAtual.setFocus()
                        else:
                            MessageBox_Msg1 = QMessageBox.about(self,'Aviso','Existe outra operação em andamento.\nTente novamente mais tarde.')
                    elif self.PageIDAux == '17a':
                        self.ClearPage()
                        self.PageIDAux = '9'
                        self.InsertGridLayout(1, 0, 1, 3)
                        self.CreatePage9()
                        self.PutHeader()
                else:
                    self.ClearPage()
                    self.PageIDAux = '9'
                    self.InsertGridLayout(1, 0, 1, 3)
                    self.CreatePage9()
                    self.PutHeader()
            #%% PageID == '10'
            elif PageID == '10':
                self.ClearPage()
                self.PageID = '10'
                self.PageIDAux = ''
                self.InsertGridLayout(1, 0, 1, 20)
                self.HMI_Trades.CreatePage10()
                self.PutHeader()
            #%% PageID == '11'
            elif PageID == '11':
                self.ClearPage()
                self.PageID = '11'
                self.PageIDAux = ''
                self.InsertGridLayout(1, 0, 1, 20)
                self.HMI_ARCA.CreatePage11()
                self.PutHeader()
            #%% PageID == '12'
            elif PageID == '12':
                self.ClearPage()
                self.PageID = '12'
                self.PageIDAux = ''
                self.InsertGridLayout(1, 0, 1, 20)
                self.HMI_Desempenho.CreatePage12()
                self.PutHeader()
            #%% PageID == '13'
            elif PageID == '13':
                self.ClearPage()
                self.PageID = '13'
                self.PageIDAux = ''
                self.InsertGridLayout(1, 0, 1, 20)
                self.HMI_Mercado.CreatePage13()
                self.PutHeader()
            #%% PageID == '14'
            elif PageID == '14': MessageBox_Msg1 = QMessageBox.about(self,'Informação','Funcionalidade ainda não implementada.') # Derivação do perfil, Tem ID especial
            #%% PageID == '15'
            elif PageID == '15': # Derivação do perfil, Tem ID especial
                self.ClearPage()
                self.PageIDAux = '15'
                self.InsertGridLayout(1, 0, 1, 20)
                self.CreatePage15()
                self.PutHeader()
            #%% PageID == '16'
            elif PageID == '16': # MessageBox, não tem ID
                self.CreatePage16()
                self.PutHeader()
            #%% PageID == '17a'
            elif PageID == '17a': # Derivação do perfil, Tem ID especial
                self.ClearPage()
                self.PageIDAux = '17a'
                Email = '@'
                NomeDeUsuario = 'c'
                self.InsertGridLayout(1, 0, 1, 20)
                self.CreatePage17a(Email, NomeDeUsuario)
                self.PutHeader()
            #%% PageID == '17b'
            elif PageID == '17b': # Derivação do perfil, Tem ID especial
                self.ClearPage()
                self.PageIDAux = '17b'
                self.InsertGridLayout(1, 0, 1, 20)
                self.CreatePage17b()
                self.PutHeader()
            #%% PageID == '18'
            elif PageID in ['18','18_']:
                if PageID == '18_':
                    if self.PageID == '20':
                        if len(self.TextBox_NovaCorretora.text())>0 and len(self.TextBox_NovaCorretora.text())<9:
                            if not self.DBManager.FLAG:
                                Executado = self.DBManager.AddCorretora()
                                if Executado:
                                    self.ClearPage()
                                    self.PageID = '18'
                                    self.PageIDAux = ''
                                    self.InsertGridLayout(1, 0, 1, 20)
                                    self.HMI_Trades.HMI_Trades_Bolsa.CreatePage18()
                                    self.PutHeader()
                                else:
                                    try: self.Label_Erro1.deleteLater()
                                    except: pass
                                    self.Label_Erro1 = QLabel()
                                    self.Label_Erro1.setStyleSheet('color: red')
                                    self.Label_Erro1.setFont(self.font12)
                                    self.Label_Erro1.setAlignment(Qt.AlignCenter)
                                    self.Label_Erro1.setText('Corretora já existe.') # Substitui a string do erro mais provável
                                    self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 11, 0, 1, 1, Qt.AlignCenter)
                            else:
                                MessageBox_Msg1 = QMessageBox.about(self,'Aviso','Existe outra operação em andamento.\nTente novamente mais tarde.')
                        else:
                            try: self.Label_Erro1.deleteLater()
                            except: pass
                            self.Label_Erro1 = QLabel()
                            self.Label_Erro1.setStyleSheet('color: red')
                            self.Label_Erro1.setFont(self.font12)
                            self.Label_Erro1.setAlignment(Qt.AlignCenter)
                            self.Label_Erro1.setText('Insira um nome válido com até 8 caracteres.') # Substitui a string do erro mais provável
                            self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 11, 0, 1, 1, Qt.AlignCenter)
                    elif self.PageID == '22':
                        if len(self.TextBox_NovoNomeCorretora.text())>0 and len(self.TextBox_NovoNomeCorretora.text())<9:
                            if not self.DBManager.FLAG and not self.AtualizarCotacoes_running:
                                Executado = self.DBManager.RenameCorretora()
                                if Executado:
                                    self.ClearPage()
                                    self.PageID = '18'
                                    self.PageIDAux = ''
                                    self.InsertGridLayout(1, 0, 1, 20)
                                    self.HMI_Trades.HMI_Trades_Bolsa.CreatePage18()
                                    self.PutHeader()
                                else:
                                     try: self.Label_Erro1.deleteLater()
                                     except: pass
                                     self.Label_Erro1 = QLabel()
                                     self.Label_Erro1.setStyleSheet('color: red')
                                     self.Label_Erro1.setFont(self.font12)
                                     self.Label_Erro1.setAlignment(Qt.AlignCenter)
                                     self.Label_Erro1.setText('Corretora já existe.') # Substitui a string do erro mais provável
                                     self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 11, 0, 1, 1, Qt.AlignCenter)
                            else:
                                MessageBox_Msg1 = QMessageBox.about(self,'Aviso','Existe outra operação em andamento.\nPode ser que a tabela de cotações\nesteja sendo atualizada\n\nTente novamente mais tarde.')
                        else:
                            try: self.Label_Erro1.deleteLater()
                            except: pass
                            self.Label_Erro1 = QLabel()
                            self.Label_Erro1.setStyleSheet('color: red')
                            self.Label_Erro1.setFont(self.font12)
                            self.Label_Erro1.setAlignment(Qt.AlignCenter)
                            self.Label_Erro1.setText('Insira um nome válido com até 8 caracteres.') # Substitui a string do erro mais provável
                            self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 11, 0, 1, 1, Qt.AlignCenter)
                    elif self.PageID == '25':
                        try:
                            if not self.DBManager.FLAG:
                                self.DBManager.UpdateContaCorrente()
                                self.ClearPage()
                                self.PageID = '18'
                                self.PageIDAux = ''
                                self.InsertGridLayout(1, 0, 1, 20)
                                self.HMI_Trades.HMI_Trades_Bolsa.CreatePage18()
                                self.PutHeader()
                            else:
                                MessageBox_Msg1 = QMessageBox.about(self,'Aviso','Existe outra operação em andamento.\nTente novamente mais tarde.')
                        except:
                            try: self.Label_Erro1.deleteLater()
                            except: pass
                            self.Label_Erro1 = QLabel()
                            self.Label_Erro1.setStyleSheet('color: red')
                            self.Label_Erro1.setFont(self.font12)
                            self.Label_Erro1.setAlignment(Qt.AlignCenter)
                            self.Label_Erro1.setText('Insira um valor válido.') # Substitui a string do erro mais provável
                            self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 4, 1, 1, 1, Qt.AlignCenter)
                else:
                    self.ClearPage()
                    self.PageID = '18'
                    self.PageIDAux = ''
                    self.InsertGridLayout(1, 0, 1, 20)
                    self.HMI_Trades.HMI_Trades_Bolsa.CreatePage18()
                    self.PutHeader()
            #%% PageID == '19'
            elif PageID in ['19','19_']:
                if PageID == '19_':
                    if self.PageID == '37':
                        if len(self.TextBox_NovaCorretora.text())>0 and len(self.TextBox_NovaCorretora.text())<12:
                            if not self.DBManager.FLAG:
                                Executado = self.DBManager.AddCorretora()
                                if Executado:
                                    self.ClearPage()
                                    self.PageID = '19'
                                    self.PageIDAux = ''
                                    self.HMI_Trades.HMI_Trades_Criptomoedas.CreatePage19()
                                    self.PutHeader()
                                else:
                                    try: self.Label_Erro1.deleteLater()
                                    except: pass
                                    self.Label_Erro1 = QLabel()
                                    self.Label_Erro1.setStyleSheet('color: red')
                                    self.Label_Erro1.setFont(self.font12)
                                    self.Label_Erro1.setAlignment(Qt.AlignCenter)
                                    self.Label_Erro1.setText('Corretora já existe.') # Substitui a string do erro mais provável
                                    self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 14, 0, 1, 1, Qt.AlignCenter)
                            else:
                                MessageBox_Msg1 = QMessageBox.about(self,'Aviso','Existe outra operação em andamento.\nTente novamente mais tarde.')
                        else:
                            try: self.Label_Erro1.deleteLater()
                            except: pass
                            self.Label_Erro1 = QLabel()
                            self.Label_Erro1.setStyleSheet('color: red')
                            self.Label_Erro1.setFont(self.font12)
                            self.Label_Erro1.setAlignment(Qt.AlignCenter)
                            self.Label_Erro1.setText('Insira um nome válido com até 11 caracteres.') # Substitui a string do erro mais provável
                            self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 14, 0, 1, 1, Qt.AlignCenter)
                    elif self.PageID == '39':
                        if len(self.TextBox_NovoNomeCorretora.text())>0 and len(self.TextBox_NovoNomeCorretora.text())<12:
                            if not self.DBManager.FLAG and not self.AtualizarCotacoes_running:
                                Executado = self.DBManager.RenameCorretora()
                                if Executado:
                                    self.ClearPage()
                                    self.PageID = '19'
                                    self.PageIDAux = ''
                                    self.HMI_Trades.HMI_Trades_Criptomoedas.CreatePage19()
                                    self.PutHeader()
                                else:
                                     try: self.Label_Erro1.deleteLater()
                                     except: pass
                                     self.Label_Erro1 = QLabel()
                                     self.Label_Erro1.setStyleSheet('color: red')
                                     self.Label_Erro1.setFont(self.font12)
                                     self.Label_Erro1.setAlignment(Qt.AlignCenter)
                                     self.Label_Erro1.setText('Corretora já existe.') # Substitui a string do erro mais provável
                                     self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 11, 0, 1, 1, Qt.AlignCenter)
                            else:
                                MessageBox_Msg1 = QMessageBox.about(self,'Aviso','Existe outra operação em andamento.\nPode ser que a tabela de cotações\nesteja sendo atualizada\n\nTente novamente mais tarde.')
                        else:
                            try: self.Label_Erro1.deleteLater()
                            except: pass
                            self.Label_Erro1 = QLabel()
                            self.Label_Erro1.setStyleSheet('color: red')
                            self.Label_Erro1.setFont(self.font12)
                            self.Label_Erro1.setAlignment(Qt.AlignCenter)
                            self.Label_Erro1.setText('Insira um nome válido com até 11 caracteres.') # Substitui a string do erro mais provável
                            self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 11, 0, 1, 1, Qt.AlignCenter)
                else:
                    self.ClearPage()
                    self.PageID = '19'
                    self.PageIDAux = ''
                    self.HMI_Trades.HMI_Trades_Criptomoedas.CreatePage19()
                    self.PutHeader()
            #%% PageID == 'Bancos'
            elif PageID in ['Bancos', 'Bancos_']:
                if PageID == 'Bancos_':
                    if self.PageID == 'BancosAdd':
                        if len(self.TextBox_NovoBanco.text())>0 and len(self.TextBox_NovoBanco.text())<12:
                            if not self.DBManager.FLAG:
                                Executado = self.DBManager.AddBanco()
                                if Executado:
                                    self.ClearPage()
                                    self.PageID = 'Bancos'
                                    self.PageIDAux = ''
                                    self.InsertGridLayout(1, 0, 1, 20)
                                    self.HMI_Trades.CreatePageBancos()
                                    self.PutHeader()
                                else:
                                    try: self.Label_Erro1.deleteLater()
                                    except: pass
                                    self.Label_Erro1 = QLabel()
                                    self.Label_Erro1.setStyleSheet('color: red')
                                    self.Label_Erro1.setFont(self.font12)
                                    self.Label_Erro1.setAlignment(Qt.AlignCenter)
                                    self.Label_Erro1.setText('Banco já existe.') # Substitui a string do erro mais provável
                                    self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 14, 0, 1, 1, Qt.AlignCenter)
                            else:
                                MessageBox_Msg1 = QMessageBox.about(self,'Aviso','Existe outra operação em andamento.\nTente novamente mais tarde.')
                        else:
                            try: self.Label_Erro1.deleteLater()
                            except: pass
                            self.Label_Erro1 = QLabel()
                            self.Label_Erro1.setStyleSheet('color: red')
                            self.Label_Erro1.setFont(self.font12)
                            self.Label_Erro1.setAlignment(Qt.AlignCenter)
                            self.Label_Erro1.setText('Insira um nome válido com até 11 caracteres.') # Substitui a string do erro mais provável
                            self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 14, 0, 1, 1, Qt.AlignCenter)
                    elif self.PageID == 'BancosRename':
                        if len(self.TextBox_NovoNomeBanco.text())>0 and len(self.TextBox_NovoNomeBanco.text())<12:
                            if not self.DBManager.FLAG:
                                Executado = self.DBManager.RenameBanco()
                                if Executado:
                                    self.ClearPage()
                                    self.PageID = 'Bancos'
                                    self.PageIDAux = ''
                                    self.InsertGridLayout(1, 0, 1, 20)
                                    self.HMI_Trades.CreatePageBancos()
                                    self.PutHeader()
                                else:
                                     try: self.Label_Erro1.deleteLater()
                                     except: pass
                                     self.Label_Erro1 = QLabel()
                                     self.Label_Erro1.setStyleSheet('color: red')
                                     self.Label_Erro1.setFont(self.font12)
                                     self.Label_Erro1.setAlignment(Qt.AlignCenter)
                                     self.Label_Erro1.setText('Banco já existe.') # Substitui a string do erro mais provável
                                     self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 11, 0, 1, 1, Qt.AlignCenter)
                            else:
                                MessageBox_Msg1 = QMessageBox.about(self,'Aviso','Existe outra operação em andamento.\nTente novamente mais tarde.')
                        else:
                            try: self.Label_Erro1.deleteLater()
                            except: pass
                            self.Label_Erro1 = QLabel()
                            self.Label_Erro1.setStyleSheet('color: red')
                            self.Label_Erro1.setFont(self.font12)
                            self.Label_Erro1.setAlignment(Qt.AlignCenter)
                            self.Label_Erro1.setText('Insira um nome válido com até 11 caracteres.') # Substitui a string do erro mais provável
                            self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 11, 0, 1, 1, Qt.AlignCenter)
                else:
                    self.ClearPage()
                    self.PageID = 'Bancos'
                    self.PageIDAux = ''
                    self.InsertGridLayout(1, 0, 1, 20)
                    self.HMI_Trades.CreatePageBancos()
                    self.PutHeader()
            #%% PageID == 'BancosAdd'
            elif PageID == 'BancosAdd':
                self.ClearPage()
                self.PageID = 'BancosAdd'
                self.PageIDAux = ''
                self.InsertGridLayout(1, 0, 1, 20)
                self.HMI_Trades.CreatePageBancosAdd()
                self.PutHeader()
            #%% PageID == 'BancosDelete'
            elif PageID == 'BancosDelete': self.HMI_Trades.CreatePageBancosDelete()
            #%% PageID == 'BancosRename'
            elif PageID == 'BancosRename':
                self.ClearPage()
                self.PageID = 'BancosRename'
                self.PageIDAux = ''
                self.InsertGridLayout(1, 0, 1, 20)
                self.HMI_Trades.CreatePageBancosRename()
                self.PutHeader()
            #%% PageID == '20'
            elif PageID == '20':
                self.ClearPage()
                self.PageID = '20'
                self.PageIDAux = ''
                self.InsertGridLayout(1, 0, 1, 20)
                self.HMI_Trades.HMI_Trades_Bolsa.CreatePage20()
                self.PutHeader()
            #%% PageID == '21'
            elif PageID == '21': # MessageBox
                if not self.DBManager.FLAG and not self.AtualizarCotacoes_running:
                    self.HMI_Trades.HMI_Trades_Bolsa.CreatePage21()
                else:
                    MessageBox_Msg1 = QMessageBox.about(self,'Aviso','Existe outra operação em andamento.\nPode ser que a tabela de cotações\nesteja sendo atualizada\n\nTente novamente mais tarde.')
            #%% PageID == '22'
            elif PageID == '22':
                self.ClearPage()
                self.PageID = '22'
                self.PageIDAux = ''
                self.InsertGridLayout(1, 0, 1, 20)
                self.HMI_Trades.HMI_Trades_Bolsa.CreatePage22()
                self.PutHeader()
            #%% PageID == '23'
            elif PageID in ['23', '23_', '_23_']:
                if self.PageID == '34b':
                    if not self.DBManager.FLAG and not self.AtualizarCotacoes_running:
                        if PageID == '_23_':
                            self.Modo = 'Manual'
                            self.DBManager.RenameAtivo()
                        elif PageID == '23_':
                            self.Modo = 'Auto'
                            self.DBManager.RenameAtivo()
                    else:
                        MessageBox_Msg1 = QMessageBox.about(self,'Aviso','Existe outra operação em andamento.\nPode ser que a tabela de cotações\nesteja sendo atualizada\n\nTente novamente mais tarde.')
                elif self.PageID == '35b':
                    if not self.DBManager.FLAG and not self.AtualizarCotacoes_running:
                        if PageID == '23_':
                            self.Modo = 'Manual'
                            self.DBManager.RenameAtivo()
                    else:
                        MessageBox_Msg1 = QMessageBox.about(self,'Aviso','Existe outra operação em andamento.\nPode ser que a tabela de cotações\nesteja sendo atualizada\n\nTente novamente mais tarde.')
                self.ClearPage()
                self.PageID = '23'
                self.PageIDAux = ''
                self.InsertGridLayout(1, 0, 1, 20)
                self.HMI_Trades.HMI_Trades_Bolsa.CreatePage23()
                self.PutHeader()

                if self.DBManager.VerificarSeTemEstoqueNegativo(): MessageBox_Msg1 = QMessageBox.about(self,'Aviso','ESTOQUE NEGATIVO! VERIFIQUE AS SUAS OPERAÇÕES!')
                elif self.DBManager.VerificarSeTemCustoNegativo(): MessageBox_Msg1 = QMessageBox.about(self,'Aviso','CUSTO NEGATIVO! VERIFIQUE AS SUAS OPERAÇÕES!')
            #%% PageID == '24'
            elif PageID in ['24', '24_']:
                if PageID  == '24_':
                    try:
                        Data = str(self.Selecao_Ano+'/'+self.Selecao_Mes+'/'+self.Selecao_Dia+' '+self.Selecao_Hora+':'+self.Selecao_Minuto+':'+self.Selecao_Segundo)
                        Data = datetime.strptime(Data, '%Y/%m/%d %H:%M:%S')
                        sinal = 1 if self.Deposito else -1
                        Valor = float(self.TextBox_Valor.text())*sinal
                        if not self.DBManager.FLAG:
                            if self.DBManager.AddDepositoOuSaque(Data, Valor):
                                self.TextBox_Valor.setText('')
                                self.OnButtonPressed('Deposito')
                                self.HMI_Trades.HMI_Trades_Bolsa.CreateTable_24_2()
                                self.TextBox_Valor.setFocus()
                            else:
                                MessageBox_Msg1 = QMessageBox.about(self,'Informação','Movimentação não registrada.\nJá tem registro nessa data.')
                                self.TextBox_Valor.setFocus()
                        else:
                            MessageBox_Msg1 = QMessageBox.about(self,'Aviso','Existe outra operação em andamento.\nTente novamente mais tarde.')
                    except Exception as e:
                        print("Erro: Bolsa; Depósitos e Saques:\n",e)
                        self.TextBox_Valor.setStyleSheet('color: red')
                        self.TextBox_Valor.setFocus()
                else:
                    self.ClearPage()
                    self.PageID = '24'
                    self.PageIDAux = ''
                    self.InsertGridLayout(1, 0, 1, 20)
                    self.HMI_Trades.HMI_Trades_Bolsa.CreatePage24()
                    self.PutHeader()
            #%% PageID == '25'
            elif PageID == '25':
                self.ClearPage()
                self.PageID = '25'
                self.PageIDAux = ''
                self.InsertGridLayout(1, 0, 1, 20)
                self.HMI_Trades.HMI_Trades_Bolsa.CreatePage25()
                self.PutHeader()
            #%% PageID == '26'
            elif PageID == '26':
                self.ClearPage()
                self.PageID = '26'
                self.PageIDAux = ''
                self.InsertGridLayout(1, 0, 1, 20)
                self.HMI_Trades.HMI_Trades_Bolsa.CreatePage26()
                self.PutHeader()
            #%% PageID == '27'
            elif PageID == '27':
                self.ClearPage()
                self.PageID = '27'
                self.PageIDAux = ''
                self.InsertGridLayout(1, 0, 1, 20)
                self.HMI_Trades.HMI_Trades_Bolsa.CreatePage27()
                self.PutHeader()
            #%% PageID == '28'
            elif PageID == '28':
                try: self.Label_Erro1.deleteLater()
                except: pass
                if self.PageID == '28':
                    # Registrar na DBManager a operação
                    try:
                        # Verificar se os dados inseridos são coerentes e só então registrar na DB
                        data = self.Selecao_Dia+"/"+self.Selecao_Mes+"/"+self.Selecao_Ano+" "+self.Selecao_Hora+":"+self.Selecao_Minuto+":"+self.Selecao_Segundo
                        data = datetime.strptime(data, '%d/%m/%Y %H:%M:%S')
                        if data<=datetime.now():
                            if "%" in self.TextBox_TaxaB3.text():
                                TaxaB3Per = float(self.TextBox_TaxaB3.text().replace("%",""))
                                TaxaB3 = round(TaxaB3Per/100*int(self.TextBox_Qqt.text())*float(self.TextBox_Preco.text()),2)
                            else:
                                TaxaB3 = float(self.TextBox_TaxaB3.text())
                                TaxaB3Per = round(TaxaB3/int(self.TextBox_Qqt.text())/float(self.TextBox_Preco.text())*100,2)
                            check = int(self.TextBox_Qqt.text())
                            check = float(self.TextBox_Preco.text())
                            check = float(self.TextBox_Corretagem.text())
                            if not self.DBManager.FLAG and not self.AtualizarCotacoes_running:
                                deubom = self.DBManager.AddOperacao()
                                if deubom:
                                    self.Msg = OperacaoRegistradaComSucesso.OperacaoRegistradaComSucesso(self, 4)
                                    # MessageBox_Msg1 = QMessageBox.about(self,'Informação','Operação registrada com sucesso.')
                                    self.LimparPage()
                                else:
                                    try: self.Label_Erro1.deleteLater()
                                    except: pass
                                    self.Label_Erro1 = QLabel()
                                    self.Label_Erro1.setStyleSheet('color: red')
                                    self.Label_Erro1.setFont(self.font12)
                                    self.Label_Erro1.setAlignment(Qt.AlignCenter)
                                    self.Label_Erro1.setText('Data inválida.\nTalvez já tenha uma operação nesse momento.') # Substitui a string do erro mais provável
                                    self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 6, 4, 1, 3, Qt.AlignRight)
                            else:
                                MessageBox_Msg1 = QMessageBox.about(self,'Aviso','Existe outra operação em andamento.\nPode ser que a tabela de cotações\nesteja sendo atualizada\n\nTente novamente mais tarde.')
                        else:
                            try: self.Label_Erro1.deleteLater()
                            except: pass
                            self.Label_Erro1 = QLabel()
                            self.Label_Erro1.setStyleSheet('color: red')
                            self.Label_Erro1.setFont(self.font12)
                            self.Label_Erro1.setAlignment(Qt.AlignCenter)
                            self.Label_Erro1.setText('Data inválida.') # Substitui a string do erro mais provável
                            self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 6, 4, 1, 3, Qt.AlignRight)
                    except:
                        try: self.Label_Erro1.deleteLater()
                        except: pass
                        self.Label_Erro1 = QLabel()
                        self.Label_Erro1.setStyleSheet('color: red')
                        self.Label_Erro1.setFont(self.font12)
                        self.Label_Erro1.setAlignment(Qt.AlignCenter)
                        self.Label_Erro1.setText('Confira os valores inseridos.') # Substitui a string do erro mais provável
                        self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 6, 4, 1, 3, Qt.AlignRight)

                elif self.PageID in ['36a','36b']:
                    try: self.Label_Erro1.deleteLater()
                    except: pass
                    # Registrar na DBManager o ativo TextBox_NomeNovoAtivo
                    try:
                        # Verificar se os dados inseridos são coerentes e só então registrar na DB
                        data = self.Selecao_Dia+"/"+self.Selecao_Mes+"/"+self.Selecao_Ano+" "+self.Selecao_Hora+":"+self.Selecao_Minuto+":"+self.Selecao_Segundo
                        data = datetime.strptime(data, '%d/%m/%Y %H:%M:%S')
                        if data<=datetime.now():
                            if "%" in self.TextBox_TaxaB3.text():
                                TaxaB3Per = float(self.TextBox_TaxaB3.text().replace("%",""))
                                TaxaB3 = round(TaxaB3Per/100*int(self.TextBox_Qqt.text())*float(self.TextBox_Preco.text()),2)
                            else:
                                TaxaB3 = float(self.TextBox_TaxaB3.text())
                                TaxaB3Per = round(TaxaB3/int(self.TextBox_Qqt.text())/float(self.TextBox_Preco.text())*100,2)
                            check = int(self.TextBox_Qqt.text())
                            check = float(self.TextBox_Preco.text())
                            check = float(self.TextBox_Corretagem.text())

                            if not self.DBManager.FLAG and not self.AtualizarCotacoes_running:
                                self.DBManager.AddAtivo()
                                self.DBManager.AddOperacao() # Sempre vai dar bom na criação do ativo
                                self.AcabouDeRegistrarNovoAtivo = True
                                self.ComboBox_Ativo.clear()
                                ativos = self.DBManager.GetAtivos()
                                for idx, ativo in enumerate(ativos):
                                    self.ComboBox_Ativo.addItem(ativo)
                                    if ativo == self.TextBox_NomeNovoAtivo.text():
                                        self.idxAtivo = idx
                                self.ComboBox_Ativo.setCurrentIndex(self.idxAtivo)

                                self.ClearPage()
                                self.PageID = '28'
                                self.PageIDAux = ''
                                self.InsertGridLayout(1, 0, 1, 20)
                                self.HMI_Trades.HMI_Trades_Bolsa.CreatePage28()
                                self.PutHeader()
                                self.Msg = OperacaoRegistradaComSucesso.OperacaoRegistradaComSucesso(self, 7)
                            else:
                                MessageBox_Msg1 = QMessageBox.about(self,'Aviso','Existe outra operação em andamento.\nPode ser que a tabela de cotações\nesteja sendo atualizada\n\nTente novamente mais tarde.')
                        else:
                            try: self.Label_Erro1.deleteLater()
                            except: pass
                            self.Label_Erro1 = QLabel()
                            self.Label_Erro1.setStyleSheet('color: red')
                            self.Label_Erro1.setFont(self.font12)
                            self.Label_Erro1.setAlignment(Qt.AlignCenter)
                            self.Label_Erro1.setText('Data inválida.') # Substitui a string do erro mais provável
                            self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 6, 5, 1, 1, Qt.AlignCenter)
                    except:
                        try: self.Label_Erro1.deleteLater()
                        except: pass
                        self.Label_Erro1 = QLabel()
                        self.Label_Erro1.setStyleSheet('color: red')
                        self.Label_Erro1.setFont(self.font12)
                        self.Label_Erro1.setAlignment(Qt.AlignCenter)
                        self.Label_Erro1.setText('Confira os valores inseridos.') # Substitui a string do erro mais provável
                        self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 6, 5, 1, 1, Qt.AlignCenter)

                elif len(self.ComboBox_Ativo)>0:
                    self.ClearPage()
                    self.PageID = '28'
                    self.PageIDAux = ''
                    self.InsertGridLayout(1, 0, 1, 20)
                    self.HMI_Trades.HMI_Trades_Bolsa.CreatePage28()
                    self.PutHeader()
            #%% PageID == '29'
            elif PageID == '29':
                self.ClearPage()
                self.PageID = '29'
                self.PageIDAux = ''
                self.InsertGridLayout(1, 0, 1, 20)
                self.HMI_Trades.HMI_Trades_Bolsa.CreatePage29()
                self.PutHeader()
            #%% PageID == '30'
            elif PageID == '30':
                self.ClearPage()
                self.PageID = '30'
                self.PageIDAux = ''
                self.InsertGridLayout(1, 0, 1, 20)
                self.HMI_Trades.HMI_Trades_Bolsa.CreatePage30()
                self.PutHeader()
            #%% PageID == '31a'
            elif PageID == '31a': MessageBox_Msg1 = QMessageBox.about(self,'Informação','Funcionalidade ainda não implementada.')
            #%% PageID == '31b'
            elif PageID == '31b': MessageBox_Msg1 = QMessageBox.about(self,'Informação','Funcionalidade ainda não implementada.')
            #%% PageID == '32a'
            elif PageID == '32a': MessageBox_Msg1 = QMessageBox.about(self,'Informação','Funcionalidade ainda não implementada.')
            #%% PageID == '32b'
            elif PageID == '32b': MessageBox_Msg1 = QMessageBox.about(self,'Informação','Funcionalidade ainda não implementada.')
            #%% PageID == '33'
            elif PageID == '33':
                try:
                    self.LW = LoadingWindow.LoadingWindow(self)
                    self.LW.show()
                    self.Thread_BuscarSetoresExistentes = WorkerThread('Thread_BuscarSetoresExistentes', self)
                    self.Thread_BuscarSetoresExistentes.start()
                except Exception as e: print(e)
            #%% PageID == '34a'
            elif PageID == '34a':
                self.ClearPage()
                self.PageID = '34a'
                self.PageIDAux = ''
                self.InsertGridLayout(1, 0, 1, 20)
                self.HMI_Trades.HMI_Trades_Bolsa.CreatePage34a()
                self.PutHeader()
            #%% PageID == '34b'
            elif PageID == '34b':
                self.ClearPage()
                self.PageID = '34b'
                self.PageIDAux = ''
                self.InsertGridLayout(1, 0, 1, 20)
                self.HMI_Trades.HMI_Trades_Bolsa.CreatePage34b()
                self.PutHeader()
            #%% PageID == '35a'
            elif PageID == '35a':
                self.ClearPage()
                self.PageID = '35a'
                self.PageIDAux = ''
                self.InsertGridLayout(1, 0, 1, 20)
                self.HMI_Trades.HMI_Trades_Bolsa.CreatePage35a()
                self.PutHeader()
            #%% PageID == '35b'
            elif PageID == '35b':
                self.ClearPage()
                self.PageID = '35b'
                self.PageIDAux = ''
                self.InsertGridLayout(1, 0, 1, 20)
                self.HMI_Trades.HMI_Trades_Bolsa.CreatePage35b()
                self.PutHeader()
            #%% PageID == '36a_Auto'
            elif PageID == '36a_Auto':
                self.ClearPage()
                self.PageID = '36a'
                self.PageIDAux = ''
                self.InsertGridLayout(1, 0, 1, 20)
                self.ModoManualAtivo = False
                self.HMI_Trades.HMI_Trades_Bolsa.CreatePage36a()
                self.PutHeader()
            #%% PageID == '36a_Manual'
            elif PageID == '36a_Manual':
                self.ClearPage()
                self.PageID = '36a'
                self.PageIDAux = ''
                self.InsertGridLayout(1, 0, 1, 20)
                self.ModoManualAtivo = True
                self.HMI_Trades.HMI_Trades_Bolsa.CreatePage36a()
                self.PutHeader()
            #%% PageID == '36b_Manual'
            elif PageID == '36b_Manual':
                self.ClearPage()
                self.PageID = '36b'
                self.PageIDAux = ''
                self.InsertGridLayout(1, 0, 1, 20)
                self.ModoManualAtivo = True
                self.HMI_Trades.HMI_Trades_Bolsa.CreatePage36a()
                self.PutHeader()
            #%% PageID == '37'
            elif PageID == '37':
                self.ClearPage()
                self.PageID = '37'
                self.PageIDAux = ''
                self.InsertGridLayout(1, 0, 1, 20)
                self.HMI_Trades.HMI_Trades_Criptomoedas.CreatePage37()
                self.PutHeader()
            #%% PageID == '38'
            elif PageID == '38': # MessageBox
                self.HMI_Trades.HMI_Trades_Criptomoedas.CreatePage38()
                self.PutHeader()
            #%% PageID == '39'
            elif PageID == '39':
                self.ClearPage()
                self.PageID = '39'
                self.PageIDAux = ''
                self.InsertGridLayout(1, 0, 1, 20)
                self.HMI_Trades.HMI_Trades_Criptomoedas.CreatePage39()
                self.PutHeader()
            #%% PageID == '40'
            elif PageID == '40':
                self.ClearPage()
                self.PageID = '40'
                self.PageIDAux = ''
                self.InsertGridLayout(1, 0, 1, 20)
                self.HMI_Trades.HMI_Trades_Criptomoedas.CreatePage40()
                self.PutHeader()

                if self.DBManager.VerificarPrimeiraOperacao(): MessageBox_Msg1 = QMessageBox.about(self,'Aviso','A primeira operação deve ser do tipo DEPÓSITO ou DROP.')
                elif self.DBManager.VerificarSeTemEstoqueNegativo(): MessageBox_Msg1 = QMessageBox.about(self,'Aviso','ESTOQUE NEGATIVO! VERIFIQUE AS SUAS OPERAÇÕES!')
                elif self.DBManager.VerificarSeTemCustoNegativo(): MessageBox_Msg1 = QMessageBox.about(self,'Aviso','CUSTO NEGATIVO! VERIFIQUE AS SUAS OPERAÇÕES!')
            #%% PageID == '41'
            elif PageID == '41':
                try: self.Label_Erro1.deleteLater()
                except: pass
                if self.PageID == '41':
                    # Registrar na DBManager a operação
                    try:
                        # Verificar se os dados inseridos são coerentes e só então registrar na DB
                        data = self.Selecao_Dia+"/"+self.Selecao_Mes+"/"+self.Selecao_Ano+" "+self.Selecao_Hora+":"+self.Selecao_Minuto+":"+self.Selecao_Segundo
                        data = datetime.strptime(data, '%d/%m/%Y %H:%M:%S')
                        if len(self.TextBox_Taxa.text())>0: check = float(self.TextBox_Taxa.text())
                        check = float(self.TextBox_Qqt.text())
                        check = float(self.TextBox_Preco.text())
                        if len(self.TextBox_Conversao.text())>0: check = float(self.TextBox_Conversao.text())
                        if not self.TextBox_Par_Esquerda.text() == self.TextBox_Par_Direita.text():
                            if not self.TextBox_Par_Esquerda.text() == '' or not self.TextBox_Par_Direita.text() == '':
                                if not self.DBManager.FLAG and not self.AtualizarCotacoes_running:
                                    deubom = self.DBManager.AddOperacao()
                                    if deubom:
                                        try:
                                            self.Msg = OperacaoRegistradaComSucesso.OperacaoRegistradaComSucesso(self, 5)
                                        except Exception as e:
                                            print(e)
                                        self.LimparPage()
                                    else:
                                        try: self.Label_Erro1.deleteLater()
                                        except: pass
                                        self.Label_Erro1 = QLabel()
                                        self.Label_Erro1.setStyleSheet('color: red')
                                        self.Label_Erro1.setFont(self.font12)
                                        self.Label_Erro1.setAlignment(Qt.AlignCenter)
                                        self.Label_Erro1.setText('Data inválida.\nTalvez já tenha uma operação nesse momento.') # Substitui a string do erro mais provável
                                        self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 6, 4, 1, 3, Qt.AlignCenter)
                                else:
                                    MessageBox_Msg1 = QMessageBox.about(self,'Aviso','Existe outra operação em andamento.\nPode ser que a tabela de cotações\nesteja sendo atualizada\n\nTente novamente mais tarde.')
                            else:
                                try: self.Label_Erro1.deleteLater()
                                except: pass
                                self.Label_Erro1 = QLabel()
                                self.Label_Erro1.setStyleSheet('color: red')
                                self.Label_Erro1.setFont(self.font12)
                                self.Label_Erro1.setAlignment(Qt.AlignCenter)
                                self.Label_Erro1.setText('As moedas devem ser declaradas.') # Substitui a string do erro mais provável
                                self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 6, 4, 1, 3, Qt.AlignCenter)
                        else:
                            try: self.Label_Erro1.deleteLater()
                            except: pass
                            self.Label_Erro1 = QLabel()
                            self.Label_Erro1.setStyleSheet('color: red')
                            self.Label_Erro1.setFont(self.font12)
                            self.Label_Erro1.setAlignment(Qt.AlignCenter)
                            self.Label_Erro1.setText('As moedas nos pares não podem ser iguais.') # Substitui a string do erro mais provável
                            self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 6, 4, 1, 3, Qt.AlignCenter)
                    except:
                        try: self.Label_Erro1.deleteLater()
                        except: pass
                        self.Label_Erro1 = QLabel()
                        self.Label_Erro1.setStyleSheet('color: red')
                        self.Label_Erro1.setFont(self.font12)
                        self.Label_Erro1.setAlignment(Qt.AlignCenter)
                        self.Label_Erro1.setText('Confira os valores inseridos.') # Substitui a string do erro mais provável
                        self.F_GLayout[self.GCount-1].addWidget(self.Label_Erro1, 6, 4, 1, 3, Qt.AlignCenter)
                else:
                    self.ClearPage()
                    self.PageID = '41'
                    self.PageIDAux = ''
                    self.InsertGridLayout(1, 0, 1, 20)
                    self.HMI_Trades.HMI_Trades_Criptomoedas.CreatePage41()
                    self.PutHeader()
            #%% PageID == '42'
            elif PageID == '42': # Ideia abandonada
                # self.ClearPage()
                # self.PageID = '42'
                # self.PageIDAux = ''
                # self.InsertGridLayout(1, 0, 1, 20)
                # self.HMI_Trades.HMI_Trades_Criptomoedas.CreatePage42()
                # self.PutHeader()
                pass
            #%% PageID == '43'
            elif PageID == '43': # Alter Cripto Operation
                self.ClearPage()
                self.PageID = '43'
                self.PageIDAux = ''
                self.InsertGridLayout(1, 0, 1, 20)
                self.HMI_Trades.HMI_Trades_Criptomoedas.CreatePage43()
                self.PutHeader()
            #%% PageID == '44'
            elif PageID == '44': # Delete Cripto Operation
                self.ClearPage()
                self.PageID = '44'
                self.PageIDAux = ''
                self.InsertGridLayout(1, 0, 1, 20)
                self.HMI_Trades.HMI_Trades_Criptomoedas.CreatePage44()
                self.PutHeader()
            #%% PageID == '45'
            elif PageID == '45':
                if not self.DBManager.FLAG and not self.AtualizarCotacoes_running:
                    self.ClearPage()
                    self.PageIDAux = '45' # Tem ID especial, porque é compartilhado com outros ramos
                    self.InsertGridLayout(1, 0, 1, 20)
                    self.CreatePage45()
                    self.PutHeader()
                else:
                    MessageBox_Msg1 = QMessageBox.about(self,'Aviso','Existe outra operação em andamento.\nPode ser que a tabela de cotações\nesteja sendo atualizada\n\nTente novamente mais tarde.')
            #%% PageID == '46'
            elif PageID == '46':
                self.ClearPage()
                self.PageID = '46'
                self.PageIDAux = ''
                self.InsertGridLayout(1, 0, 1, 20)
                self.HMI_ARCA.CreatePage46()
                self.PutHeader()
            #%% PageID == '47'
            elif PageID == '47':
                self.ClearPage()
                self.PageID = '47'
                self.PageIDAux = ''
                self.InsertGridLayout(1, 0, 1, 20)
                self.HMI_ARCA.CreatePage47()
                self.PutHeader()
            #%% PageID == '48'
            elif PageID == '48':
                self.ClearPage()
                self.PageID = '48'
                self.PageIDAux = ''
                self.InsertGridLayout(1, 0, 1, 20)
                self.HMI_ARCA.CreatePage48()
                self.PutHeader()
            #%% PageID == '49'
            elif PageID == '49':
                self.ClearPage()
                self.PageID = '49'
                self.PageIDAux = ''
                self.InsertGridLayout(1, 0, 1, 20)
                self.HMI_ARCA.CreatePage49()
                self.PutHeader()
            #%% PageID == '50'
            elif PageID == '50':
                self.ClearPage()
                self.PageID = '50'
                self.PageIDAux = ''
                self.InsertGridLayout(1, 0, 1, 20)
                self.HMI_ARCA.CreatePage50()
                self.PutHeader()
            #%% PageID == '51'
            elif PageID == '51':
                self.ClearPage()
                self.PageID = '51'
                self.PageIDAux = ''
                self.InsertGridLayout(1, 0, 1, 20)
                self.HMI_Tributacao.CreatePage51()
                self.PutHeader()
            #%% PageID == 'ChooseCurrency'
            elif PageID == 'ChooseCurrency':
                self.ClearPage()
                self.PageIDAux = 'ChooseCurrency' # Derivação do perfil, Tem ID especial
                self.InsertGridLayout(1, 0, 1, 20)
                self.CreatePageChooseCurrency()
                self.PutHeader()
            #%% Confirmar email
            elif PageID =='Confirmar email':
                pass
            #%% Inserir nova senha
            elif PageID == 'Inserir nova senha':
                pass
            #%% PageID == 'Info'
            elif PageID == 'Info1': self.CreatePageInfo1()
            elif PageID == 'Info2': self.CreatePageInfo2()
            elif PageID == 'Info3': self.CreatePageInfo3()
            elif PageID == 'Info4': self.CreatePageInfo4()
            elif PageID == 'Info5': self.CreatePageInfo5()
            elif PageID == 'Info6': self.CreatePageInfo6()
            elif PageID == 'Info7': self.CreatePageInfo7()
            elif PageID == 'Info8': self.CreatePageInfo8()
            elif PageID == 'Info9': self.CreatePageInfo9()
            elif PageID == 'Info10': self.CreatePageInfo10()
            elif PageID == 'Info11': self.CreatePageInfo11()
            elif PageID == 'Info12': self.CreatePageInfo12()
            elif PageID == 'Info13': self.CreatePageInfo13()
            elif PageID == 'Info14': self.CreatePageInfo14()
            elif PageID == 'Info15': self.CreatePageInfo15()
            elif PageID == 'Info16': self.CreatePageInfo16()
            elif PageID == 'Info17': self.CreatePageInfo17()
            elif PageID == 'Info18': self.CreatePageInfo18()
            elif PageID == 'Info19': self.CreatePageInfo19()
            elif PageID == 'Info20': self.CreatePageInfo20()
            elif PageID == 'Info21': self.CreatePageInfo21()
            elif PageID == 'Info22': self.CreatePageInfo22()
            elif PageID == 'Info23': self.CreatePageInfo23()
            elif PageID == 'Info24': self.CreatePageInfo24()
            elif PageID == 'Info25': self.CreatePageInfo25()
            elif PageID == 'Info26': self.CreatePageInfo26()
            elif PageID == 'Info27': self.CreatePageInfo27()
            elif PageID == 'Info28': self.CreatePageInfo28()

    #%% Limpar Page
    def LimparPage(self):
        if self.PageID in ['28', '29', '36a']:
            self.TextBox_Qqt.setText('')
            self.TextBox_Preco.setText('')
            self.TextBox_Corretagem.setText('')
            if "%" in self.TextBox_TaxaB3.text(): self.DBManager.UpdateTaxa(float(self.TextBox_TaxaB3.text().replace('%','')))
            else: self.TextBox_TaxaB3.setText(self.DBManager.GetTaxaB3Per())
            self.TextBox_Obs.setText('')
            self.TextBox_Qqt.setFocus()
            try: self.Label_Erro1.deleteLater()
            except: pass
        elif self.PageID in ['41']:
            self.TextBox_Par_Esquerda.setText('')
            self.TextBox_Par_Direita.setText('')
            self.TextBox_Qqt.setText('')
            if self.TipoDeOperacao in ['Compra', 'Venda']: self.TextBox_Preco.setText('')
            else: self.TextBox_Preco.setText('1')
            self.TextBox_Taxa.setText('')
            self.TextBox_MoedaDaTaxa.setText('')
            self.TextBox_Conversao.setText('')
            self.TextBox_Obs.setText('')
            self.TextBox_Par_Esquerda.setFocus()
            try: self.Label_Erro1.deleteLater()
            except: pass

    #%% CreatePage
    def CreatePage0(self):
        black = 'rgb(0, 0, 0)'
        self.Background_1 = QLabel()
        self.Background_1.setStyleSheet("background-color: "+black)

        self.Background_Entrar = QLabel()
        Background_Entrar = 'rgb(20, 20, 30)'
        self.Background_Entrar.setStyleSheet("background-color: "+Background_Entrar)

        self.Label_Titulo = QLabel()
        self.Label_Titulo.setText('Entrar')
        self.Label_Titulo.setStyleSheet('color: white')
        self.Label_Titulo.setFont(self.font30)

        self.Label_NomeDeUsuario = QLabel()
        self.Label_NomeDeUsuario.setText('  Nome de usuário:')
        self.Label_NomeDeUsuario.setStyleSheet('color: white')
        self.Label_NomeDeUsuario.setFont(self.font16)

        self.TextBox_NomeDeUsuario = QLineEdit()
        self.TextBox_NomeDeUsuario.returnPressed.connect(lambda: self.CreatePage('1'))
        self.TextBox_NomeDeUsuario.setFixedWidth(int(self.frameGeometry().width()/3*0.9))

        self.Label_Senha = QLabel()
        self.Label_Senha.setText('  Senha:')
        self.Label_Senha.setStyleSheet('color: white')
        self.Label_Senha.setFont(self.font16)

        self.TextBox_Senha = QLineEdit()
        self.TextBox_Senha.setEchoMode(QLineEdit.Password)
        self.TextBox_Senha.returnPressed.connect(lambda: self.CreatePage('1'))
        self.TextBox_Senha.setFixedWidth(int(self.frameGeometry().width()/3*0.9))

        self.Button_LogIn = QPushButton()
        self.Button_LogIn.pressed.connect(lambda: self.CreatePage('1'))
        self.Button_LogIn.setFont(self.font16)
        self.Button_LogIn.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
        self.Button_LogIn.setFixedWidth(int(self.frameGeometry().width()/3*0.9))
        self.Button_LogIn.setIcon(QIcon("./images/log_in.png"))
        self.Button_LogIn.setCursor(QCursor(Qt.PointingHandCursor))
        self.Button_LogIn.setStyleSheet("background-color: rgb(20, 120, 30); border: none; outline: none;")

        self.Button_EsqueciASenha = QPushButton("Esqueci a senha")
        self.Button_EsqueciASenha.pressed.connect(lambda: self.CreatePage('2'))
        self.Button_EsqueciASenha.setFont(self.font12)
        self.Button_EsqueciASenha.setCursor(QCursor(Qt.PointingHandCursor))

        self.Button_ImportarUsuario = QPushButton("Importar usuário")
        # self.Button_ImportarUsuario.pressed.connect(lambda: self.CreatePage('3'))
        self.Button_ImportarUsuario.pressed.connect(lambda: self.OnButtonPressed('ImportFromDB'))
        self.Button_ImportarUsuario.setFont(self.font12)
        self.Button_ImportarUsuario.setCursor(QCursor(Qt.PointingHandCursor))

        self.Button_CriarUsuario = QPushButton("Criar novo usuário")
        self.Button_CriarUsuario.pressed.connect(lambda: self.CreatePage('4'))
        self.Button_CriarUsuario.setFont(self.font12)
        self.Button_CriarUsuario.setCursor(QCursor(Qt.PointingHandCursor))

        self.Button_VersionInfo = QPushButton()
        self.Button_VersionInfo.pressed.connect(lambda: QMessageBox.about(self,'Informação', self.APPVersion))
        self.Button_VersionInfo.setIcon(QIcon("./images/Info.png"))
        self.Button_VersionInfo.setCursor(QCursor(Qt.PointingHandCursor))
        self.Button_VersionInfo.setStyleSheet("background-color: none; border: none; outline: none;")

        self.F_GLayout[self.GCount-1].addWidget(self.Background_Entrar, 0, 0, 20, 2)
        self.F_GLayout[self.GCount-1].addWidget(self.Label_Titulo, 5, 0, 1, 2, Qt.AlignCenter)
        self.F_GLayout[self.GCount-1].addWidget(self.Label_NomeDeUsuario, 7, 0, Qt.AlignLeft | Qt.AlignBottom)
        self.F_GLayout[self.GCount-1].addWidget(self.TextBox_NomeDeUsuario, 8, 0, 1, 2, Qt.AlignHCenter | Qt.AlignTop)
        self.F_GLayout[self.GCount-1].addWidget(self.Label_Senha, 9, 0, Qt.AlignLeft | Qt.AlignBottom)
        self.F_GLayout[self.GCount-1].addWidget(self.TextBox_Senha, 10, 0, 1, 2, Qt.AlignHCenter | Qt.AlignTop)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_LogIn, 11, 0, 2, 2, Qt.AlignCenter)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_EsqueciASenha, 17, 0, Qt.AlignLeft)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_ImportarUsuario, 18, 0, Qt.AlignLeft)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_CriarUsuario, 19, 0, Qt.AlignLeft)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_VersionInfo, 19, 1, Qt.AlignRight)

        self.Gif_Mercado = QLabel()
        self.Movie_Mercado = QMovie("./images/AnaliseDoMercadoGifIcone.gif")
        self.Movie_Mercado.setScaledSize(QSize().scaled(int(self.frameGeometry().width()*3/4*.90),int(self.frameGeometry().height()*3/4*.95), Qt.KeepAspectRatio))
        self.Gif_Mercado.setMovie(self.Movie_Mercado)
        self.Movie_Mercado.start()

        self.Button_MercadoInvisivel = QPushButton()
        self.Button_MercadoInvisivel.pressed.connect(lambda: self.CreatePage('5'))
        self.Button_MercadoInvisivel.setStyleSheet('background-color: rgba(0, 0, 0, 0); color: white; border: 1px solid rgba(0, 0, 0, 0); border: none; outline: none;')
        self.Button_MercadoInvisivel.setFixedSize(int(self.frameGeometry().width()*3/4*.90),int(self.frameGeometry().height()*3/4*.95))
        self.Button_MercadoInvisivel.setCursor(QCursor(Qt.PointingHandCursor))

        self.Button_Mercado = QPushButton("Análise do\nMercado")
        self.Button_Mercado.pressed.connect(lambda: self.CreatePage('5'))
        self.Button_Mercado.setFont(self.font30)
        self.Button_Mercado.setStyleSheet('background-color: black; color: white; border: none; outline: none')
        self.Button_Mercado.setFixedSize(int(self.frameGeometry().width()*3/4*.9),int(self.frameGeometry().height()/4*.95))
        self.Button_Mercado.setCursor(QCursor(Qt.PointingHandCursor))

        self.F_GLayout[self.GCount-1].addWidget(self.Background_1, 0, 2, 20, 1)
        self.F_GLayout[self.GCount-1].addWidget(self.Gif_Mercado, 0, 2, 15, 1, Qt.AlignCenter)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_MercadoInvisivel, 0, 2, 15, 1, Qt.AlignCenter)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_Mercado, 15, 2, 5, 1, Qt.AlignCenter)

        self.TextBox_NomeDeUsuario.setFocus()

        self.unsetCursor()

    def CreatePage1(self):
        self.Background_1 = QLabel()
        Background_1 = 'rgb(0, 0, 0)'
        self.Background_1.setStyleSheet("background-color: "+Background_1)

        self.Background_2 = QLabel()
        Background_2 = 'rgb(0, 0, 0)'
        self.Background_2.setStyleSheet("background-color: "+Background_2)

        self.Background_3 = QLabel()
        Background_3 = 'rgb(0, 0, 0)'
        self.Background_3.setStyleSheet("background-color: "+Background_3)

        self.Background_4 = QLabel()
        Background_4 = 'rgb(0, 0, 0)'
        self.Background_4.setStyleSheet("background-color: "+Background_4)

        self.Label_Titulo = QLabel('Minha carteira')
        self.Label_Titulo.setStyleSheet('color: white')
        self.Label_Titulo.setFont(self.font30)
        self.Label_Titulo.setAlignment(Qt.AlignCenter)

        self.Gif_Negociações = QLabel()
        self.Movie_Negociações = QMovie("./images/Trade_Icon.gif")
        self.Movie_Negociações.setScaledSize(QSize().scaled(int(self.frameGeometry().width()/3*.80),int(self.frameGeometry().height()/5*.95-20), Qt.KeepAspectRatio))
        self.Gif_Negociações.setMovie(self.Movie_Negociações)
        self.Movie_Negociações.start()

        self.Button_Negociações = QPushButton("Negociações")
        self.Button_Negociações.pressed.connect(lambda: self.CreatePage('10'))
        self.Button_Negociações.setFont(self.font24)
        self.Button_Negociações.setStyleSheet('background-color: black; color: white; border: 1px solid rgb(0, 0, 0)')
        self.Button_Negociações.setFixedSize(int(self.frameGeometry().width()/3*.96),int(self.frameGeometry().height()/5))
        self.Button_Negociações.setCursor(QCursor(Qt.PointingHandCursor))

        self.Button_NegociaçõesInvisivel = QPushButton()
        self.Button_NegociaçõesInvisivel.pressed.connect(lambda: self.CreatePage('10'))
        self.Button_NegociaçõesInvisivel.setStyleSheet('background-color: rgba(0, 0, 0, 0); color: white; border: 1px solid rgba(0, 0, 0, 0); border: none; outline: none;')
        self.Button_NegociaçõesInvisivel.setFixedSize(int(self.frameGeometry().width()/3*.98),int(self.frameGeometry().height()*4/5-160))
        self.Button_NegociaçõesInvisivel.setCursor(QCursor(Qt.PointingHandCursor))

        self.Gif_Tributacao = QLabel()
        self.Movie_Tributacao = QMovie("./images/GIF_Tributacao.gif")
        self.Movie_Tributacao.setScaledSize(QSize().scaled(int(self.frameGeometry().width()/3*.80),int(self.frameGeometry().height()/5*.95-20), Qt.KeepAspectRatio))
        self.Gif_Tributacao.setMovie(self.Movie_Tributacao)
        self.Movie_Tributacao.start()

        self.Button_Tributacao = QPushButton("Tributação")
        self.Button_Tributacao.pressed.connect(lambda: self.CreatePage('51'))
        self.Button_Tributacao.setFont(self.font24)
        self.Button_Tributacao.setStyleSheet('background-color: black; color: white; border: 1px solid rgb(0, 0, 0)')
        self.Button_Tributacao.setFixedSize(int(self.frameGeometry().width()/3*.96),int(self.frameGeometry().height()/5))
        self.Button_Tributacao.setCursor(QCursor(Qt.PointingHandCursor))

        self.Button_TributacaoInvisivel = QPushButton()
        self.Button_TributacaoInvisivel.pressed.connect(lambda: self.CreatePage('51'))
        self.Button_TributacaoInvisivel.setStyleSheet('background-color: rgba(0, 0, 0, 0); color: white; border: 1px solid rgba(0, 0, 0, 0); border: none; outline: none;')
        self.Button_TributacaoInvisivel.setFixedSize(int(self.frameGeometry().width()/3*.98),int(self.frameGeometry().height()*4/5-160))
        self.Button_TributacaoInvisivel.setCursor(QCursor(Qt.PointingHandCursor))

        self.Gif_ARCA = QLabel()
        self.Movie_ARCA = QMovie("./images/ArcaGifIcone.gif")
        self.Movie_ARCA.setScaledSize(QSize().scaled(int(self.frameGeometry().width()/3*.80),int(self.frameGeometry().height()/5*.95-20), Qt.KeepAspectRatio))
        self.Gif_ARCA.setMovie(self.Movie_ARCA)
        self.Movie_ARCA.start()

        self.Button_ARCA = QPushButton("ARCA")
        self.Button_ARCA.pressed.connect(lambda: self.CreatePage('11'))
        self.Button_ARCA.setFont(self.font24)
        self.Button_ARCA.setStyleSheet('background-color: black; color: white; border: 1px solid rgb(0, 0, 0)')
        self.Button_ARCA.setFixedSize(int(self.frameGeometry().width()/3*.96),int(self.frameGeometry().height()/5))
        self.Button_ARCA.setCursor(QCursor(Qt.PointingHandCursor))

        self.Button_ARCAInvisivel = QPushButton()
        self.Button_ARCAInvisivel.pressed.connect(lambda: self.CreatePage('11'))
        self.Button_ARCAInvisivel.setStyleSheet('background-color: rgba(0, 0, 0, 0); color: white; border: 1px solid rgba(0, 0, 0, 0); border: none; outline: none;')
        self.Button_ARCAInvisivel.setFixedSize(int(self.frameGeometry().width()/3*.80),int(self.frameGeometry().height()/5*.95-20))
        self.Button_ARCAInvisivel.setCursor(QCursor(Qt.PointingHandCursor))

        self.Gif_Desempenho = QLabel()
        self.Movie_Desempenho = QMovie("./images/Análise_de_Desempenho_Icon.gif")
        self.Movie_Desempenho.setScaledSize(QSize().scaled(int(self.frameGeometry().width()/3*.80),int(self.frameGeometry().height()/5*.95-20), Qt.KeepAspectRatio))
        self.Gif_Desempenho.setMovie(self.Movie_Desempenho)
        self.Movie_Desempenho.start()

        self.Button_Desempenho = QPushButton("Desempenho")
        self.Button_Desempenho.pressed.connect(lambda: self.CreatePage('12'))
        self.Button_Desempenho.setFont(self.font24)
        self.Button_Desempenho.setStyleSheet('background-color: black; color: white; border: 1px solid rgb(0, 0, 0)')
        self.Button_Desempenho.setFixedSize(int(self.frameGeometry().width()/3*.96),int(self.frameGeometry().height()/5))
        self.Button_Desempenho.setCursor(QCursor(Qt.PointingHandCursor))

        self.Button_DesempenhoInvisivel = QPushButton()
        self.Button_DesempenhoInvisivel.pressed.connect(lambda: self.CreatePage('12'))
        self.Button_DesempenhoInvisivel.setStyleSheet('background-color: rgba(0, 0, 0, 0); color: white; border: 1px solid rgba(0, 0, 0, 0); border: none; outline: none;')
        self.Button_DesempenhoInvisivel.setFixedSize(int(self.frameGeometry().width()/3*.80),int(self.frameGeometry().height()/5*.95-20))
        self.Button_DesempenhoInvisivel.setCursor(QCursor(Qt.PointingHandCursor))

        self.Gif_Mercado = QLabel()
        self.Movie_Mercado = QMovie("./images/AnaliseDoMercadoGifIcone.gif")
        self.Movie_Mercado.setScaledSize(QSize().scaled(int(self.frameGeometry().width()/3*.98),int(self.frameGeometry().height()*4/5-110), Qt.KeepAspectRatio))
        self.Gif_Mercado.setMovie(self.Movie_Mercado)
        self.Movie_Mercado.start()

        self.Button_Mercado = QPushButton("Análise do\nMercado")
        self.Button_Mercado.pressed.connect(lambda: self.CreatePage('13'))
        self.Button_Mercado.setFont(self.font24)
        self.Button_Mercado.setStyleSheet('background-color: black; color: white; border: 1px solid rgb(0, 0, 0)')
        self.Button_Mercado.setFixedSize(int(self.frameGeometry().width()/3*.96),int(self.frameGeometry().height()/5))
        self.Button_Mercado.setCursor(QCursor(Qt.PointingHandCursor))

        self.Button_MercadoInvisivel = QPushButton()
        self.Button_MercadoInvisivel.pressed.connect(lambda: self.CreatePage('13'))
        self.Button_MercadoInvisivel.setStyleSheet('background-color: rgba(0, 0, 0, 0); color: white; border: 1px solid rgba(0, 0, 0, 0); border: none; outline: none;')
        self.Button_MercadoInvisivel.setFixedSize(int(self.frameGeometry().width()/3*.98),int(self.frameGeometry().height()*4/5-110))
        self.Button_MercadoInvisivel.setCursor(QCursor(Qt.PointingHandCursor))

        self.F_GLayout[self.GCount-1].addWidget(self.Background_1, 0, 0, 1, 2)
        self.F_GLayout[self.GCount-1].addWidget(self.Background_2, 0, 0, 5, 1)
        self.F_GLayout[self.GCount-1].addWidget(self.Background_3, 0, 1, 5, 1)
        self.F_GLayout[self.GCount-1].addWidget(self.Background_4, 0, 2, 5, 1)
        self.F_GLayout[self.GCount-1].addWidget(self.Label_Titulo, 0, 0, 1, 2, Qt.AlignCenter | Qt.AlignTop)
        self.F_GLayout[self.GCount-1].addWidget(self.Gif_Negociações, 1, 0, 1, 1, Qt.AlignCenter)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_NegociaçõesInvisivel, 1, 0, 1, 1, Qt.AlignCenter)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_Negociações, 2, 0, 1, 1)
        self.F_GLayout[self.GCount-1].addWidget(self.Gif_Tributacao, 3, 0, 1, 1, Qt.AlignCenter)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_TributacaoInvisivel, 3, 0, 1, 1, Qt.AlignCenter)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_Tributacao, 4, 0, 1, 1)
        self.F_GLayout[self.GCount-1].addWidget(self.Gif_ARCA, 1, 1, 1, 1, Qt.AlignCenter)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_ARCAInvisivel, 1, 1, 1, 1, Qt.AlignCenter)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_ARCA, 2, 1, 1, 1)
        self.F_GLayout[self.GCount-1].addWidget(self.Gif_Desempenho, 3, 1, 1, 1, Qt.AlignCenter)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_DesempenhoInvisivel, 3, 1, 1, 1, Qt.AlignCenter)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_Desempenho, 4, 1, 1, 1)
        self.F_GLayout[self.GCount-1].addWidget(self.Gif_Mercado, 0, 2, 4, 1, Qt.AlignCenter)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_MercadoInvisivel, 0, 2, 4, 1, Qt.AlignCenter)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_Mercado, 4, 2, 1, 1)

    def CreatePage2(self):
        self.Background_1 = QLabel()
        Background_1 = 'rgb(0, 0, 0)'
        self.Background_1.setStyleSheet("background-color: "+Background_1)

        self.Label_Titulo = QLabel('Auxílio de senha')
        self.Label_Titulo.setStyleSheet('color: white')
        self.Label_Titulo.setFont(self.font30)
        self.Label_Titulo.setAlignment(Qt.AlignCenter)

        self.Label_Msg1 = QLabel('Insira o email associado à sua conta ou o nome de usuário:')
        self.Label_Msg1.setStyleSheet('color: white')
        self.Label_Msg1.setFont(self.font16)
        self.Label_Msg1.setAlignment(Qt.AlignCenter)

        self.TextBox_EmailOuUsuario = QLineEdit()
        self.TextBox_EmailOuUsuario.returnPressed.connect(lambda: self.CreatePage('6'))
        self.TextBox_EmailOuUsuario.setFixedWidth(int(self.frameGeometry().width()/3*0.9))

        self.Button_ConfirmacaoDeEmail = QPushButton('Seguinte')
        self.Button_ConfirmacaoDeEmail.pressed.connect(lambda: self.CreatePage('6'))
        self.Button_ConfirmacaoDeEmail.setFont(self.font16)
        self.Button_ConfirmacaoDeEmail.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
        self.Button_ConfirmacaoDeEmail.setFixedWidth(int(self.frameGeometry().width()/3*0.9))
        self.Button_ConfirmacaoDeEmail.setIcon(QIcon("./images/log_in.png"))
        self.Button_ConfirmacaoDeEmail.setCursor(QCursor(Qt.PointingHandCursor))
        self.Button_ConfirmacaoDeEmail.setStyleSheet("background-color: rgb(20, 120, 30)")

        self.F_GLayout[self.GCount-1].addWidget(self.Background_1, 0, 0, 20, 1)
        self.F_GLayout[self.GCount-1].addWidget(self.Label_Titulo, 6, 0, 1, 1, Qt.AlignCenter)
        self.F_GLayout[self.GCount-1].addWidget(self.Label_Msg1, 9, 0, 1, 1, Qt.AlignHCenter | Qt.AlignBottom)
        self.F_GLayout[self.GCount-1].addWidget(self.TextBox_EmailOuUsuario, 10, 0, 1, 1, Qt.AlignCenter | Qt.AlignTop)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_ConfirmacaoDeEmail, 12, 0, 2, 1, Qt.AlignHCenter)
        self.TextBox_EmailOuUsuario.setFocus()

    def CreatePage3(self):
        self.Background_1 = QLabel()
        Background_1 = 'rgb(0, 0, 0)'
        self.Background_1.setStyleSheet("background-color: "+Background_1)

        self.Background_2 = QLabel()
        Background_2 = 'rgb(0, 0, 0)'
        self.Background_2.setStyleSheet("background-color: "+Background_2)

        self.Background_3 = QLabel()
        Background_3 = 'rgb(0, 0, 0)'
        self.Background_3.setStyleSheet("background-color: "+Background_3)

        self.Background_4 = QLabel()
        Background_4 = 'rgb(0, 0, 0)'
        self.Background_4.setStyleSheet("background-color: "+Background_4)

        self.Label_Titulo = QLabel('Importar usuário')
        self.Label_Titulo.setStyleSheet('color: white')
        self.Label_Titulo.setFont(self.font30)
        self.Label_Titulo.setAlignment(Qt.AlignCenter)

        self.Button_ExportDBToExcel = QPushButton()
        self.Button_ExportDBToExcel.pressed.connect(lambda: self.OnButtonPressed('ImportFromExcel'))
        self.Button_ExportDBToExcel.setStyleSheet('background-color: rgb(0, 0, 0); color: white; border: 1px solid rgb(0, 0, 0)')
        self.Button_ExportDBToExcel.setFixedSize(int(self.frameGeometry().width()*2/5),int(self.frameGeometry().height()-160))
        self.Button_ExportDBToExcel.setIconSize(QSize(int(self.frameGeometry().width()*2/5),int(self.frameGeometry().height())))
        self.Button_ExportDBToExcel.setIcon(QIcon("./images/import_from_excel_icon.png"))
        self.Button_ExportDBToExcel.setCursor(QCursor(Qt.PointingHandCursor))

        self.Button_ExportDBToDB = QPushButton()
        self.Button_ExportDBToDB.pressed.connect(lambda: self.OnButtonPressed('ImportFromDB'))
        self.Button_ExportDBToDB.setStyleSheet('background-color: rgb(0, 0, 0); color: white; border: 1px solid rgb(0, 0, 0)')
        self.Button_ExportDBToDB.setFixedSize(int(self.frameGeometry().width()*2/5),int(self.frameGeometry().height()-160))
        self.Button_ExportDBToDB.setIconSize(QSize(int(self.frameGeometry().width()*2/5),int(self.frameGeometry().height())))
        self.Button_ExportDBToDB.setIcon(QIcon("./images/import_from_db_icon.png"))
        self.Button_ExportDBToDB.setCursor(QCursor(Qt.PointingHandCursor))

        self.F_GLayout[self.GCount-1].addWidget(self.Background_2, 0, 0, 1, 2)
        self.F_GLayout[self.GCount-1].addWidget(self.Background_3, 0, 0, 2, 1)
        self.F_GLayout[self.GCount-1].addWidget(self.Background_4, 0, 1, 2, 1)
        self.F_GLayout[self.GCount-1].addWidget(self.Label_Titulo, 0, 0, 1, 2, Qt.AlignCenter)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_ExportDBToExcel, 1, 0, 2, 1, Qt.AlignTop)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_ExportDBToDB, 1, 1, 2, 1, Qt.AlignTop)

    def CreatePage4(self):
        self.Background_1 = QLabel()
        Background_1 = 'rgb(0, 0, 0)'
        self.Background_1.setStyleSheet("background-color: "+Background_1)

        self.Label_Titulo = QLabel('Criação de usuário')
        self.Label_Titulo.setStyleSheet('color: white')
        self.Label_Titulo.setFont(self.font30)
        self.Label_Titulo.setAlignment(Qt.AlignCenter)

        self.Label_NomeDeUsuario = QLabel()
        self.Label_NomeDeUsuario.setStyleSheet('color: white')
        self.Label_NomeDeUsuario.setFont(self.font16)
        self.Label_NomeDeUsuario.setAlignment(Qt.AlignCenter)

        minimo = 24
        maximo = 36
        a = (int(self.screen_width*0.5))/((1/minimo)-(1/maximo))
        b = self.screen_width-(a/minimo)
        razao = (self.frameGeometry().width()-b)/a
        self.Label_NomeDeUsuario.setText('Nome de usuário:'+str((' ')*int(self.frameGeometry().width()*razao)))

        self.TextBox_NomeDeUsuario = QLineEdit()
        self.TextBox_NomeDeUsuario.returnPressed.connect(lambda: self.CreatePage('8'))
        self.TextBox_NomeDeUsuario.setFixedWidth(int(self.frameGeometry().width()/3*0.9))

        self.Label_Email = QLabel()
        self.Label_Email.setStyleSheet('color: white')
        self.Label_Email.setFont(self.font16)
        self.Label_Email.setAlignment(Qt.AlignCenter)
        minimo = 19
        maximo = 21
        a = (int(self.screen_width*0.5))/((1/minimo)-(1/maximo))
        b = self.screen_width-(a/minimo)
        razao = (self.frameGeometry().width()-b)/a
        self.Label_Email.setText('Email:'+str((' ')*int(self.frameGeometry().width()*razao)))

        self.TextBox_Email = QLineEdit()
        self.TextBox_Email.returnPressed.connect(lambda: self.CreatePage('8'))
        self.TextBox_Email.setFixedWidth(int(self.frameGeometry().width()/3*0.9))

        self.Label_Senha1 = QLabel()
        self.Label_Senha1.setStyleSheet('color: white')
        self.Label_Senha1.setFont(self.font16)
        self.Label_Senha1.setAlignment(Qt.AlignLeft)
        minimo = 19
        maximo = 21
        a = (int(self.screen_width*0.5))/((1/minimo)-(1/maximo))
        b = self.screen_width-(a/minimo)
        razao = (self.frameGeometry().width()-b)/a
        self.Label_Senha1.setText('Senha:'+str((' ')*int(self.frameGeometry().width()*razao)))

        self.TextBox_Senha1 = QLineEdit()
        self.TextBox_Senha1.returnPressed.connect(lambda: self.CreatePage('8'))
        self.TextBox_Senha1.setFixedWidth(int(self.frameGeometry().width()/3*0.9))
        self.TextBox_Senha1.setEchoMode(QLineEdit.Password)

        self.Label_Senha2 = QLabel()
        self.Label_Senha2.setStyleSheet('color: white')
        self.Label_Senha2.setFont(self.font16)
        self.Label_Senha2.setAlignment(Qt.AlignLeft)
        minimo = 24
        maximo = 39
        a = (int(self.screen_width*0.5))/((1/minimo)-(1/maximo))
        b = self.screen_width-(a/minimo)
        razao = (self.frameGeometry().width()-b)/a
        self.Label_Senha2.setText('Confirme a senha:'+str((' ')*int(self.frameGeometry().width()*razao)))

        self.TextBox_Senha2 = QLineEdit()
        self.TextBox_Senha2.returnPressed.connect(lambda: self.CreatePage('8'))
        self.TextBox_Senha2.setFixedWidth(int(self.frameGeometry().width()/3*0.9))
        self.TextBox_Senha2.setEchoMode(QLineEdit.Password)

        self.ComboBox_Currency = QComboBox()
        self.ComboBox_Currency.setFont(self.font16)
        self.ComboBox_Currency.setCursor(QCursor(Qt.PointingHandCursor))
        self.ComboBox_Currency.setStyleSheet("background-color: rgb(10, 10, 10) ; color: rgb(0, 255, 0);")
        for idx, coin in enumerate(self.DBManager.MoedasFiatBasicas):
            self.ComboBox_Currency.addItem(coin)

        self.Label_Path = QLabel('Local de armazenamento:')
        self.Label_Path.setStyleSheet('color: white')
        self.Label_Path.setFont(self.font14)
        self.Label_Path.setAlignment(Qt.AlignRight)

        self.Button_PathPadrao = QPushButton('PADRÃO')
        self.Button_PathPadrao.pressed.connect(lambda: self.OnButtonPressed('Path Padrão'))
        self.Button_PathPadrao.setStyleSheet("background-color: green; color: white;")
        self.Button_PathPadrao.setFont(self.font16)
        self.Button_PathPadrao.setCursor(QCursor(Qt.PointingHandCursor))

        self.Button_PathEspecifico = QPushButton('específico')
        self.Button_PathEspecifico.pressed.connect(lambda: self.OnButtonPressed('Path Específico'))
        self.Button_PathEspecifico.setStyleSheet("background-color: black; color: white;")
        self.Button_PathEspecifico.setFont(self.font16)
        self.Button_PathEspecifico.setCursor(QCursor(Qt.PointingHandCursor))

        self.TextBox_Path = QLineEdit()
        self.TextBox_Path.setFixedWidth(int(self.frameGeometry().width()*0.9))
        self.TextBox_Path.setEnabled(False)
        self.TextBox_Path.setAlignment(Qt.AlignCenter)
        self.TextBox_Path.setStyleSheet("background-color: black; color: white;")

        self.Button_ConfirmacaoDeEmail = QPushButton()
        self.Button_ConfirmacaoDeEmail.pressed.connect(lambda: self.CreatePage('8'))
        self.Button_ConfirmacaoDeEmail.setFont(self.font16)
        self.Button_ConfirmacaoDeEmail.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
        self.Button_ConfirmacaoDeEmail.setFixedWidth(int(self.frameGeometry().width()/3*0.9))
        self.Button_ConfirmacaoDeEmail.setIcon(QIcon("./images/log_in.png"))
        self.Button_ConfirmacaoDeEmail.setCursor(QCursor(Qt.PointingHandCursor))
        self.Button_ConfirmacaoDeEmail.setStyleSheet("background-color: rgb(20, 120, 30)")

        self.F_GLayout[self.GCount-1].addWidget(self.Background_1, 0, 0, 20, 2)
        self.F_GLayout[self.GCount-1].addWidget(self.Label_Titulo, 3, 0, 1, 2, Qt.AlignCenter)
        self.F_GLayout[self.GCount-1].addWidget(self.Label_NomeDeUsuario, 6, 0, 1, 2, Qt.AlignHCenter | Qt.AlignBottom)
        self.F_GLayout[self.GCount-1].addWidget(self.TextBox_NomeDeUsuario, 7, 0, 1, 2, Qt.AlignHCenter | Qt.AlignTop)
        self.F_GLayout[self.GCount-1].addWidget(self.Label_Email, 8, 0, 1, 2, Qt.AlignHCenter | Qt.AlignBottom)
        self.F_GLayout[self.GCount-1].addWidget(self.TextBox_Email, 9, 0, 1, 2, Qt.AlignHCenter | Qt.AlignTop)
        self.F_GLayout[self.GCount-1].addWidget(self.Label_Senha1, 10, 0, 1, 2, Qt.AlignHCenter | Qt.AlignBottom)
        self.F_GLayout[self.GCount-1].addWidget(self.TextBox_Senha1, 11, 0, 1, 2, Qt.AlignHCenter | Qt.AlignTop)
        self.F_GLayout[self.GCount-1].addWidget(self.Label_Senha2, 12, 0, 1, 2, Qt.AlignHCenter | Qt.AlignBottom)
        self.F_GLayout[self.GCount-1].addWidget(self.TextBox_Senha2, 13, 0, 1, 2, Qt.AlignHCenter | Qt.AlignTop)
        self.F_GLayout[self.GCount-1].addWidget(self.ComboBox_Currency, 14, 0, 1, 2, Qt.AlignHCenter | Qt.AlignTop)
        self.F_GLayout[self.GCount-1].addWidget(self.Label_Path, 15, 0, 1, 2, Qt.AlignHCenter | Qt.AlignTop)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_PathPadrao, 16, 0, Qt.AlignRight)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_PathEspecifico, 16, 1, Qt.AlignLeft)
        self.F_GLayout[self.GCount-1].addWidget(self.TextBox_Path, 17, 0, 1, 2, Qt.AlignCenter)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_ConfirmacaoDeEmail, 18, 0, 2, 2, Qt.AlignCenter)
        self.TextBox_NomeDeUsuario.setFocus()

    def CreatePage6(self, Email, NomeDeUsuario):
        def Thread(threadname, parametros):
            self.setCursor(Qt.WaitCursor)
            Email, NomeDeUsuario = parametros
            # Geração do código de confirmação
            self.CodigoDeConfirmacaoGerado = str(random.randint(0,9))+str(random.randint(0,9))+str(random.randint(0,9))+str(random.randint(0,9))+str(random.randint(0,9))+str(random.randint(0,9))
            self.EnviarEmailDeConfirmacao(Email, NomeDeUsuario, self.CodigoDeConfirmacaoGerado)
            self.unsetCursor()
            self.TextBox_CodigoDeConfirmacao.setFocus()
        _thread.start_new_thread(Thread, ('NiceThread',(Email, NomeDeUsuario)))

        self.Background_1 = QLabel()
        Background_1 = 'rgb(0, 0, 0)'
        self.Background_1.setStyleSheet("background-color: "+Background_1)

        self.Label_Titulo = QLabel('Confirmação de email')
        self.Label_Titulo.setStyleSheet('color: white')
        self.Label_Titulo.setFont(self.font30)
        self.Label_Titulo.setAlignment(Qt.AlignCenter)

        self.Label_Msg1 = QLabel('Mandamos para '+Email+' um código de confirmação da conta.')
        self.Label_Msg1.setStyleSheet('color: white')
        self.Label_Msg1.setFont(self.font16)
        self.Label_Msg1.setAlignment(Qt.AlignCenter)

        self.TextBox_CodigoDeConfirmacao = QLineEdit()
        self.TextBox_CodigoDeConfirmacao.returnPressed.connect(lambda: self.CreatePage('7'))
        self.TextBox_CodigoDeConfirmacao.setFixedWidth(int(self.frameGeometry().width()/3*0.9))

        self.Button_ResendEmail = QPushButton("Reenviar email")
        self.Button_ResendEmail.pressed.connect(lambda: _thread.start_new_thread(Thread, ('NiceThread',(Email, NomeDeUsuario))))
        self.Button_ResendEmail.setFont(self.font12)
        self.Button_ResendEmail.setCursor(QCursor(Qt.PointingHandCursor))
        self.Button_ResendEmail.setFixedHeight(int(self.frameGeometry().height()/10*.99))
        self.Button_ResendEmail.setFixedWidth(int(self.frameGeometry().width()/3*0.9*2/3))

        self.Button_ConfirmacaoDeEmail = QPushButton('Seguinte')
        self.Button_ConfirmacaoDeEmail.pressed.connect(lambda: self.CreatePage('7'))
        self.Button_ConfirmacaoDeEmail.setFont(self.font16)
        self.Button_ConfirmacaoDeEmail.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
        self.Button_ConfirmacaoDeEmail.setFixedWidth(int(self.frameGeometry().width()/3*0.9*2/3))
        self.Button_ConfirmacaoDeEmail.setIcon(QIcon("./images/log_in.png"))
        self.Button_ConfirmacaoDeEmail.setCursor(QCursor(Qt.PointingHandCursor))
        self.Button_ConfirmacaoDeEmail.setStyleSheet("background-color: rgb(20, 120, 30)")

        self.F_GLayout[self.GCount-1].addWidget(self.Background_1, 0, 0, 20, 2)
        self.F_GLayout[self.GCount-1].addWidget(self.Label_Titulo, 6, 0, 1, 2, Qt.AlignCenter)
        self.F_GLayout[self.GCount-1].addWidget(self.Label_Msg1, 9, 0, 1, 2, Qt.AlignHCenter | Qt.AlignBottom)
        self.F_GLayout[self.GCount-1].addWidget(self.TextBox_CodigoDeConfirmacao, 10, 0, 1, 2, Qt.AlignCenter | Qt.AlignTop)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_ResendEmail, 12, 0, 2, 1, Qt.AlignHCenter | Qt.AlignRight)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_ConfirmacaoDeEmail, 12, 1, 2, 1, Qt.AlignHCenter | Qt.AlignLeft)
        self.TextBox_CodigoDeConfirmacao.setFocus()

    def CreatePage7(self):
        self.Background_1 = QLabel()
        Background_1 = 'rgb(0, 0, 0)'
        self.Background_1.setStyleSheet("background-color: "+Background_1)

        self.Label_Titulo = QLabel('Alteração de senha')
        self.Label_Titulo.setStyleSheet('color: white')
        self.Label_Titulo.setFont(self.font30)
        self.Label_Titulo.setAlignment(Qt.AlignCenter)

        self.Label_Msg1 = QLabel('Digite uma nova senha:')
        self.Label_Msg1.setStyleSheet('color: white')
        self.Label_Msg1.setFont(self.font16)
        self.Label_Msg1.setAlignment(Qt.AlignCenter)

        self.TextBox_NovaSenha1 = QLineEdit()
        self.TextBox_NovaSenha1.returnPressed.connect(lambda: self.CreatePage('1'))
        self.TextBox_NovaSenha1.setFixedWidth(int(self.frameGeometry().width()/3*0.9))
        self.TextBox_NovaSenha1.setEchoMode(QLineEdit.Password)

        self.Label_Msg2 = QLabel('Digite novamente a senha:')
        self.Label_Msg2.setStyleSheet('color: white')
        self.Label_Msg2.setFont(self.font16)
        self.Label_Msg2.setAlignment(Qt.AlignCenter)

        self.TextBox_NovaSenha2 = QLineEdit()
        self.TextBox_NovaSenha2.returnPressed.connect(lambda: self.CreatePage('1'))
        self.TextBox_NovaSenha2.setFixedWidth(int(self.frameGeometry().width()/3*0.9))
        self.TextBox_NovaSenha2.setEchoMode(QLineEdit.Password)

        self.Button_ConfirmacaoDeEmail = QPushButton('Recuperar')
        self.Button_ConfirmacaoDeEmail.pressed.connect(lambda: self.CreatePage('1'))
        self.Button_ConfirmacaoDeEmail.setFont(self.font16)
        self.Button_ConfirmacaoDeEmail.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
        self.Button_ConfirmacaoDeEmail.setFixedWidth(int(self.frameGeometry().width()/3*0.9))
        self.Button_ConfirmacaoDeEmail.setIcon(QIcon("./images/log_in.png"))
        self.Button_ConfirmacaoDeEmail.setCursor(QCursor(Qt.PointingHandCursor))
        self.Button_ConfirmacaoDeEmail.setStyleSheet("background-color: rgb(20, 120, 30)")

        self.F_GLayout[self.GCount-1].addWidget(self.Background_1, 0, 0, 20, 1)
        self.F_GLayout[self.GCount-1].addWidget(self.Label_Titulo, 6, 0, Qt.AlignCenter)
        self.F_GLayout[self.GCount-1].addWidget(self.Label_Msg1, 9, 0, Qt.AlignHCenter | Qt.AlignBottom)
        self.F_GLayout[self.GCount-1].addWidget(self.TextBox_NovaSenha1, 10, 0, Qt.AlignCenter | Qt.AlignTop)
        self.F_GLayout[self.GCount-1].addWidget(self.Label_Msg2, 11, 0, Qt.AlignHCenter | Qt.AlignBottom)
        self.F_GLayout[self.GCount-1].addWidget(self.TextBox_NovaSenha2, 12, 0, Qt.AlignCenter | Qt.AlignTop)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_ConfirmacaoDeEmail, 14, 0, Qt.AlignHCenter)
        self.TextBox_NovaSenha1.setFocus()

    def CreatePage8(self):
        Email = self.TextBox_Email.text()
        NomeDeUsuario = self.TextBox_NomeDeUsuario.text()

        # Envio de email com o código
        def Thread(threadname, parametros):
            self.setCursor(Qt.WaitCursor)
            Email, NomeDeUsuario = parametros
            # Geração do código de confirmação
            self.CodigoDeConfirmacaoGerado = str(random.randint(0,9))+str(random.randint(0,9))+str(random.randint(0,9))+str(random.randint(0,9))+str(random.randint(0,9))+str(random.randint(0,9))
            self.EnviarEmailDeConfirmacao(Email, NomeDeUsuario, self.CodigoDeConfirmacaoGerado)
            self.unsetCursor()
            self.TextBox_CodigoDeConfirmacao.setFocus()
        _thread.start_new_thread(Thread, ('NiceThread',(Email, NomeDeUsuario)))

        self.Background_1 = QLabel()
        Background_1 = 'rgb(0, 0, 0)'
        self.Background_1.setStyleSheet("background-color: "+Background_1)

        self.Label_Titulo = QLabel('Confirmação de email')
        self.Label_Titulo.setStyleSheet('color: white')
        self.Label_Titulo.setFont(self.font30)
        self.Label_Titulo.setAlignment(Qt.AlignCenter)

        self.Label_Msg1 = QLabel('Mandamos para '+Email+' um código de confirmação da conta.')
        self.Label_Msg1.setStyleSheet('color: white')
        self.Label_Msg1.setFont(self.font16)
        self.Label_Msg1.setAlignment(Qt.AlignCenter)

        self.Label_Msg2 = QLabel('Insira no campo abaixo o código de confirmação:')
        self.Label_Msg2.setStyleSheet('color: white')
        self.Label_Msg2.setFont(self.font16)
        self.Label_Msg2.setAlignment(Qt.AlignCenter)

        self.TextBox_CodigoDeConfirmacao = QLineEdit()
        self.TextBox_CodigoDeConfirmacao.returnPressed.connect(lambda: self.CreatePage('1'))
        self.TextBox_CodigoDeConfirmacao.setFixedWidth(int(self.frameGeometry().width()/3*0.9))

        self.Button_ResendEmail = QPushButton("Reenviar email")
        self.Button_ResendEmail.pressed.connect(lambda: _thread.start_new_thread(Thread, ('NiceThread',(Email, NomeDeUsuario))))
        self.Button_ResendEmail.setFont(self.font12)
        self.Button_ResendEmail.setCursor(QCursor(Qt.PointingHandCursor))
        self.Button_ResendEmail.setFixedHeight(int(self.frameGeometry().height()/10*.99))
        self.Button_ResendEmail.setFixedWidth(int(self.frameGeometry().width()/3*0.9*2/3))

        self.Button_ConfirmacaoDeEmail = QPushButton('Confirmar')
        self.Button_ConfirmacaoDeEmail.pressed.connect(lambda: self.CreatePage('1'))
        self.Button_ConfirmacaoDeEmail.setFont(self.font16)
        self.Button_ConfirmacaoDeEmail.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
        self.Button_ConfirmacaoDeEmail.setFixedWidth(int(self.frameGeometry().width()/3*0.9*2/3))
        self.Button_ConfirmacaoDeEmail.setIcon(QIcon("./images/log_in.png"))
        self.Button_ConfirmacaoDeEmail.setCursor(QCursor(Qt.PointingHandCursor))
        self.Button_ConfirmacaoDeEmail.setStyleSheet("background-color: rgb(20, 120, 30)")

        self.F_GLayout[self.GCount-1].addWidget(self.Background_1, 0, 0, 20, 2)
        self.F_GLayout[self.GCount-1].addWidget(self.Label_Titulo, 6, 0, 1, 2, Qt.AlignCenter)
        self.F_GLayout[self.GCount-1].addWidget(self.Label_Msg1, 8, 0, 1, 2, Qt.AlignHCenter)
        self.F_GLayout[self.GCount-1].addWidget(self.Label_Msg2, 10, 0, 1, 2, Qt.AlignHCenter | Qt.AlignBottom)
        self.F_GLayout[self.GCount-1].addWidget(self.TextBox_CodigoDeConfirmacao, 11, 0, 1, 2, Qt.AlignCenter | Qt.AlignTop)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_ResendEmail, 13, 0, 2, 1, Qt.AlignHCenter | Qt.AlignRight)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_ConfirmacaoDeEmail, 13, 1, 2, 1, Qt.AlignHCenter | Qt.AlignLeft)
        self.TextBox_CodigoDeConfirmacao.setFocus()

    def CreatePage9(self):
        self.Background_1 = QLabel()
        Background_1 = 'rgb(0, 0, 0)'
        self.Background_1.setStyleSheet("background-color: "+Background_1)

        self.Background_2 = QLabel()
        Background_2 = 'rgb(0, 0, 0)'
        self.Background_2.setStyleSheet("background-color: "+Background_2)

        self.Background_3 = QLabel()
        Background_3 = 'rgb(0, 0, 0)'
        self.Background_3.setStyleSheet("background-color: "+Background_3)

        self.Background_4 = QLabel()
        Background_4 = 'rgb(0, 0, 0)'
        self.Background_4.setStyleSheet("background-color: "+Background_4)

        self.Label_Titulo1 = QLabel('Perfil')
        self.Label_Titulo1.setStyleSheet('color: white')
        self.Label_Titulo1.setFont(self.font30)
        self.Label_Titulo1.setAlignment(Qt.AlignCenter)

        self.Button_ProfilePhoto = QPushButton()
        self.Button_ProfilePhoto.pressed.connect(lambda: self.OnButtonPressed("ChangeProfilePhoto"))
        if os.path.exists(os.getcwd()+"\\images\\ProfilePhoto_"+self.User+".jpg"):
            self.Button_ProfilePhoto.setIcon(QIcon("./images/ProfilePhoto_"+self.User+".jpg"))
        else:
            self.Button_ProfilePhoto.setIcon(QIcon("./images/ProfilePhoto.jpg"))
        self.Button_ProfilePhoto.setFixedSize(int(self.frameGeometry().width()/8),int(self.frameGeometry().width()/8))
        self.Button_ProfilePhoto.setIconSize(QSize(int(self.frameGeometry().width()/8*.95),int(self.frameGeometry().width()/8*.95)))
        self.Button_ProfilePhoto.setCursor(QCursor(Qt.PointingHandCursor))

        self.Label_UserEmail = QLabel()
        self.Label_UserEmail.setText(self.User+'\n'+self.Email)
        self.Label_UserEmail.setStyleSheet('color: white')
        self.Label_UserEmail.setFont(self.font12)
        self.Label_UserEmail.setAlignment(Qt.AlignCenter)

        self.Button_MudarEmail = QPushButton("Mudar o email")
        self.Button_MudarEmail.pressed.connect(lambda: self.CreatePage('14'))
        self.Button_MudarEmail.setStyleSheet('background-color: rgb(10, 10, 20); color: white')
        self.Button_MudarEmail.setFont(self.font12)
        self.Button_MudarEmail.setCursor(QCursor(Qt.PointingHandCursor))

        self.Button_MudarSenha = QPushButton("Mudar a senha")
        self.Button_MudarSenha.pressed.connect(lambda: self.CreatePage('15'))
        self.Button_MudarSenha.setStyleSheet('background-color: rgb(10, 10, 20); color: white')
        self.Button_MudarSenha.setFont(self.font12)
        self.Button_MudarSenha.setCursor(QCursor(Qt.PointingHandCursor))

        self.Button_MudarPath = QPushButton("Mudar o caminho")
        self.Button_MudarPath.pressed.connect(lambda: self.OnButtonPressed('Change UserPath'))
        self.Button_MudarPath.setStyleSheet('background-color: rgb(10, 10, 20); color: white')
        self.Button_MudarPath.setFont(self.font12)
        self.Button_MudarPath.setCursor(QCursor(Qt.PointingHandCursor))

        self.Button_EndAccount = QPushButton('Encerrar a conta')
        self.Button_EndAccount.pressed.connect(lambda: self.CreatePage('16'))
        self.Button_EndAccount.setStyleSheet('background-color: rgb(20, 10, 10); color: white')
        self.Button_EndAccount.setFixedHeight(40)
        self.Button_EndAccount.setIconSize(QSize(35, 35))
        self.Button_EndAccount.setIcon(QIcon("./images/closeAccount_icon.png"))
        self.Button_EndAccount.setCursor(QCursor(Qt.PointingHandCursor))

        self.Button_VersionInfo = QPushButton()
        self.Button_VersionInfo.pressed.connect(lambda: QMessageBox.about(self,'Informação', self.APPVersion+"\n\nCaminho do usuário:\n"+self.DBManager.UserPath+"/"+self.User+".db"))
        self.Button_VersionInfo.setIcon(QIcon("./images/Info.png"))
        self.Button_VersionInfo.setCursor(QCursor(Qt.PointingHandCursor))
        self.Button_VersionInfo.setStyleSheet("background-color: black; border: none; outline: none;")

        self.F_GLayout[self.GCount-1].addWidget(self.Background_1, 0, 0, 10, 1)
        self.F_GLayout[self.GCount-1].addWidget(self.Label_Titulo1, 0, 0, 1, 1, Qt.AlignLeft | Qt.AlignTop)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_ProfilePhoto, 2, 0, 2, 1, Qt.AlignBottom | Qt.AlignHCenter)
        self.F_GLayout[self.GCount-1].addWidget(self.Label_UserEmail, 4, 0, 1, 1, Qt.AlignTop | Qt.AlignHCenter)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_MudarEmail, 5, 0, 1, 1, Qt.AlignCenter | Qt.AlignBottom)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_MudarSenha, 6, 0, 1, 1, Qt.AlignCenter)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_MudarPath, 7, 0, 1, 1, Qt.AlignCenter | Qt.AlignTop)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_EndAccount, 9, 0, 1, 1, Qt.AlignBottom | Qt.AlignLeft)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_VersionInfo, 9, 0, Qt.AlignRight)

        self.InsertGridLayout(1, 3, 1, 17)

        self.Label_Titulo2 = QLabel('Exportar Base de Dados')
        self.Label_Titulo2.setStyleSheet('color: white')
        self.Label_Titulo2.setFont(self.font30)
        self.Label_Titulo2.setAlignment(Qt.AlignCenter)

        self.Button_ExportDBToExcel = QPushButton()
        self.Button_ExportDBToExcel.pressed.connect(lambda: self.OnButtonPressed('ExportToExcel'))
        self.Button_ExportDBToExcel.setStyleSheet('background-color: rgb(0, 0, 0); color: white; border: none; outline: none;')
        self.Button_ExportDBToExcel.setFixedSize(int(self.frameGeometry().width()*2/5),int(self.frameGeometry().height()-160))
        self.Button_ExportDBToExcel.setIconSize(QSize(int(self.frameGeometry().width()*2/5),int(self.frameGeometry().height())))
        self.Button_ExportDBToExcel.setIcon(QIcon("./images/export_to_excel_icon.png"))
        self.Button_ExportDBToExcel.setCursor(QCursor(Qt.PointingHandCursor))

        self.Button_ExportDBToDB = QPushButton()
        self.Button_ExportDBToDB.pressed.connect(lambda: self.OnButtonPressed('ExportToDB'))
        self.Button_ExportDBToDB.setStyleSheet('background-color: rgb(0, 0, 0); color: white; border: none; outline: none;')
        self.Button_ExportDBToDB.setFixedSize(int(self.frameGeometry().width()*2/5),int(self.frameGeometry().height()-160))
        self.Button_ExportDBToDB.setIconSize(QSize(int(self.frameGeometry().width()*2/5),int(self.frameGeometry().height())))
        self.Button_ExportDBToDB.setIcon(QIcon("./images/export_to_db_icon.png"))
        self.Button_ExportDBToDB.setCursor(QCursor(Qt.PointingHandCursor))

        self.F_GLayout[self.GCount-1].addWidget(self.Background_2, 0, 0, 1, 2)
        self.F_GLayout[self.GCount-1].addWidget(self.Background_3, 0, 0, 2, 1)
        self.F_GLayout[self.GCount-1].addWidget(self.Background_4, 0, 1, 2, 1)
        self.F_GLayout[self.GCount-1].addWidget(self.Label_Titulo2, 0, 0, 1, 2, Qt.AlignCenter | Qt.AlignTop)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_ExportDBToExcel, 1, 0, 2, 1, Qt.AlignTop)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_ExportDBToDB, 1, 1, 2, 1, Qt.AlignTop)

    def CreatePage14(self):
        self.Background_1 = QLabel()
        Background_1 = 'rgb(0, 0, 0)'
        self.Background_1.setStyleSheet("background-color: "+Background_1)

        self.Label_Titulo = QLabel('Alteração de email')
        self.Label_Titulo.setStyleSheet('color: white')
        self.Label_Titulo.setFont(self.font30)
        self.Label_Titulo.setAlignment(Qt.AlignCenter)

        self.Label_Msg1 = QLabel('Seu email atual é o\n'+self.Email)
        self.Label_Msg1.setStyleSheet('color: white')
        self.Label_Msg1.setFont(self.font16)
        self.Label_Msg1.setAlignment(Qt.AlignCenter)

        self.Label_Msg2 = QLabel('Digite o seu novo email:')
        self.Label_Msg2.setStyleSheet('color: white')
        self.Label_Msg2.setFont(self.font16)
        self.Label_Msg2.setAlignment(Qt.AlignCenter)

        self.TextBox_Email = QLineEdit()
        self.TextBox_Email.returnPressed.connect(lambda: self.CreatePage('17a'))
        self.TextBox_Email.setFixedWidth(int(self.frameGeometry().width()/3*0.9))

        self.Button_ConfirmacaoDeEmail = QPushButton('Seguinte')
        self.Button_ConfirmacaoDeEmail.pressed.connect(lambda: self.CreatePage('17a'))
        self.Button_ConfirmacaoDeEmail.setFont(self.font16)
        self.Button_ConfirmacaoDeEmail.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
        self.Button_ConfirmacaoDeEmail.setFixedWidth(int(self.frameGeometry().width()/3*0.9))
        self.Button_ConfirmacaoDeEmail.setIcon(QIcon("./images/log_in.png"))
        self.Button_ConfirmacaoDeEmail.setCursor(QCursor(Qt.PointingHandCursor))
        self.Button_ConfirmacaoDeEmail.setStyleSheet("background-color: rgb(20, 120, 30)")

        self.F_GLayout[self.GCount-1].addWidget(self.Background_1, 0, 0, 20, 1)
        self.F_GLayout[self.GCount-1].addWidget(self.Label_Titulo, 6, 0, 1, 1, Qt.AlignCenter)
        self.F_GLayout[self.GCount-1].addWidget(self.Label_Msg1, 8, 0, 1, 1, Qt.AlignHCenter)
        self.F_GLayout[self.GCount-1].addWidget(self.Label_Msg2, 9, 0, 1, 1, Qt.AlignHCenter | Qt.AlignBottom)
        self.F_GLayout[self.GCount-1].addWidget(self.TextBox_Email, 10, 0, 1, 1, Qt.AlignCenter | Qt.AlignTop)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_ConfirmacaoDeEmail, 12, 0, 2, 1, Qt.AlignHCenter)
        self.TextBox_Email.setFocus()

    def CreatePage15(self):
        self.Background_1 = QLabel()
        Background_1 = 'rgb(0, 0, 0)'
        self.Background_1.setStyleSheet("background-color: "+Background_1)

        self.Label_Titulo = QLabel('Alteração de senha')
        self.Label_Titulo.setStyleSheet('color: white')
        self.Label_Titulo.setFont(self.font30)
        self.Label_Titulo.setAlignment(Qt.AlignCenter)

        self.Label_SenhaAtual = QLabel()
        self.Label_SenhaAtual.setStyleSheet('color: white')
        self.Label_SenhaAtual.setFont(self.font16)
        self.Label_SenhaAtual.setAlignment(Qt.AlignLeft)
        minimo = 24
        maximo = 43
        a = (int(self.screen_width*0.5))/((1/minimo)-(1/maximo))
        b = self.screen_width-(a/minimo)
        razao = (self.frameGeometry().width()-b)/a
        self.Label_SenhaAtual.setText('Insira a senha atual:'+str((' ')*int(self.frameGeometry().width()*razao)))

        self.TextBox_SenhaAtual = QLineEdit()
        self.TextBox_SenhaAtual.returnPressed.connect(lambda: self.CreatePage('9_'))
        self.TextBox_SenhaAtual.setFixedWidth(int(self.frameGeometry().width()/3*0.9))
        self.TextBox_SenhaAtual.setEchoMode(QLineEdit.Password)

        self.Label_Senha1 = QLabel()
        self.Label_Senha1.setStyleSheet('color: white')
        self.Label_Senha1.setFont(self.font16)
        self.Label_Senha1.setAlignment(Qt.AlignLeft)
        minimo = 24
        maximo = 47
        a = (int(self.screen_width*0.5))/((1/minimo)-(1/maximo))
        b = self.screen_width-(a/minimo)
        razao = (self.frameGeometry().width()-b)/a
        self.Label_Senha1.setText('Crie uma nova senha:'+str((' ')*int(self.frameGeometry().width()*razao)))

        self.TextBox_Senha1 = QLineEdit()
        self.TextBox_Senha1.returnPressed.connect(lambda: self.CreatePage('9_'))
        self.TextBox_Senha1.setFixedWidth(int(self.frameGeometry().width()/3*0.9))
        self.TextBox_Senha1.setEchoMode(QLineEdit.Password)

        self.Label_Senha2 = QLabel()
        self.Label_Senha2.setStyleSheet('color: white')
        self.Label_Senha2.setFont(self.font16)
        self.Label_Senha2.setAlignment(Qt.AlignLeft)
        minimo = 23
        maximo = 43
        a = (int(self.screen_width*0.5))/((1/minimo)-(1/maximo))
        b = self.screen_width-(a/minimo)
        razao = (self.frameGeometry().width()-b)/a
        self.Label_Senha2.setText('Confirme a senha:'+str((' ')*int(self.frameGeometry().width()*razao)))

        self.TextBox_Senha2 = QLineEdit()
        self.TextBox_Senha2.returnPressed.connect(lambda: self.CreatePage('9_'))
        self.TextBox_Senha2.setFixedWidth(int(self.frameGeometry().width()/3*0.9))
        self.TextBox_Senha2.setEchoMode(QLineEdit.Password)

        self.Button_Alterar = QPushButton()
        self.Button_Alterar.pressed.connect(lambda: self.CreatePage('9_'))
        self.Button_Alterar.setFont(self.font16)
        self.Button_Alterar.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
        self.Button_Alterar.setFixedWidth(int(self.frameGeometry().width()/3*0.9))
        self.Button_Alterar.setIcon(QIcon("./images/log_in.png"))
        self.Button_Alterar.setCursor(QCursor(Qt.PointingHandCursor))
        self.Button_Alterar.setStyleSheet("background-color: rgb(20, 120, 30)")

        self.F_GLayout[self.GCount-1].addWidget(self.Background_1, 0, 0, 20, 1)
        self.F_GLayout[self.GCount-1].addWidget(self.Label_Titulo, 3, 0, Qt.AlignCenter)
        self.F_GLayout[self.GCount-1].addWidget(self.Label_SenhaAtual, 7, 0, Qt.AlignHCenter | Qt.AlignBottom)
        self.F_GLayout[self.GCount-1].addWidget(self.TextBox_SenhaAtual, 8, 0, Qt.AlignHCenter | Qt.AlignTop)
        self.F_GLayout[self.GCount-1].addWidget(self.Label_Senha1, 9, 0, Qt.AlignHCenter | Qt.AlignBottom)
        self.F_GLayout[self.GCount-1].addWidget(self.TextBox_Senha1, 10, 0, Qt.AlignHCenter | Qt.AlignTop)
        self.F_GLayout[self.GCount-1].addWidget(self.Label_Senha2, 11, 0, Qt.AlignHCenter | Qt.AlignBottom)
        self.F_GLayout[self.GCount-1].addWidget(self.TextBox_Senha2, 12, 0, Qt.AlignHCenter | Qt.AlignTop)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_Alterar, 14, 0, 2, 1, Qt.AlignCenter)
        self.TextBox_SenhaAtual.setFocus()

    def CreatePage16(self):
        MessageBox_Msg1 = QMessageBox()
        MessageBox_Msg1.setWindowTitle("Encerrar conta")
        MessageBox_Msg1.setText("Tem certeza que deseja encerrar a sua conta?\n\nNote que esse processo é irreversível!\n")
        MessageBox_Msg1.setIcon(QMessageBox.Warning)
        MessageBox_Msg1.setStandardButtons(QMessageBox.Yes|QMessageBox.Cancel)
        MessageBox_Msg1.setDefaultButton(QMessageBox.Cancel)
        MessageBox_Msg1.setWindowIcon(QIcon(".\images\logoARCA.png"))

        returnValue = MessageBox_Msg1.exec()
        if returnValue == QMessageBox.Yes:
            self.CreatePage('17b')

        self.unsetCursor()

    def CreatePage17a(self, Email, NomeDeUsuario):
        self.Background_1 = QLabel()
        Background_1 = 'rgb(0, 0, 0)'
        self.Background_1.setStyleSheet("background-color: "+Background_1)

        self.Label_Titulo = QLabel('Confirmação de email')
        self.Label_Titulo.setStyleSheet('color: white')
        self.Label_Titulo.setFont(self.font30)
        self.Label_Titulo.setAlignment(Qt.AlignCenter)

        self.Label_Msg1 = QLabel('Mandamos para '+Email+' um código de confirmação da conta.')
        self.Label_Msg1.setStyleSheet('color: white')
        self.Label_Msg1.setFont(self.font16)
        self.Label_Msg1.setAlignment(Qt.AlignCenter)

        self.Label_Msg2 = QLabel('Insira no campo abaixo o código de confirmação:')
        self.Label_Msg2.setStyleSheet('color: white')
        self.Label_Msg2.setFont(self.font16)
        self.Label_Msg2.setAlignment(Qt.AlignCenter)

        self.TextBox_CodigoDeConfirmacao = QLineEdit()
        self.TextBox_CodigoDeConfirmacao.returnPressed.connect(lambda: self.CreatePage('9_'))
        self.TextBox_CodigoDeConfirmacao.setFixedWidth(int(self.frameGeometry().width()/3*0.9))

        self.Button_ResendEmail = QPushButton("Reenviar email")
        # self.Button_ResendEmail.pressed.connect(lambda: self.EnviarEmail(Email, NomeDeUsuario, CodigoDeConfirmacaoGerado))
        self.Button_ResendEmail.setFont(self.font12)
        self.Button_ResendEmail.setCursor(QCursor(Qt.PointingHandCursor))
        self.Button_ResendEmail.setFixedHeight(int(self.frameGeometry().height()/10*.99))
        self.Button_ResendEmail.setFixedWidth(int(self.frameGeometry().width()/3*0.9*2/3))

        self.Button_ConfirmacaoDeEmail = QPushButton('Alterar')
        self.Button_ConfirmacaoDeEmail.pressed.connect(lambda: self.CreatePage('9_'))
        self.Button_ConfirmacaoDeEmail.setFont(self.font16)
        self.Button_ConfirmacaoDeEmail.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
        self.Button_ConfirmacaoDeEmail.setFixedWidth(int(self.frameGeometry().width()/3*0.9*2/3))
        self.Button_ConfirmacaoDeEmail.setIcon(QIcon("./images/log_in.png"))
        self.Button_ConfirmacaoDeEmail.setCursor(QCursor(Qt.PointingHandCursor))
        self.Button_ConfirmacaoDeEmail.setStyleSheet("background-color: rgb(20, 120, 30)")

        self.F_GLayout[self.GCount-1].addWidget(self.Background_1, 0, 0, 20, 2)
        self.F_GLayout[self.GCount-1].addWidget(self.Label_Titulo, 6, 0, 1, 2, Qt.AlignCenter)
        self.F_GLayout[self.GCount-1].addWidget(self.Label_Msg1, 8, 0, 1, 2, Qt.AlignHCenter)
        self.F_GLayout[self.GCount-1].addWidget(self.Label_Msg2, 10, 0, 1, 2, Qt.AlignHCenter | Qt.AlignBottom)
        self.F_GLayout[self.GCount-1].addWidget(self.TextBox_CodigoDeConfirmacao, 11, 0, 1, 2, Qt.AlignCenter | Qt.AlignTop)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_ResendEmail, 13, 0, 2, 1, Qt.AlignHCenter | Qt.AlignRight)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_ConfirmacaoDeEmail, 13, 1, 2, 1, Qt.AlignHCenter | Qt.AlignLeft)
        self.TextBox_CodigoDeConfirmacao.setFocus()

    def CreatePage17b(self):
        Email = self.Email
        NomeDeUsuario = self.User

        # Envio de email com o código
        def Thread(threadname, parametros):
            self.setCursor(Qt.WaitCursor)
            Email, NomeDeUsuario = parametros
            # Geração do código de confirmação
            self.CodigoDeConfirmacaoGerado = str(random.randint(0,9))+str(random.randint(0,9))+str(random.randint(0,9))+str(random.randint(0,9))+str(random.randint(0,9))+str(random.randint(0,9))
            self.EnviarEmailDeConfirmacao(Email, NomeDeUsuario, self.CodigoDeConfirmacaoGerado)
            self.unsetCursor()
            self.TextBox_CodigoDeConfirmacao.setFocus()
        _thread.start_new_thread(Thread, ('NiceThread',(Email, NomeDeUsuario)))

        self.Background_1 = QLabel()
        Background_1 = 'rgb(0, 0, 0)'
        self.Background_1.setStyleSheet("background-color: "+Background_1)

        self.Label_Titulo = QLabel('Confirmação de email')
        self.Label_Titulo.setStyleSheet('color: white')
        self.Label_Titulo.setFont(self.font30)
        self.Label_Titulo.setAlignment(Qt.AlignCenter)

        self.Label_Msg1 = QLabel('Mandamos para '+Email+' um código.\nEssa etapa é só pra ter certeza que é você mesmo que está decidindo encerrar a conta.')
        self.Label_Msg1.setStyleSheet('color: white')
        self.Label_Msg1.setFont(self.font16)
        self.Label_Msg1.setAlignment(Qt.AlignCenter)

        self.Label_Msg2 = QLabel('Insira no campo abaixo o código de confirmação:')
        self.Label_Msg2.setStyleSheet('color: white')
        self.Label_Msg2.setFont(self.font16)
        self.Label_Msg2.setAlignment(Qt.AlignCenter)

        self.TextBox_CodigoDeConfirmacao = QLineEdit()
        self.TextBox_CodigoDeConfirmacao.returnPressed.connect(lambda: self.CreatePage('0'))
        self.TextBox_CodigoDeConfirmacao.setFixedWidth(int(self.frameGeometry().width()/3*0.9))

        self.Button_ResendEmail = QPushButton("Reenviar email")
        self.Button_ResendEmail.pressed.connect(lambda: _thread.start_new_thread(Thread, ('NiceThread',(Email, NomeDeUsuario))))
        self.Button_ResendEmail.setFont(self.font12)
        self.Button_ResendEmail.setCursor(QCursor(Qt.PointingHandCursor))
        self.Button_ResendEmail.setFixedHeight(int(self.frameGeometry().height()/10*.99))
        self.Button_ResendEmail.setFixedWidth(int(self.frameGeometry().width()/3*0.9*2/3))

        self.Button_ConfirmacaoDeEmail = QPushButton('Encerrar conta')
        self.Button_ConfirmacaoDeEmail.pressed.connect(lambda: self.CreatePage('0'))
        self.Button_ConfirmacaoDeEmail.setFont(self.font16)
        self.Button_ConfirmacaoDeEmail.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
        self.Button_ConfirmacaoDeEmail.setFixedWidth(int(self.frameGeometry().width()/3*0.9*2/3))
        self.Button_ConfirmacaoDeEmail.setIcon(QIcon("./images/log_in.png"))
        self.Button_ConfirmacaoDeEmail.setCursor(QCursor(Qt.PointingHandCursor))
        self.Button_ConfirmacaoDeEmail.setStyleSheet("background-color: rgb(20, 120, 30)")

        self.F_GLayout[self.GCount-1].addWidget(self.Background_1, 0, 0, 20, 2)
        self.F_GLayout[self.GCount-1].addWidget(self.Label_Titulo, 6, 0, 1, 2, Qt.AlignCenter)
        self.F_GLayout[self.GCount-1].addWidget(self.Label_Msg1, 8, 0, 1, 2, Qt.AlignHCenter)
        self.F_GLayout[self.GCount-1].addWidget(self.Label_Msg2, 10, 0, 1, 2, Qt.AlignHCenter | Qt.AlignBottom)
        self.F_GLayout[self.GCount-1].addWidget(self.TextBox_CodigoDeConfirmacao, 11, 0, 1, 2, Qt.AlignCenter | Qt.AlignTop)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_ResendEmail, 13, 0, 2, 1, Qt.AlignHCenter | Qt.AlignRight)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_ConfirmacaoDeEmail, 13, 1, 2, 1, Qt.AlignHCenter | Qt.AlignLeft)
        self.TextBox_CodigoDeConfirmacao.setFocus()

    def CreatePage45(self):
        self.Background_1 = QLabel()
        Background_1 = 'rgb(0, 0, 0)'
        self.Background_1.setStyleSheet("background-color: "+Background_1)

        self.Label_Titulo = QLabel('Inserção das cotações manualmente')
        self.Label_Titulo.setStyleSheet('color: white')
        self.Label_Titulo.setFont(self.font30)
        self.Label_Titulo.setAlignment(Qt.AlignCenter)

        self.Button_Pronto = QPushButton('Pronto')
        self.Button_Pronto.setFont(self.font22)
        self.Button_Pronto.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
        self.Button_Pronto.setIcon(QIcon("./images/log_in.png"))
        self.Button_Pronto.setCursor(QCursor(Qt.PointingHandCursor))
        self.Button_Pronto.setStyleSheet("background-color: rgb(20, 120, 30)")

        self.Button_Voltar = QPushButton('Voltar')
        self.Button_Voltar.setFont(self.font22)
        self.Button_Voltar.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
        self.Button_Voltar.setIcon(QIcon("./images/voltar.png"))
        self.Button_Voltar.setCursor(QCursor(Qt.PointingHandCursor))
        self.Button_Voltar.setStyleSheet("background-color: rgb(120, 20, 30)")
        self.Button_Voltar.pressed.connect(lambda: self.CreatePage(self.PageID))

        self.F_GLayout[self.GCount-1].addWidget(self.Background_1, 0, 0, 6, 4)
        self.F_GLayout[self.GCount-1].addWidget(self.Label_Titulo, 0, 0, 1, 4, Qt.AlignCenter)
        self.F_GLayout[self.GCount-1].addWidget(self.Button_Pronto, 2, 3, 2, 1, Qt.AlignCenter)

        Ativos, Datas, Valores = self.DBManager.GetCotacoesManuais()
        if len(Ativos)>0:
            self.Button_Pronto.pressed.connect(lambda: self.OnButtonPressed('Atualizar Tabela de Cotações'))
            self.F_GLayout[self.GCount-1].addWidget(self.Button_Voltar, 4, 3, 2, 1, Qt.AlignCenter)
            HHeader = ['Ativo',
                       'Última atualização',
                       'Cotação em '+self.DBManager.GetUserCoinCurrency()]
            self.Table_CotaçõesManuaisHeader = QTableWidget()
            self.Table_CotaçõesManuaisHeader.setVisible(False)
            self.Table_CotaçõesManuaisHeader.setShowGrid(False)
            self.Table_CotaçõesManuaisHeader.setFont(self.font20)
            self.Table_CotaçõesManuaisHeader.setStyleSheet('''QTableWidget {background-color: rgb(0, 0, 0);
                                                                            color: white;
                                                                            border: 1px solid rgba(0, 0, 0, 0);}
                                                              QTableView {border-top: 2px solid white;
                                                                          border-right: 2px dashed white;
                                                                          border-left: 2px dashed white;}''') # gridline-color: rgba(255, 255, 255, 0)
            self.Table_CotaçõesManuaisHeader.setRowCount(1)
            self.Table_CotaçõesManuaisHeader.setColumnCount(3)
            self.Table_CotaçõesManuaisHeader.verticalHeader().hide()
            self.Table_CotaçõesManuaisHeader.horizontalHeader().hide()

            celula0 = QTableWidgetItem(str(HHeader[0]))
            celula0.setFlags(Qt.ItemFlags(Qt.ItemIsSelectable))
            celula0.setTextAlignment(Qt.AlignHCenter | Qt.AlignBottom)
            self.Table_CotaçõesManuaisHeader.setItem(0, 0, celula0)
            celula1 = QTableWidgetItem(str(HHeader[1]))
            celula1.setFlags(Qt.ItemFlags(Qt.ItemIsSelectable))
            celula1.setTextAlignment(Qt.AlignHCenter | Qt.AlignBottom)
            self.Table_CotaçõesManuaisHeader.setItem(0, 1, celula1)
            celula2 = QTableWidgetItem(str(HHeader[2]))
            celula2.setFlags(Qt.ItemFlags(Qt.ItemIsSelectable))
            celula2.setTextAlignment(Qt.AlignHCenter | Qt.AlignBottom)
            self.Table_CotaçõesManuaisHeader.setItem(0, 2, celula2)

            self.Table_CotaçõesManuaisHeader.setSizeAdjustPolicy(QAbstractScrollArea.AdjustToContents)
            self.Table_CotaçõesManuaisHeader.setFixedHeight(int(self.frameGeometry().height()*1/20))
            self.Table_CotaçõesManuaisHeader.resizeColumnsToContents()

            self.Table_CotaçõesManuais = QTableWidget()
            self.Table_CotaçõesManuais.setVisible(False)
            self.Table_CotaçõesManuais.setShowGrid(False)
            self.Table_CotaçõesManuais.setFont(self.font16)
            self.Table_CotaçõesManuais.setStyleSheet('''QTableWidget {background-color: rgb(0, 0, 0);
                                                                      color: white;
                                                                      border: 1px solid rgba(0, 0, 0, 0);}
                                                        QTableView {border-bottom: 2px dashed white;
                                                                    border-right: 1px solid white;
                                                                    border-left: 1px solid white;}
                                                        QTableView::item {border-bottom: 1px dashed white;}''') # gridline-color: rgba(255, 255, 255, 0)
            self.Table_CotaçõesManuais.setRowCount(len(Datas))
            self.Table_CotaçõesManuais.setColumnCount(3)
            self.Table_CotaçõesManuais.verticalHeader().hide()
            self.Table_CotaçõesManuais.horizontalHeader().hide()

            for n, item in enumerate(Ativos):
                it = QTableWidgetItem(str(item))
                it.setFlags(Qt.ItemFlags(Qt.ItemIsSelectable))
                it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.Table_CotaçõesManuais.setItem(n, 0, it)

            for n, item in enumerate(Datas):
                it = QTableWidgetItem(str(item))
                it.setFlags(Qt.ItemFlags(Qt.ItemIsSelectable))
                it.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
                item = datetime.strptime(item, "%Y-%m-%d %H:%M:%S")
                if datetime.now() - item <= timedelta(days=2): it.setForeground(QBrush(QColor('green')))
                elif datetime.now() - item <= timedelta(days=5): it.setForeground(QBrush(QColor('orange')))
                else: it.setForeground(QBrush(QColor('red')))
                self.Table_CotaçõesManuais.setItem(n, 1, it)

            for n, item in enumerate(Valores):
                it = QTableWidgetItem(str(item))
                it.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                if item <= 0: it.setForeground(QBrush(QColor('red')))
                self.Table_CotaçõesManuais.setItem(n, 2, it)

            self.data_45_1 = []
            for i, item in enumerate(Datas):
                self.data_45_1.append((Ativos[i], Datas[i], Valores[i]))

            self.Table_CotaçõesManuais.setSizeAdjustPolicy(QAbstractScrollArea.AdjustToContents)
            self.Table_CotaçõesManuais.setFixedHeight(int(self.frameGeometry().height()*6/10))
            self.Table_CotaçõesManuais.resizeColumnsToContents()

            SizeCol0, SizeCol1, SizeCol2 = self.GetSizeOfTableColumns(self.data_45_1, HHeader)

            self.SetTableWidth_45_1(SizeCol0, SizeCol1, SizeCol2)

            self.F_GLayout[self.GCount-1].addWidget(self.Table_CotaçõesManuaisHeader, 1, 0, 1, 3, Qt.AlignHCenter | Qt.AlignBottom)
            self.F_GLayout[self.GCount-1].addWidget(self.Table_CotaçõesManuais, 2, 0, 3, 3, Qt.AlignHCenter | Qt.AlignTop)
        else:
            self.Label_Msg1 = QLabel('Legal!\nEstá tudo no automático pra você!\nNão há nada para fazer aqui :)')
            self.Label_Msg1.setStyleSheet('color: white')
            self.Label_Msg1.setFont(self.font24)
            self.Label_Msg1.setAlignment(Qt.AlignCenter)
            self.F_GLayout[self.GCount-1].addWidget(self.Label_Msg1, 1, 0, 2, 3, Qt.AlignCenter)

            self.Button_Pronto.pressed.connect(lambda: self.CreatePage(self.PageID))

            self.Button_Joinha = QPushButton()
            self.Button_Joinha.setFixedSize(300,300)
            self.Button_Joinha.setIconSize(QSize(295, 295))
            self.Button_Joinha.setStyleSheet('background-color: black; color: white; border: 1px solid black')
            self.Button_Joinha.setIcon(QIcon("./images/Joinha.png"))
            self.F_GLayout[self.GCount-1].addWidget(self.Button_Joinha, 3, 0, 2, 3, Qt.AlignHCenter | Qt.AlignTop)

        # self.DBManager.AtualizarCotacoes()

    def CreatePageChooseCurrency(self):
        self.Label_Titulo = QLabel('Escolha a sua moeda corrente')
        self.Label_Titulo.setStyleSheet('color: white')
        self.Label_Titulo.setFont(self.font30)
        self.Label_Titulo.setAlignment(Qt.AlignCenter)

        self.F_GLayout[self.GCount-1].addWidget(self.Label_Titulo, 0, 0, 1, 1, Qt.AlignCenter)

        self.ComboBox_Currency = QComboBox()
        self.ComboBox_Currency.setFont(self.font16)
        self.ComboBox_Currency.setCursor(QCursor(Qt.PointingHandCursor))
        self.ComboBox_Currency.setStyleSheet("background-color: rgb(10, 10, 10) ; color: rgb(0, 255, 0);")
        # self.ComboBox_Currency.activated[str].connect(lambda pct: self.AtualizarPagina())
        for idx, coin in enumerate(self.DBManager.GetMoedasFiat()):
            self.ComboBox_Currency.addItem(coin)

        self.F_GLayout[self.GCount-1].addWidget(self.ComboBox_Currency, 1, 0, 4, 1, Qt.AlignCenter)

    def CreatePageInfo1(self): MessageBox_Msg1 = QMessageBox.about(self,'Informação','    Quando um ativo tem o modo manual ativado,\nserá solicitada a sua cotação durante procedimentos\n           que atualizam a Tabela de Cotações.')

    def CreatePageInfo2(self): MessageBox_Msg1 = QMessageBox.about(self,'Informação','O que inserir em Tipo?\n\nAqui se indica se foi realizada uma operação de compra ou de venda.')

    def CreatePageInfo3(self): MessageBox_Msg1 = QMessageBox.about(self,'Informação','O que inserir em Data?\n\nAqui se indica quando a operação foi realizada.')

    def CreatePageInfo4(self): MessageBox_Msg1 = QMessageBox.about(self,'Informação','O que inserir em Qqt (Quantidade)?\n\nAqui se indica quantos papeis do ativo foi comprado ou vendido nessa operação.')

    def CreatePageInfo5(self): MessageBox_Msg1 = QMessageBox.about(self,'Informação','O que inserir em Preço?\n\nAqui se indica o preço unitário do ativo que foi comprada ou vendida nessa operação.')

    def CreatePageInfo6(self): MessageBox_Msg1 = QMessageBox.about(self,'Informação','O que inserir em Corretagem?\n\nAqui se indica, em valores brutos, quanto a corretora cobrou para realizar essa operação.')

    def CreatePageInfo7(self): MessageBox_Msg1 = QMessageBox.about(self,'Informação','O que inserir em Taxa B3?\n\nAqui se indica em valores brutos ou em porcentagem (com o simbolo "%") quanto de taxa que foi cobrada pela B3 para realizar essa operação.')

    def CreatePageInfo8(self): MessageBox_Msg1 = QMessageBox.about(self,'Informação','O que inserir em Obervação?\n\nAqui se pode deixar qualquer comentário sobre essa operação. Ela não será levada em consideração em cálculo algum.')

    def CreatePageInfo9(self): MessageBox_Msg1 = QMessageBox.about(self,'Informação','O que inserir em Conversão?\n\nAqui se indica a cotação da segunda moeda no par com relação á sua moeda corrente.\nExemplo: Se o par é BTCUSDT e a sua moeda corrente é BRL, então deve-se indicar o preço de USDTBRL no momento da operação.\nSe o par é BTCBRL e a sua moeda corrente é BRL, então deve-se inserir 1.')

    def CreatePageInfo10(self): MessageBox_Msg1 = QMessageBox.about(self,'Informação','O que inserir em Moeda da Taxa?\n\nAqui se indica qual moeda foi usada para pagar a taxa. Se não houve taxa, pode deixar vazio.')

    def CreatePageInfo11(self): MessageBox_Msg1 = QMessageBox.about(self,'Informação','O que inserir em Taxa?\n\nAqui se indica o valor cobrado na taxa, sem converter moedas. O valor será subtraído do seu estoque da Moeda da Taxa.')

    def CreatePageInfo12(self): MessageBox_Msg1 = QMessageBox.about(self,'Informação','O que inserir em Preço?\n\nAqui se indica o preço unitário do ativo no momento da operação.\nA referência de preço é a mesma do par em questão.\nPor exemplo: Se o par é BTCUSDT e a sua moeda corrente BRL, indique o preço do BTC em USDT. Apenas ignore a sua moeda corrente neste momento.')

    def CreatePageInfo13(self): MessageBox_Msg1 = QMessageBox.about(self,'Informação','O que inserir em Qqt (Quantidade)?\n\nAqui se indica quantas moedas (ou uma fração dela) foi operada.')

    def CreatePageInfo14(self): MessageBox_Msg1 = QMessageBox.about(self,'Informação','O que inserir em Tipo?\n\nAqui se indica o que fez com a moeda em questão.\nSe COMPROU, se VENDEU.\nTambém é possível declarar se tal quantidade foi colocada ou retirada de bloqueio (STAKE ou UNSTAKE), se foi segurada ou liberada (HOLD ou UNHOLD), apenas para sua organização. É NECESSÁRIO declarar todos os DEPÓSITOS e SAQUES, assim como ganhos (DROP) e perdas (BURN).\nBurn pode ser multas, roubos, furtos, confisco etc.')

    def CreatePageInfo15(self): MessageBox_Msg1 = QMessageBox.about(self,'Informação','O que inserir em Par?\n\nAqui se indica qual a moeda ou moedas que foram operadas. ATENÇÃO na importãncia da ordem. Veja o seguinte exemplo:\nBTCUSDT COMPRA -> Significa diminuir o estoque de USDT e aumentar o estoque de BTC.\nEm operações de DEPÓSITO, SAQUE, STAKE, UNSTAKE, HOLD, UNHOLD, DROP e BURN admite-se somente uma moeda no par.')

    def CreatePageInfo16(self): MessageBox_Msg1 = QMessageBox.about(self,'Informação','O que é nota de equilíbrio?\n\nNota de equilíbrio é um indicador que avalia a variedade da sua carteira.\nEla leva em consideração o volume nos diferentes\ntipos de ativos, sub-tipos e setores.')

    def CreatePageInfo17(self): MessageBox_Msg1 = QMessageBox.about(self,'Informação','O que é Gasto Mensal?\n\nGasto Mensal é aproximadamente a sua despeza mensal.\nCom esse valor, calcular-se-á o valor\nque deve ser investido ou retirado de investimentos\npara manter uma reserva segura.')

    def CreatePageInfo18(self): MessageBox_Msg1 = QMessageBox.about(self,'Informação','O que é Meses de Reserva?\n\nMeses de Reserva é o tempo que querer que\na sua reserva cubra seus Gastos Mensais em\ncaso de instabilidade econômica pessoal.')

    def CreatePageInfo19(self): MessageBox_Msg1 = QMessageBox.about(self,'Informação',"O que é 'Aplicar'?\n\nÉ o valor calculado para se aplicar ou retirar de investimento.\nA conta é:\n(Patrimônio - MesesDeReserva * GastoMensal) - MontanteNasCorretoras")

    def CreatePageInfo20(self): MessageBox_Msg1 = QMessageBox.about(self,'Informação','O que é Montante Aplicado?\n\nÉ a diferença entre a soma de depósitos e a soma de saques\nefetuados em cada corretora.')

    def CreatePageInfo21(self): MessageBox_Msg1 = QMessageBox.about(self,'Informação','O que é Montante em Aplicação?\n\nÉ a diferença entre o patrimônio nas\ncorretoras e as suas contas-correntes.\nDesconsidera-se os bancos.')

    def CreatePageInfo22(self): MessageBox_Msg1 = QMessageBox.about(self,'Informação','O que é RNB?\n\nRendimento Nominal Bruto é o rendimento percentual\nsem considerar a tributação.')

    def CreatePageInfo23(self): MessageBox_Msg1 = QMessageBox.about(self,'Informação','O que é Patrimônio?\n\nÉ a soma de todos os valores declarados.\nConsidera-se as cotações adiquiridas online\nou as cotações informadas na tabela manual\n(Atalho: tecla C).\nNão considera-se o desconto de tributação.')

    def CreatePageInfo24(self): MessageBox_Msg1 = QMessageBox.about(self,'Informação',"                                             Ações e Negócios\n\nNada mais rentável do que negócios.\nAplicação convexa, limite de perda, ganhos infinitos.\n\nO que olhar:\nLucratividade, endividamento, preço, crescimento, governança corporativa.\n\nChecklist:\nP/L; PEG RATIO; P/VP; ROE; ROA; ROIC; MG. BRUTA; MG. LUCRO;\nDL/PL; DL/EBIT; LIQUIDEZ CORRENTE; DIVIDA BRUTA/EBIT;\nCAGR EBIT; CAGR LUCRO; SALDO DOS ACIONISTAS.")

    def CreatePageInfo25(self): MessageBox_Msg1 = QMessageBox.about(self,'Informação',"                                                   Real Estate\n\nRendimentos mensais isentos de imposto de renda para pessoa física.\nFluxo de caixa, efeito bola de neve.\n\nO que olhar:\nTipo de fundo imobiliário, localização, vacância, inquilinos, rendimento por cotas.\n\nChecklist:\nP/VP; NÚMERO DE NEGÓCIOS; VACÂNCIA FINANCEIRA; VACÂNCIA\nFÍSICA; CAP RATE; DIVIDEND YIELD; VALOR DO M2/ALUGUEL.")

    def CreatePageInfo26(self): MessageBox_Msg1 = QMessageBox.about(self,'Informação',"                                                   Caixa\n\nCaixa é a parcela da sua carteira utilizada para aproveitar oportunidades.\nSanidade mental em momentos de crise.\n\nO que olhar:\nAplicações conservadoras, de alta liquidez e baixa/média rentabilidade.\n\nCheclist:\nRATING/RISCO; PRÊMIO X BENCHMARK; PRAZO.")

    def CreatePageInfo27(self): MessageBox_Msg1 = QMessageBox.about(self,'Informação',"                                             Ativos Internacionais\n\nNada mais rentável do que negócios em dólar.\nProteção à crises locais.\n\nO que olhar:\nLucratividade, endividamento, preço, crescimento, governança corporativa.\n\nChecklist:\nP/L; PEG RATIO; P/VP; ROE; ROA; ROIC; MG. BRUTA; MG. LUCRO;\nDL/PL; DL/EBIT; LIQUIDEZ CORRENTE; DIVIDA BRUTA/EBIT;\nCAGR EBIT; CAGR LUCRO; SALDO DOS ACIONISTAS.")

    def CreatePageInfo28(self): MessageBox_Msg1 = QMessageBox.about(self,'Informação',"               Criptomoedas\n\nExposição de altíssimo risco.\nMuito volátil, aplicação convexa.\n\nDica de estratégia saudável:\n50 a 60% em BTC\n10% em ETH\n10% em BNB\nO restante divide nas pimentinhas.")

    #%% UpdatePage
    def UpdatePage(self):
        if self.PageID == '18':
            self.idxCorretora = self.ComboBox_Corretoras.currentIndex()
            self.Label_Title2.setText(self.ComboBox_Corretoras.currentText())
            self.TextBox_Msg2.setText(self.DBManager.GetCorretoraCoinCurrency()+' ~'+str(self.DBManager.GetPatrimonioTotalCorretora()))
            self.TextBox_Msg3.setText(self.DBManager.GetCorretoraCoinCurrency()+' '+str(self.DBManager.GetValorEmContaCorrente()))
            self.TextBox_Msg4.setText(self.DBManager.GetDataUltimaOPRegistrada())
            self.TextBox_Msg5.setText(self.DBManager.GetCorretoraCoinCurrency()+' '+str(self.DBManager.GetUltimoDepositoOuSaqueRegistrado()))
        elif self.PageID == '19':
            self.Label_Title2.setText(self.ComboBox_Corretoras.currentText())
            self.TextBox_Msg2.setText(self.DBManager.GetCorretoraCoinCurrency()+' ~'+str(self.DBManager.GetPatrimonioTotalCorretora()))
            self.TextBox_Msg3.setText(self.DBManager.GetCorretoraCoinCurrency()+' '+str(self.DBManager.GetValorEmContaCorrente()))
            self.TextBox_Msg4.setText(self.DBManager.GetDataUltimaOPRegistrada())
            self.TextBox_Msg5.setText(self.DBManager.GetCorretoraCoinCurrency()+' '+str(self.DBManager.GetUltimoDepositoOuSaqueRegistrado()))
        elif self.PageID == 'Bancos':
            self.idxBanco = self.ComboBox_Bancos.currentIndex()
        elif self.PageID == '23':
            self.Label_Msg1.setText('Operações realizadas com '+self.ComboBox_Ativo.currentText())
            self.HMI_Trades.HMI_Trades_Bolsa.CreateTable_23_1() # Table_Operacoes_Realizadas_com_AtivoHeader
            self.HMI_Trades.HMI_Trades_Bolsa.CreateTable_23_2() # Table_Operacoes_Realizadas_com_Ativo (Criada com uma thread)
            self.idxAtivo = self.ComboBox_Ativo.currentIndex()
            self.TextBox_Msg2.setText(str(self.DBManager.GetEstoque()))
        elif self.PageID == '24':
            self.ComboBox_Dia.clear()
            for idx, dia in enumerate(range(1,monthrange(int(self.ComboBox_Ano.currentText()),int(self.ComboBox_Mes.currentText()))[1]+1)):
                self.ComboBox_Dia.addItem(str(dia))
        elif self.PageID == '26':
            self.ComboBox_SubtipoDeAtivo.clear()
            for idx, Subtipo in enumerate(self.DBManager.GetSubtiposDeAtivo()):
                self.ComboBox_SubtipoDeAtivo.addItem(Subtipo)
                try:
                    if Subtipo == self.TextBox_SubtipoDeAtivo.text(): self.idxSubtipoDeAtivo = idx
                except: pass
            if self.idxSubtipoDeAtivo < len(self.DBManager.GetSubtiposDeAtivo()):
                self.ComboBox_SubtipoDeAtivo.setCurrentIndex(self.idxSubtipoDeAtivo)
            self.ComboBox_SetorAtivo.clear()
            for idx, Setor in enumerate(self.DBManager.GetSetoresDeAtivo()):
                self.ComboBox_SetorAtivo.addItem(Setor)
                try:
                    if Setor == self.TextBox_SetorAtivo.text(): self.idxSetorAtivo = idx
                except: pass
            if self.idxSetorAtivo < len(self.DBManager.GetSetoresDeAtivo()):
                self.ComboBox_SetorAtivo.setCurrentIndex(self.idxSetorAtivo)
            if self.ComboBox_TipoDeAtivo.currentText() == "Real Estate":
                self.ComboBox_SetorAtivo.setCurrentIndex(self.ComboBox_SetorAtivo.count()-1)
        elif self.PageID == '27':
            self.ComboBox_SubtipoDeAtivo.clear()
            for idx, Subtipo in enumerate(self.DBManager.GetSubtiposDeAtivo()):
                self.ComboBox_SubtipoDeAtivo.addItem(Subtipo)
                try:
                    if Subtipo == self.TextBox_SubtipoDeAtivo.text(): self.idxSubtipoDeAtivo = idx
                except: pass
            if self.idxSubtipoDeAtivo < len(self.DBManager.GetSubtiposDeAtivo()):
                self.ComboBox_SubtipoDeAtivo.setCurrentIndex(self.idxSubtipoDeAtivo)
            self.ComboBox_SetorAtivo.clear()
            for idx, Setor in enumerate(self.DBManager.GetSetoresDeAtivo()):
                self.ComboBox_SetorAtivo.addItem(Setor)
                try:
                    if Setor == self.TextBox_SetorAtivo.text(): self.idxSetorAtivo = idx
                except: pass
            if self.idxSetorAtivo < len(self.DBManager.GetSetoresDeAtivo()):
                self.ComboBox_SetorAtivo.setCurrentIndex(self.idxSetorAtivo)
            if self.ComboBox_TipoDeAtivo.currentText() == "Real Estate":
                self.ComboBox_SetorAtivo.setCurrentIndex(self.ComboBox_SetorAtivo.count()-1)
        elif self.PageID == '28':
            self.ComboBox_Dia.clear()
            for idx, dia in enumerate(range(1,monthrange(int(self.ComboBox_Ano.currentText()),int(self.ComboBox_Mes.currentText()))[1]+1)):
                self.ComboBox_Dia.addItem(str(dia))
        elif self.PageID == '36a':
            self.ComboBox_Dia.clear()
            for idx, dia in enumerate(range(1,monthrange(int(self.ComboBox_Ano.currentText()),int(self.ComboBox_Mes.currentText()))[1]+1)):
                self.ComboBox_Dia.addItem(str(dia))

    #%% Ferramentas
    def EnableAllButtons(self, condition = True):
        if self.PageID == '28':
            self.Button_Compra.setEnabled(condition)
            self.Button_Venda.setEnabled(condition)
            self.Button_LimparCampos.setEnabled(condition)
            self.Button_RegistrarOperacao.setEnabled(condition)
            self.Button_Info3.setEnabled(condition)
            self.Button_Info2.setEnabled(condition)
            self.Button_Info4.setEnabled(condition)
            self.Button_Info5.setEnabled(condition)
            self.Button_Info6.setEnabled(condition)
            self.Button_Info7.setEnabled(condition)
            self.Button_Info8.setEnabled(condition)
            self.Button_Perfil.setEnabled(condition)
            self.Button_ShowHideValues.setEnabled(condition)
            self.Button_Home.setEnabled(condition)
            self.Button_Return.setEnabled(condition)

    def GetSizeOfTableColumns(self, data, HHeader):
        coluna = []
        resposta = []
        if not len(data) > 0:
            aux = (' ',)
            for i in range(len(HHeader)-1):
                aux += (' ',)
            data.append(aux)
        for j in range(len(data[0])):
            col = []
            for i, linha in enumerate(data):
                col.append(str(linha[j]))
            col.append(HHeader[j])
            coluna.append(col)

        if self.PageIDAux == '45':
            col = len(max(coluna[0], key=len))*14 # Ativo
            resposta.append(col)
            col = len(max(coluna[1], key=len))*16 # Data
            resposta.append(col)
            col = len(max(coluna[2], key=len))*15 # Cotação em CURRENCY
            resposta.append(col)

        elif self.PageID == '11':
            col = len(max(coluna[0], key=len))*9 # Corretora
            resposta.append(col)
            col = len(max(coluna[1], key=len))*8 # Conta-corrente
            resposta.append(col)

        elif self.PageID in ['23', '29', '30']:
            col7 = int(self.frameGeometry().width())
            col = len(max(coluna[0], key=len))*10
            col7 -= col
            resposta.append(col)
            col = len(max(coluna[1], key=len))*15
            col7 -= col
            resposta.append(col)
            col = len(max(coluna[2], key=len))*15
            col7 -= col
            resposta.append(col)
            col = len(max(coluna[3], key=len))*18
            col7 -= col
            resposta.append(col)
            col = len(max(coluna[4], key=len))*12
            col7 -= col
            resposta.append(col)
            col = len(max(coluna[5], key=len))*15
            col7 -= col
            resposta.append(col)
            col = len(max(coluna[6], key=len))*12
            col7 -= col
            resposta.append(col)
            col7 -= 44

            if not col7 >= len(max(coluna[7], key=len))*10:
                self.Table_Operacoes_Realizadas_com_AtivoHeader.setColumnWidth(7, col7)
                col7 = len(max(coluna[7], key=len))*10
            else:
                self.Table_Operacoes_Realizadas_com_AtivoHeader.setColumnWidth(7, col7)

            resposta.append(col7)

        elif self.PageID == '24':
            col = len(max(coluna[0], key=len))*9 # Data
            resposta.append(col)
            col = len(max(coluna[1], key=len))*11 # Valores
            resposta.append(col)

        elif self.PageID in ['40', '43', '44']:
            col8 = int(self.frameGeometry().width())
            col = len(max(coluna[0], key=len))*8
            col8 -= col
            resposta.append(col)
            col = len(max(coluna[1], key=len))*12
            col8 -= col
            resposta.append(col)
            col = len(max(coluna[2], key=len))*13
            col8 -= col
            resposta.append(col)
            col = len(max(coluna[3], key=len))*13
            col8 -= col
            resposta.append(col)
            col = len(max(coluna[4], key=len))*12
            col8 -= col
            resposta.append(col)
            col = len(max(coluna[5], key=len))*12
            col8 -= col
            resposta.append(col)
            col = len(max(coluna[6], key=len))*10
            col8 -= col
            resposta.append(col)
            col = len(max(coluna[7], key=len))*10
            col8 -= col
            resposta.append(col)
            col8 -= 44

            if not col8 >= len(max(coluna[8], key=len))*10:
                self.Table_Operacoes_RealizadasHeader.setColumnWidth(8, col8)
                col8 = len(max(coluna[8], key=len))*10
            else:
                self.Table_Operacoes_RealizadasHeader.setColumnWidth(8, col8)

            resposta.append(col8)

        elif self.PageID == '51':
            if not self.DTeST_51 == "DTeST":
                if int(self.frameGeometry().width()*6/20*1.4) < 510:
                    col = len(max(coluna[0], key=len))*4 # MesAno
                    resposta.append(col)
                    col = len(max(coluna[1], key=len))*7 # Resultado Líquido
                    resposta.append(col)
                    col = len(max(coluna[2], key=len))*8 # Imposto devido
                    resposta.append(col)
                    col = len(max(coluna[3], key=len))*8 # Alíquota
                    resposta.append(col)
                    col = len(max(coluna[4], key=len))*7 # Lucro mínimo taxável
                    resposta.append(col)
                    col = len(max(coluna[5], key=len))*8 # Resultado bruto
                    resposta.append(col)
                    col = len(max(coluna[6], key=len))*8 # Prejuízo acumulado
                    resposta.append(col)
                    col = len(max(coluna[7], key=len))*7 # Resultado final
                    resposta.append(col)
                else:
                    col = len(max(coluna[0], key=len))*5 #
                    resposta.append(col)
                    col = len(max(coluna[1], key=len))*10 #
                    resposta.append(col)
                    col = len(max(coluna[2], key=len))*10 #
                    resposta.append(col)
                    col = len(max(coluna[3], key=len))*11 #
                    resposta.append(col)
                    col = len(max(coluna[4], key=len))*10 #
                    resposta.append(col)
                    col = len(max(coluna[5], key=len))*10 #
                    resposta.append(col)
                    col = len(max(coluna[6], key=len))*10 #
                    resposta.append(col)
                    col = len(max(coluna[7], key=len))*10 #
                    resposta.append(col)
            else:
                col = len(max(coluna[0], key=len))*6 #
                resposta.append(col)
                col = len(max(coluna[1], key=len))*10 #
                resposta.append(col)
                col = len(max(coluna[2], key=len))*10 #
                resposta.append(col)
                col = len(max(coluna[3], key=len))*10 #
                resposta.append(col)
                col = len(max(coluna[4], key=len))*10 #
                resposta.append(col)
                col = len(max(coluna[5], key=len))*10 #
                resposta.append(col)

        return resposta

    def SetTableWidth_45_1(self, SizeCol0, SizeCol1, SizeCol2):
            self.Table_CotaçõesManuais.setVisible(False)
            self.Table_CotaçõesManuaisHeader.setVisible(False)

            self.Table_CotaçõesManuais.adjustSize()
            self.Table_CotaçõesManuais.setColumnWidth(0, SizeCol0)
            self.Table_CotaçõesManuais.setColumnWidth(1, SizeCol1)
            self.Table_CotaçõesManuais.setColumnWidth(2, SizeCol2)

            width = self.Table_CotaçõesManuais.verticalHeader().width()
            width += self.Table_CotaçõesManuais.horizontalHeader().length()
            if self.Table_CotaçõesManuais.horizontalScrollBar().isVisible():
                width += self.Table_CotaçõesManuais.horizontalScrollBar().width()
            width += self.Table_CotaçõesManuais.frameWidth() * 2
            self.Table_CotaçõesManuais.setFixedWidth(width)

            self.Table_CotaçõesManuaisHeader.adjustSize()
            self.Table_CotaçõesManuaisHeader.setColumnWidth(0, SizeCol0)
            self.Table_CotaçõesManuaisHeader.setColumnWidth(1, SizeCol1)
            self.Table_CotaçõesManuaisHeader.setColumnWidth(2, SizeCol2)

            self.Table_CotaçõesManuais.setVisible(True)
            self.Table_CotaçõesManuaisHeader.setVisible(True)

    def EnviarEmailDeConfirmacao(self,Email,NomeDeUsuario,CodigoDeConfirmacaoGerado):
        import smtplib, ssl

        smtp_server = "smtp.gmail.com"
        port = 587  # For starttls

        sender_email = "confirmacaodeemailfinancas@gmail.com"
        password = "SenhaDasMinhasFinancas2021"

        # Create a secure SSL context
        context = ssl.create_default_context()

        # Try to log in to server and send email
        try:
            server = smtplib.SMTP(smtp_server,port)
            server.ehlo() # Can be omitted
            server.starttls(context=context) # Secure the connection
            server.ehlo() # Can be omitted
            server.login(sender_email, password)

            # Send email here
            message = ("Subject: Confirmação de email\n\nOlá "+NomeDeUsuario+"\n\nEste é o seu codigo de verifição de email:\n\n"+CodigoDeConfirmacaoGerado+
                       "\n\n------------------------------------\nEmail Automático do APP Finanças para Confirmação de Email").encode('utf-8').strip()
            receiver_email = Email
            server.sendmail(sender_email, receiver_email, message)
            server.quit()
        except Exception as e:
            # print(e) # Print any error messages
            try:server.quit()
            except:pass

    def ClearPage(self):
        try: # Cabeçalho
            self.Button_Return.deleteLater()
            self.Button_Home.deleteLater()
            self.Button_ShowHideValues.deleteLater()
            self.Label_Patrimonio.deleteLater()
            self.TextBox_Patrimonio.deleteLater()
            self.Label_Rendimento.deleteLater()
            self.TextBox_Rendimento.deleteLater()
            self.Button_Sair.deleteLater()
            self.Button_AttCotacoes.deleteLater()
            self.Button_AttCotacoesManuais.deleteLater()
            self.Button_Perfil.deleteLater()
            self.Button_Currency.deleteLater()
        except:pass
        try:
            self.GLayout.removeItem(self.HBoxLayout_Header)
        except:pass

        if self.PageID == "11" and self.PageIDAux == '':
            try:
                self.HMI_ARCA.AbortPlotings = True
                self.Button_Msg3.deleteLater()
                self.TextBox_MontanteEmAplicacao.deleteLater()
                self.HBoxLayout_Montantes.removeItem(self.HBoxLayout_Montantes1)
                self.Button_Msg4.deleteLater()
                self.TextBox_MontanteAplicado.deleteLater()
                self.HBoxLayout_Montantes.removeItem(self.HBoxLayout_Montantes2)
                self.F_GLayout[0].removeItem(self.HBoxLayout_Montantes)

                self.Label_Msg9.deleteLater()
                self.Label_Msg8.deleteLater()
                self.HBoxLayout_Reserva.removeItem(self.HBoxLayout_Reserva1)
                self.Button_Msg5.deleteLater()
                self.TextBox_GastoMensal.deleteLater()
                self.HBoxLayout_Reserva.removeItem(self.HBoxLayout_Reserva2)
                self.Button_Msg6.deleteLater()
                self.TextBox_MesesDeReserva.deleteLater()
                self.HBoxLayout_Reserva.removeItem(self.HBoxLayout_Reserva3)
                self.Button_Msg7.deleteLater()
                self.TextBox_Aplicar.deleteLater()
                self.HBoxLayout_Reserva.removeItem(self.HBoxLayout_Reserva4)
                self.F_GLayout[0].removeItem(self.HBoxLayout_Reserva)

                self.Label_Subtitulo1_.deleteLater()
                self.GraphWidget_AcoesENegocios.deleteLater()
                self.Label_Subtitulo2_.deleteLater()
                self.GraphWidget_RealEstate.deleteLater()
                self.Label_Subtitulo3_.deleteLater()
                self.GraphWidget_Caixa.deleteLater()
                self.Label_Subtitulo4_.deleteLater()
                self.GraphWidget_AtivosInternacionais.deleteLater()
                self.Label_Subtitulo5_.deleteLater()
                self.GraphWidget_Especifica.deleteLater()
                self.HBoxLayout_Graficos_Subtipos.removeItem(self.VBoxLayout_Graficos_Subtipos_AcoesENegocios)
                self.HBoxLayout_Graficos_Subtipos.removeItem(self.VBoxLayout_Graficos_Subtipos_RealEstate)
                self.HBoxLayout_Graficos_Subtipos.removeItem(self.VBoxLayout_Graficos_Subtipos_Caixa)
                self.HBoxLayout_Graficos_Subtipos.removeItem(self.VBoxLayout_Graficos_Subtipos_AtivosInternacionais)
                self.HBoxLayout_Graficos_Subtipos.removeItem(self.VBoxLayout_Graficos_Subtipos_Especifica)
                self.F_GLayout[0].removeItem(self.HBoxLayout_Graficos_Subtipos)

                self.Label_Subtitulo1.deleteLater()
                self.Label_Subtitulo2.deleteLater()
                self.Label_Subtitulo3.deleteLater()
                self.Label_Subtitulo4.deleteLater()
                self.Label_Subtitulo5.deleteLater()
                self.Button_AcoesENegociosInvisivel.deleteLater()
                self.Button_RealEstateInvisivel.deleteLater()
                self.Button_CaixaInvisivel.deleteLater()
                self.Button_AcoesInternacionaisInvisivel.deleteLater()
                self.Button_EspecificaInvisivel.deleteLater()
                self.HBoxLayout_Graficos_Subtipos_Invisiveis.removeItem(self.VBoxLayout_Graficos_Subtipos_AcoesENegociosInvisiveis)
                self.HBoxLayout_Graficos_Subtipos_Invisiveis.removeItem(self.VBoxLayout_Graficos_Subtipos_RealEstateInvisiveis)
                self.HBoxLayout_Graficos_Subtipos_Invisiveis.removeItem(self.VBoxLayout_Graficos_Subtipos_CaixaInvisiveis)
                self.HBoxLayout_Graficos_Subtipos_Invisiveis.removeItem(self.VBoxLayout_Graficos_Subtipos_AtivosInternacionaisInvisiveis)
                self.HBoxLayout_Graficos_Subtipos_Invisiveis.removeItem(self.VBoxLayout_Graficos_Subtipos_EspecificaInvisiveis)
                self.F_GLayout[0].removeItem(self.HBoxLayout_Graficos_Subtipos_Invisiveis)
            except Exception as e:print(e) # Debug
        elif self.PageID == "24" and self.PageIDAux == '':
            try:
                self.Button_Calendario.deleteLater()
                self.Button_Relogio.deleteLater()
                self.F_GLayout[0].removeItem(self.HBoxLayout_Data)
            except:pass
        elif self.PageID in ['28', '29', '36a', '36b'] and self.PageIDAux == '':
            try:
                self.Button_Calendario.deleteLater()
                self.Button_Relogio.deleteLater()
                self.Button_Info3.deleteLater()
                self.F_GLayout[0].removeItem(self.HBoxLayout_Data)
                self.Button_Compra.deleteLater()
                self.Button_Venda.deleteLater()
                self.Button_Info2.deleteLater()
                self.F_GLayout[0].removeItem(self.HBoxLayout_Tipo)
                self.TextBox_Qqt.deleteLater()
                self.Button_Info4.deleteLater()
                self.F_GLayout[0].removeItem(self.HBoxLayout_Qqt)
                self.TextBox_Preco.deleteLater()
                self.Button_Info5.deleteLater()
                self.F_GLayout[0].removeItem(self.HBoxLayout_Preco)
                self.TextBox_Corretagem.deleteLater()
                self.Button_Info6.deleteLater()
                self.F_GLayout[0].removeItem(self.HBoxLayout_Corretagem)
                self.TextBox_TaxaB3.deleteLater()
                self.Button_Info7.deleteLater()
                self.F_GLayout[0].removeItem(self.HBoxLayout_TaxaB3)
                self.TextBox_Obs.deleteLater()
                self.Button_Info8.deleteLater()
                self.F_GLayout[0].removeItem(self.HBoxLayout_Obs)
            except: pass
        elif self.PageID in ['41', '43'] and self.PageIDAux == '':
            try:
                self.Button_Calendario.deleteLater()
                self.Button_Relogio.deleteLater()
                self.Button_Info3.deleteLater()
                self.F_GLayout[0].removeItem(self.HBoxLayout_Data)
                self.Button_Compra.deleteLater()
                self.Button_Venda.deleteLater()
                self.Button_Deposito.deleteLater()
                self.Button_Saque.deleteLater()
                self.Button_Drop.deleteLater()
                self.Button_Burn.deleteLater()
                self.Button_Stake.deleteLater()
                self.Button_Unstake.deleteLater()
                self.Button_Hold.deleteLater()
                self.Button_Unhold.deleteLater()
                self.Button_Info14.deleteLater()
                self.F_GLayout[0].removeItem(self.HBoxLayout_Tipo)
                self.TextBox_Par_Esquerda.deleteLater()
                self.Label_Msg12.deleteLater()
                self.TextBox_Par_Direita.deleteLater()
                self.Button_Info15.deleteLater()
                self.F_GLayout[0].removeItem(self.HBoxLayout_Par)
                self.TextBox_Qqt.deleteLater()
                self.Button_Info13.deleteLater()
                self.F_GLayout[0].removeItem(self.HBoxLayout_Qqt)
                self.TextBox_Preco.deleteLater()
                self.Button_Info12.deleteLater()
                self.F_GLayout[0].removeItem(self.HBoxLayout_Preco)
                self.TextBox_Taxa.deleteLater()
                self.Button_Info11.deleteLater()
                self.F_GLayout[0].removeItem(self.HBoxLayout_Taxa)
                self.TextBox_MoedaDaTaxa.deleteLater()
                self.Button_Info10.deleteLater()
                self.F_GLayout[0].removeItem(self.HBoxLayout_MoedaDaTaxa)
                self.TextBox_Conversao.deleteLater()
                self.Button_Info9.deleteLater()
                self.F_GLayout[0].removeItem(self.HBoxLayout_Conversao)
                self.TextBox_Obs.deleteLater()
                self.Button_Info8.deleteLater()
                self.F_GLayout[0].removeItem(self.HBoxLayout_Obs)
            except: pass

        for i in range(self.GCount):
            for j in reversed(range(self.F_GLayout[i].count())): # Clear all Widgets from the sub Layout
                self.F_GLayout[i].itemAt(j).widget().setParent(None)
            self.GLayout.removeItem(self.F_GLayout[i]) # Delete each sub Layout
            self.GCount -= 1

        try:
            for i in reversed(range(self.GLayout.count())): # Clear all Widgets from the main Layout
                self.GLayout.itemAt(i).widget().setParent(None)
        except:pass

    def InsertGridLayout(self, row=1, column=0, panRow=1, panColumn=20): # Adiciona um sub-GridLayout no Grid Principal
        self.F_GLayout.append(QGridLayout())
        self.GCount += 1
        self.GLayout.addLayout(self.F_GLayout[self.GCount-1], row, column, panRow, panColumn)

    #%% Header
    def AtualizarCotacoes(self, whatever): self.DBManager.AtualizarCotacoes()

    def PutHeader(self):
        self.Background_Header = QLabel()
        Background_Header = 'rgb(20, 20, 40)'
        self.Background_Header.setStyleSheet("background-color: "+Background_Header)
        self.Background_Header.setFixedHeight(50)

        self.Button_Return = QPushButton()
        self.Button_Return.setFixedSize(40,40)
        self.Button_Return.setIconSize(QSize(35, 35))
        self.Button_Return.setCursor(QCursor(Qt.PointingHandCursor))
        self.Button_Return.setStyleSheet("background-color: "+Background_Header+"; border: none; outline: none;")
        self.Button_Return.setIcon(QIcon("./images/return_icon.png"))

        self.Button_Home = QPushButton()
        self.Button_Home.pressed.connect(lambda: self.CreatePage('1'))
        self.Button_Home.setFixedSize(40,40)
        self.Button_Home.setIconSize(QSize(35, 35))
        self.Button_Home.setCursor(QCursor(Qt.PointingHandCursor))
        self.Button_Home.setStyleSheet("background-color: "+Background_Header+"; border: none; outline: none;")
        self.Button_Home.setIcon(QIcon("./images/home_icon.png"))

        self.Button_ShowHideValues = QPushButton()
        self.Button_ShowHideValues.pressed.connect(lambda: self.OnButtonPressed('ShowHideValues'))
        self.Button_ShowHideValues.setFixedSize(40,40)
        self.Button_ShowHideValues.setIconSize(QSize(35, 35))
        self.Button_ShowHideValues.setCursor(QCursor(Qt.PointingHandCursor))
        self.Button_ShowHideValues.setStyleSheet("background-color: "+Background_Header+"; border: none; outline: none;")
        if not self.ShowValues: self.Button_ShowHideValues.setIcon(QIcon("./images/iconClosedEye.png"))
        else: self.Button_ShowHideValues.setIcon(QIcon("./images/iconOpenEye.png"))

        self.Label_Patrimonio = QPushButton('Patrimônio:')
        self.Label_Patrimonio.pressed.connect(lambda: self.CreatePage('Info23'))
        self.Label_Patrimonio.setStyleSheet('background-color: rgba(0, 0, 0, 0); color: white; border: 1px solid rgba(0, 0, 0, 0); border: none; outline: none; alignment: right')
        self.Label_Patrimonio.setFont(self.font14)

        self.TextBox_Patrimonio = QLineEdit()
        self.TextBox_Patrimonio.setFont(self.font14)
        self.TextBox_Patrimonio.setEnabled(False)
        self.TextBox_Patrimonio.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        if isinstance(self.Patrimonio,float):
            self.TextBox_Patrimonio.setText(self.DBManager.GetUserCoinCurrency()+str(self.Patrimonio))
            self.TextBox_Patrimonio.setStyleSheet('QLineEdit {background-color: '+Background_Header+'; color: green; border: 1px solid '+Background_Header+'}')
        else:
            self.TextBox_Patrimonio.setText(str(self.Patrimonio))
            self.TextBox_Patrimonio.setStyleSheet('QLineEdit {background-color: '+Background_Header+'; color: orange; border: 1px solid '+Background_Header+'}')
        if self.ShowValues: self.TextBox_Patrimonio.setEchoMode(QLineEdit.Normal)
        else: self.TextBox_Patrimonio.setEchoMode(QLineEdit.Password)

        self.Label_Rendimento = QPushButton('RNB%:')
        self.Label_Rendimento.pressed.connect(lambda: self.CreatePage('Info22'))
        self.Label_Rendimento.setStyleSheet('background-color: rgba(0, 0, 0, 0); color: white; border: 1px solid rgba(0, 0, 0, 0); border: none; outline: none; alignment: right')
        self.Label_Rendimento.setFont(self.font14)

        self.TextBox_Rendimento = QLineEdit()
        self.TextBox_Rendimento.setFont(self.font14)
        self.TextBox_Rendimento.setEnabled(False)
        self.TextBox_Rendimento.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        if isinstance(self.Rendimento, float):
            self.TextBox_Rendimento.setText(str(self.Rendimento)+'%')
            if self.Rendimento >= 0: self.TextBox_Rendimento.setStyleSheet('QLineEdit {background-color: '+Background_Header+'; color: green; border: 1px solid '+Background_Header+'}')
            else: self.TextBox_Rendimento.setStyleSheet('QLineEdit {background-color: '+Background_Header+'; color: red; border: 1px solid '+Background_Header+'}')
        else:
            self.TextBox_Rendimento.setText(str(self.Rendimento))
            self.TextBox_Rendimento.setStyleSheet('QLineEdit {background-color: '+Background_Header+'; color: orange; border: 1px solid '+Background_Header+'}')

        self.Button_Sair = QPushButton()
        self.Button_Sair.pressed.connect(lambda: self.CreatePage('0'))
        self.Button_Sair.setFixedSize(40,40)
        self.Button_Sair.setIconSize(QSize(35, 35))
        self.Button_Sair.setCursor(QCursor(Qt.PointingHandCursor))
        self.Button_Sair.setStyleSheet("background-color: "+Background_Header+"; border: none; outline: none;")
        self.Button_Sair.setIcon(QIcon("./images/exit_icon.png"))

        self.Button_AttCotacoesManuais = QPushButton()
        self.Button_AttCotacoesManuais.pressed.connect(lambda: self.CreatePage('45'))
        self.Button_AttCotacoesManuais.setFixedSize(40,40)
        self.Button_AttCotacoesManuais.setIconSize(QSize(35, 35))
        self.Button_AttCotacoesManuais.setCursor(QCursor(Qt.PointingHandCursor))
        self.Button_AttCotacoesManuais.setStyleSheet("background-color: "+Background_Header+"; border: none; outline: none;")
        self.Button_AttCotacoesManuais.setIcon(QIcon("./images/iconeAdjustDB.png"))

        self.Button_AttCotacoes = QLabel()
        self.Button_AttCotacoes.mousePressEvent = self.AtualizarCotacoes
        self.Movie_AttCotacoes = QMovie("./images/Updating.gif")
        self.Movie_AttCotacoes.setScaledSize(QSize(50, 50))
        self.Button_AttCotacoes.setFixedSize(50,50)
        self.Button_AttCotacoes.setMovie(self.Movie_AttCotacoes)
        self.Movie_AttCotacoes.jumpToFrame(0)
        if self.AtualizarCotacoes_running: self.Movie_AttCotacoes.start()
        else: self.Button_AttCotacoes.setCursor(QCursor(Qt.PointingHandCursor))

        self.Button_Perfil = QPushButton()
        self.Button_Perfil.pressed.connect(lambda: self.CreatePage('9'))
        self.Button_Perfil.setFixedSize(40,40)
        self.Button_Perfil.setIconSize(QSize(35, 35))
        self.Button_Perfil.setCursor(QCursor(Qt.PointingHandCursor))
        self.Button_Perfil.setStyleSheet("background-color: "+Background_Header+"; border: none; outline: none;")
        try:
            if os.path.exists(os.getcwd()+"\\images\\ProfilePhoto_"+self.User+".jpg"):
                self.Button_Perfil.setIcon(QIcon("./images/ProfilePhoto_"+self.User+".jpg"))
            else:
                self.Button_Perfil.setIcon(QIcon("./images/ProfilePhoto.jpg"))
        except: self.Button_Perfil.setIcon(QIcon("./images/ProfilePhoto.jpg"))

        self.Button_Currency = QPushButton('Moeda corrente: '+self.DBManager.GetUserCoinCurrency())
        self.Button_Currency.pressed.connect(lambda: self.CreatePage('ChooseCurrency'))
        self.Button_Currency.setFixedHeight(40)
        self.Button_Currency.setIconSize(QSize(35, 35))
        self.Button_Currency.setCursor(QCursor(Qt.PointingHandCursor))
        self.Button_Currency.setStyleSheet("background-color: "+Background_Header+" ; color: rgb(255, 255, 255); border: none; outline: none;")
        self.Button_Currency.setIcon(QIcon("./images/iconeAdjustDB.jpg"))

        self.HBoxLayout_Header = QHBoxLayout()

        if self.PageID in ['0']: pass # Essa não tem cabeçalho

        elif self.PageIDAux in ['9']:
            if self.PageID in ['28', '29', '34a', '34b', '35a', '35b', '36a', '36b']: self.Button_Return.pressed.connect(lambda: self.CreatePage('23'))
            elif self.PageID in ['41', '43']: self.Button_Return.pressed.connect(lambda: self.CreatePage('40'))
            else: self.Button_Return.pressed.connect(lambda: self.CreatePage(self.PageID))
            self.GLayout.addWidget(self.Background_Header, 0, 0, 1, 20)
            self.HMI_ARCA.Flag_RecalcularGraficos = True

            self.HBoxLayout_Header.addWidget(self.Button_Return, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_AttCotacoes, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_Home, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_ShowHideValues, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Label_Patrimonio)
            self.HBoxLayout_Header.addWidget(self.TextBox_Patrimonio, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Label_Rendimento)
            self.HBoxLayout_Header.addWidget(self.TextBox_Rendimento, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_AttCotacoesManuais, Qt.AlignRight)
            self.HBoxLayout_Header.addWidget(self.Button_Currency)
            self.GLayout.addLayout(self.HBoxLayout_Header, 0, 0, 1, 20)

        elif self.PageIDAux in ['45']:
            if self.PageID in ['28', '29', '34a', '34b', '35a', '35b', '36a', '36b']: self.Button_Return.pressed.connect(lambda: self.CreatePage('23'))
            elif self.PageID in ['41', '43']: self.Button_Return.pressed.connect(lambda: self.CreatePage('40'))
            else: self.Button_Return.pressed.connect(lambda: self.CreatePage(self.PageID))
            self.GLayout.addWidget(self.Background_Header, 0, 0, 1, 20)
            self.HMI_ARCA.Flag_RecalcularGraficos = True

            self.HBoxLayout_Header.addWidget(self.Button_Return, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_AttCotacoes, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_Home, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_ShowHideValues, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Label_Patrimonio)
            self.HBoxLayout_Header.addWidget(self.TextBox_Patrimonio, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Label_Rendimento)
            self.HBoxLayout_Header.addWidget(self.TextBox_Rendimento, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_AttCotacoesManuais, Qt.AlignRight)
            self.HBoxLayout_Header.addWidget(self.Button_Perfil, Qt.AlignRight)
            self.GLayout.addLayout(self.HBoxLayout_Header, 0, 0, 1, 20)

        elif self.PageIDAux in ['14', '15', '17a', '17b', 'ChooseCurrency']:
            self.Button_Return.pressed.connect(lambda: self.CreatePage('9'))
            self.GLayout.addWidget(self.Background_Header, 0, 0, 1, 20)

            self.HBoxLayout_Header.addWidget(self.Button_Return, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_AttCotacoes, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_Home, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_ShowHideValues, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Label_Patrimonio)
            self.HBoxLayout_Header.addWidget(self.TextBox_Patrimonio, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Label_Rendimento)
            self.HBoxLayout_Header.addWidget(self.TextBox_Rendimento, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_AttCotacoesManuais, Qt.AlignRight)
            self.HBoxLayout_Header.addWidget(self.Button_Perfil, Qt.AlignRight)
            self.GLayout.addLayout(self.HBoxLayout_Header, 0, 0, 1, 20)

        elif self.PageID in ['2', '3', '4', '7']:
            self.Button_Return.pressed.connect(lambda: self.CreatePage('0'))
            self.GLayout.addWidget(self.Background_Header, 0, 0, 1, 20)

            self.HBoxLayout_Header.addWidget(self.Button_Return, Qt.AlignLeft)
            self.GLayout.addLayout(self.HBoxLayout_Header, 0, 0, 1, 20, Qt.AlignLeft)

        elif self.PageID in ['5']:
            self.Button_Return.pressed.connect(lambda: self.CreatePage('0'))
            self.GLayout.addWidget(self.Background_Header, 0, 0, 1, 20)

            self.HBoxLayout_Header.addWidget(self.Button_Return, Qt.AlignLeft)
            self.GLayout.addLayout(self.HBoxLayout_Header, 0, 0, 1, 20, Qt.AlignLeft)

        elif self.PageID in ['6']:
            self.Button_Return.pressed.connect(lambda: self.CreatePage('2'))
            self.GLayout.addWidget(self.Background_Header, 0, 0, 1, 20)

            self.HBoxLayout_Header.addWidget(self.Button_Return, Qt.AlignLeft)
            self.GLayout.addLayout(self.HBoxLayout_Header, 0, 0, 1, 20, Qt.AlignLeft)

        elif self.PageID in ['8']:
            self.Button_Return.pressed.connect(lambda: self.CreatePage('4'))
            self.GLayout.addWidget(self.Background_Header, 0, 0, 1, 20)

            self.HBoxLayout_Header.addWidget(self.Button_Return, Qt.AlignLeft)
            self.GLayout.addLayout(self.HBoxLayout_Header, 0, 0, 1, 20, Qt.AlignLeft)

        elif self.PageID in ['1']:
            self.GLayout.addWidget(self.Background_Header, 0, 0, 1, 20)

            self.HBoxLayout_Header.addWidget(self.Button_AttCotacoes, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_ShowHideValues, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Label_Patrimonio)
            self.HBoxLayout_Header.addWidget(self.TextBox_Patrimonio, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Label_Rendimento)
            self.HBoxLayout_Header.addWidget(self.TextBox_Rendimento, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_Sair, Qt.AlignRight)
            self.HBoxLayout_Header.addWidget(self.Button_Perfil, Qt.AlignRight)
            self.GLayout.addLayout(self.HBoxLayout_Header, 0, 0, 1, 20)

        elif self.PageID in ['11', '12']:
            self.Button_Return.pressed.connect(lambda: self.CreatePage('1'))
            self.GLayout.addWidget(self.Background_Header, 0, 0, 1, 20)

            self.HBoxLayout_Header.addWidget(self.Button_Return, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_AttCotacoes, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_Home, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_ShowHideValues, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Label_Patrimonio)
            self.HBoxLayout_Header.addWidget(self.TextBox_Patrimonio, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Label_Rendimento)
            self.HBoxLayout_Header.addWidget(self.TextBox_Rendimento, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_AttCotacoesManuais, Qt.AlignRight)
            self.HBoxLayout_Header.addWidget(self.Button_Perfil, Qt.AlignRight)
            self.GLayout.addLayout(self.HBoxLayout_Header, 0, 0, 1, 20)

        elif self.PageID in ['46', '47', '48', '49', '50']:
            self.Button_Return.pressed.connect(lambda: self.CreatePage('11'))
            self.GLayout.addWidget(self.Background_Header, 0, 0, 1, 20)

            self.HBoxLayout_Header.addWidget(self.Button_Return, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_AttCotacoes, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_Home, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_ShowHideValues, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Label_Patrimonio)
            self.HBoxLayout_Header.addWidget(self.TextBox_Patrimonio, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Label_Rendimento)
            self.HBoxLayout_Header.addWidget(self.TextBox_Rendimento, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_AttCotacoesManuais, Qt.AlignRight)
            self.HBoxLayout_Header.addWidget(self.Button_Perfil, Qt.AlignRight)
            self.GLayout.addLayout(self.HBoxLayout_Header, 0, 0, 1, 20)

        elif self.PageID in ['18', '19', 'Bancos']:
            self.Button_Return.pressed.connect(lambda: self.CreatePage('10'))
            self.GLayout.addWidget(self.Background_Header, 0, 0, 1, 20)

            self.HBoxLayout_Header.addWidget(self.Button_Return, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_AttCotacoes, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_Home, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_ShowHideValues, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Label_Patrimonio)
            self.HBoxLayout_Header.addWidget(self.TextBox_Patrimonio, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Label_Rendimento)
            self.HBoxLayout_Header.addWidget(self.TextBox_Rendimento, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_Perfil, Qt.AlignRight)
            self.GLayout.addLayout(self.HBoxLayout_Header, 0, 0, 1, 20)

        elif self.PageID in ['20', '22', '23', '24', '25']:
            self.Button_Return.pressed.connect(lambda: self.CreatePage('18'))
            self.GLayout.addWidget(self.Background_Header, 0, 0, 1, 20)

            self.HBoxLayout_Header.addWidget(self.Button_Return, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_AttCotacoes, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_Home, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_ShowHideValues, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Label_Patrimonio)
            self.HBoxLayout_Header.addWidget(self.TextBox_Patrimonio, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Label_Rendimento)
            self.HBoxLayout_Header.addWidget(self.TextBox_Rendimento, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_Perfil, Qt.AlignRight)
            self.GLayout.addLayout(self.HBoxLayout_Header, 0, 0, 1, 20)

        elif self.PageID in ['26', '27', '28', '29', '30']:
            self.Button_Return.pressed.connect(lambda: self.CreatePage('23'))
            self.GLayout.addWidget(self.Background_Header, 0, 0, 1, 20)

            self.HBoxLayout_Header.addWidget(self.Button_Return, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_AttCotacoes, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_Home, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_ShowHideValues, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Label_Patrimonio)
            self.HBoxLayout_Header.addWidget(self.TextBox_Patrimonio, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Label_Rendimento)
            self.HBoxLayout_Header.addWidget(self.TextBox_Rendimento, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_Perfil, Qt.AlignRight)
            self.GLayout.addLayout(self.HBoxLayout_Header, 0, 0, 1, 20)

        elif self.PageID in ['31a', '32a', '33a', '34a', '35a']:
            self.Button_Return.pressed.connect(lambda: self.CreatePage('26'))
            self.GLayout.addWidget(self.Background_Header, 0, 0, 1, 20)

            self.HBoxLayout_Header.addWidget(self.Button_Return, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_Home, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_ShowHideValues, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Label_Patrimonio)
            self.HBoxLayout_Header.addWidget(self.TextBox_Patrimonio, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Label_Rendimento)
            self.HBoxLayout_Header.addWidget(self.TextBox_Rendimento, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_Perfil, Qt.AlignRight)
            self.GLayout.addLayout(self.HBoxLayout_Header, 0, 0, 1, 20)

        elif self.PageID in ['36a']:
            self.Button_Return.pressed.connect(lambda: self.CreatePage('34a'))
            self.GLayout.addWidget(self.Background_Header, 0, 0, 1, 20)

            self.HBoxLayout_Header.addWidget(self.Button_Return, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_Home, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_ShowHideValues, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Label_Patrimonio)
            self.HBoxLayout_Header.addWidget(self.TextBox_Patrimonio, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Label_Rendimento)
            self.HBoxLayout_Header.addWidget(self.TextBox_Rendimento, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_Perfil, Qt.AlignRight)
            self.GLayout.addLayout(self.HBoxLayout_Header, 0, 0, 1, 20)

        elif self.PageID in ['36b']:
            self.Button_Return.pressed.connect(lambda: self.CreatePage('35a'))
            self.GLayout.addWidget(self.Background_Header, 0, 0, 1, 20)

            self.HBoxLayout_Header.addWidget(self.Button_Return, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_Home, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_ShowHideValues, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Label_Patrimonio)
            self.HBoxLayout_Header.addWidget(self.TextBox_Patrimonio, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Label_Rendimento)
            self.HBoxLayout_Header.addWidget(self.TextBox_Rendimento, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_Perfil, Qt.AlignRight)
            self.GLayout.addLayout(self.HBoxLayout_Header, 0, 0, 1, 20)

        elif self.PageID in ['31b', '32b', '33b', '34b', '35b']:
            self.Button_Return.pressed.connect(lambda: self.CreatePage('27'))
            self.GLayout.addWidget(self.Background_Header, 0, 0, 1, 20)

            self.HBoxLayout_Header.addWidget(self.Button_Return, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_Home, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_ShowHideValues, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Label_Patrimonio)
            self.HBoxLayout_Header.addWidget(self.TextBox_Patrimonio, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Label_Rendimento)
            self.HBoxLayout_Header.addWidget(self.TextBox_Rendimento, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_Perfil, Qt.AlignRight)
            self.GLayout.addLayout(self.HBoxLayout_Header, 0, 0, 1, 20)

        elif self.PageID in ['37', '39', '40']:
            self.Button_Return.pressed.connect(lambda: self.CreatePage('19'))
            self.GLayout.addWidget(self.Background_Header, 0, 0, 1, 20)

            self.HBoxLayout_Header.addWidget(self.Button_Return, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_AttCotacoes, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_Home, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_ShowHideValues, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Label_Patrimonio)
            self.HBoxLayout_Header.addWidget(self.TextBox_Patrimonio, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Label_Rendimento)
            self.HBoxLayout_Header.addWidget(self.TextBox_Rendimento, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_Perfil, Qt.AlignRight)
            self.GLayout.addLayout(self.HBoxLayout_Header, 0, 0, 1, 20)

        elif self.PageID in ['41', '43', '44']:
            self.Button_Return.pressed.connect(lambda: self.CreatePage('40'))
            self.GLayout.addWidget(self.Background_Header, 0, 0, 1, 20)

            self.HBoxLayout_Header.addWidget(self.Button_Return, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_AttCotacoes, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_Home, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_ShowHideValues, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Label_Patrimonio)
            self.HBoxLayout_Header.addWidget(self.TextBox_Patrimonio, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Label_Rendimento)
            self.HBoxLayout_Header.addWidget(self.TextBox_Rendimento, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_Perfil, Qt.AlignRight)
            self.GLayout.addLayout(self.HBoxLayout_Header, 0, 0, 1, 20)

        elif self.PageID in ['42']:
            self.Button_Return.pressed.connect(lambda: self.CreatePage('41'))
            self.GLayout.addWidget(self.Background_Header, 0, 0, 1, 20)

            self.HBoxLayout_Header.addWidget(self.Button_Return, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_AttCotacoes, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_Home, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_ShowHideValues, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Label_Patrimonio)
            self.HBoxLayout_Header.addWidget(self.TextBox_Patrimonio, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Label_Rendimento)
            self.HBoxLayout_Header.addWidget(self.TextBox_Rendimento, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_Perfil, Qt.AlignRight)
            self.GLayout.addLayout(self.HBoxLayout_Header, 0, 0, 1, 20)

        elif self.PageID in ['34c', '35c']:
            self.Button_Return.pressed.connect(lambda: self.CreatePage('42'))
            self.GLayout.addWidget(self.Background_Header, 0, 0, 1, 20)

            self.HBoxLayout_Header.addWidget(self.Button_Return, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_Home, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_ShowHideValues, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Label_Patrimonio)
            self.HBoxLayout_Header.addWidget(self.TextBox_Patrimonio, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Label_Rendimento)
            self.HBoxLayout_Header.addWidget(self.TextBox_Rendimento, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_Perfil, Qt.AlignRight)
            self.GLayout.addLayout(self.HBoxLayout_Header, 0, 0, 1, 20)

        elif self.PageID in ['10', '13', '51']:
            self.Button_Return.pressed.connect(lambda: self.CreatePage('1'))
            self.GLayout.addWidget(self.Background_Header, 0, 0, 1, 20)

            self.HBoxLayout_Header.addWidget(self.Button_Return, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_AttCotacoes, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_Home, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_ShowHideValues, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Label_Patrimonio)
            self.HBoxLayout_Header.addWidget(self.TextBox_Patrimonio, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Label_Rendimento)
            self.HBoxLayout_Header.addWidget(self.TextBox_Rendimento, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_Perfil, Qt.AlignRight)
            self.GLayout.addLayout(self.HBoxLayout_Header, 0, 0, 1, 20)

        elif self.PageID in ['BancosAdd','BancosDelete','BancosRename']:
            self.Button_Return.pressed.connect(lambda: self.CreatePage('Bancos'))
            self.GLayout.addWidget(self.Background_Header, 0, 0, 1, 20)

            self.HBoxLayout_Header.addWidget(self.Button_Return, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_AttCotacoes, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_Home, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_ShowHideValues, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Label_Patrimonio)
            self.HBoxLayout_Header.addWidget(self.TextBox_Patrimonio, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Label_Rendimento)
            self.HBoxLayout_Header.addWidget(self.TextBox_Rendimento, Qt.AlignLeft)
            self.HBoxLayout_Header.addWidget(self.Button_Perfil, Qt.AlignRight)
            self.GLayout.addLayout(self.HBoxLayout_Header, 0, 0, 1, 20)

    #%% OnButtonPressed
    def OnButtonPressed(self, ButtonPressed):
        #%% ButtonPressed == "ShowHideValues"
        if ButtonPressed == "ShowHideValues":
            self.HMI_ARCA.Flag_RecalcularGraficos = True
            if self.ShowValues:
                self.ShowValues = not self.ShowValues
                self.Button_ShowHideValues.setIcon(QIcon("./images/iconClosedEye.png"))
                self.TextBox_Patrimonio.setEchoMode(QLineEdit.Password)
            else:
                self.ShowValues = not self.ShowValues
                self.Button_ShowHideValues.setIcon(QIcon("./images/iconOpenEye.png"))
                self.TextBox_Patrimonio.setEchoMode(QLineEdit.Normal)

            if self.PageID == '11': # ARCA
                try:
                    if self.ShowValues:
                        self.TextBox_Aplicar.setEchoMode(QLineEdit.Normal)
                        self.TextBox_GastoMensal.setEchoMode(QLineEdit.Normal)
                        self.TextBox_GastoMensal.setEnabled(True)
                        self.TextBox_MontanteAplicado.setEchoMode(QLineEdit.Normal)
                        self.TextBox_MontanteEmAplicacao.setEchoMode(QLineEdit.Normal)
                    else:
                        self.TextBox_Aplicar.setEchoMode(QLineEdit.Password)
                        self.TextBox_GastoMensal.setEchoMode(QLineEdit.Password)
                        self.TextBox_GastoMensal.setEnabled(False)
                        self.TextBox_MontanteAplicado.setEchoMode(QLineEdit.Password)
                        self.TextBox_MontanteEmAplicacao.setEchoMode(QLineEdit.Password)
                except:pass
                try:
                    self.HMI_ARCA.CreateTable_11_2() # Table

                    self.GraphWidget_Carteira.deleteLater()
                    if self.HMI_ARCA.CarteiraGraph_id == 'pie': self.GraphWidget_Carteira = self.HMI_ARCA.CriarPlot_P11_Carteira_pie()
                    elif self.HMI_ARCA.CarteiraGraph_id == 'bar': self.GraphWidget_Carteira = self.HMI_ARCA.CriarPlot_P11_Carteira_bar()
                    self.F_GLayout[self.GCount-1].addWidget(self.GraphWidget_Carteira, 2, 2, 2, 2, Qt.AlignHCenter | Qt.AlignBottom)
                    self.Button_CarteiraInvisivel.deleteLater()
                    self.Button_CarteiraInvisivel = QPushButton()
                    self.Button_CarteiraInvisivel.pressed.connect(lambda: self.HMI_ARCA.OnButtonPressed('Change Carteira Graph'))
                    self.Button_CarteiraInvisivel.setStyleSheet('background-color: rgba(0, 0, 0, 0); color: white; border: 1px solid rgba(0, 0, 0, 0); border: none; outline: none;')
                    if self.frameGeometry().width()/self.frameGeometry().height() < 1.9577: proporcao = 1/6
                    elif self.frameGeometry().width()/self.frameGeometry().height() < 2.6202: proporcao = 1/8
                    else: proporcao = 1/10
                    self.GraphWidget_Carteira.setFixedSize(int(self.frameGeometry().width()*proporcao),int(self.frameGeometry().width()*proporcao))
                    self.Button_CarteiraInvisivel.setFixedSize(int(self.frameGeometry().height()*proporcao),int(self.frameGeometry().height()*proporcao))
                    self.Button_CarteiraInvisivel.setCursor(QCursor(Qt.PointingHandCursor))
                    self.F_GLayout[self.GCount-1].addWidget(self.Button_CarteiraInvisivel, 2, 2, 2, 2, Qt.AlignHCenter | Qt.AlignBottom)
                except:pass

            elif self.PageID == '18': # BolsaPage
                if self.ShowValues:
                    self.TextBox_Msg2.setEchoMode(QLineEdit.Normal)
                    self.TextBox_Msg3.setEchoMode(QLineEdit.Normal)
                    self.TextBox_Msg5.setEchoMode(QLineEdit.Normal)
                else:
                    self.TextBox_Msg2.setEchoMode(QLineEdit.Password)
                    self.TextBox_Msg3.setEchoMode(QLineEdit.Password)
                    self.TextBox_Msg5.setEchoMode(QLineEdit.Password)

            elif self.PageID == '19': # CriptomoedasPage
                if self.ShowValues:
                    self.TextBox_Msg2.setEchoMode(QLineEdit.Normal)
                    self.TextBox_Msg3.setEchoMode(QLineEdit.Normal)
                    self.TextBox_Msg5.setEchoMode(QLineEdit.Normal)
                else:
                    self.TextBox_Msg2.setEchoMode(QLineEdit.Password)
                    self.TextBox_Msg3.setEchoMode(QLineEdit.Password)
                    self.TextBox_Msg5.setEchoMode(QLineEdit.Password)

            elif self.PageID == '23': # RegistrosPage
                if self.ShowValues:
                    self.TextBox_Msg2.setEchoMode(QLineEdit.Normal) # Estoque
                else:
                    self.TextBox_Msg2.setEchoMode(QLineEdit.Password) # Estoque
                self.HMI_Trades.HMI_Trades_Bolsa.CreateTable_23_2() # Table

            elif self.PageID == '24': # Depositos e Saques Page
                if self.ShowValues:
                    self.HMI_Trades.HMI_Trades_Bolsa.CreateTable_24_2() # Table
                else:
                    self.HMI_Trades.HMI_Trades_Bolsa.CreateTable_24_2() # Table

            elif self.PageID == '25': # Conta-corrente Page
                if self.ShowValues:
                    self.TextBox_Msg2.setEchoMode(QLineEdit.Normal)
                else:
                    self.TextBox_Msg2.setEchoMode(QLineEdit.Password)

            elif self.PageID == '40':
                self.HMI_Trades.HMI_Trades_Criptomoedas.CreateTable_40_2() # Table

            elif self.PageID == '51': # Tributação Page
                if self.ShowValues:
                    self.TextBox_Msg2.setEchoMode(QLineEdit.Normal)
                    self.TextBox_Msg3.setEchoMode(QLineEdit.Normal)
                    self.TextBox_Msg4.setEchoMode(QLineEdit.Normal)
                else:
                    self.TextBox_Msg2.setEchoMode(QLineEdit.Password)
                    self.TextBox_Msg3.setEchoMode(QLineEdit.Password)
                    self.TextBox_Msg4.setEchoMode(QLineEdit.Password)
                self.HMI_Tributacao.CreateTable_51_2()

        #%% ButtonPressed == "Atualizar Tabela de Cotações"
        elif ButtonPressed == 'Atualizar Tabela de Cotações':
            for row in range(self.Table_CotaçõesManuais.rowCount()):
                if not str(self.data_45_1[row][2]) == str(self.Table_CotaçõesManuais.item(row, 2).text()):
                    # Apenas atualizar os ativos que são diferentes do início
                    self.DBManager.UpdateCotacao(self.data_45_1[row][0], self.Table_CotaçõesManuais.item(row, 2).text())
            self.DBManager.AtualizarCotacoes()

            self.CreatePage(self.PageID)

        #%% ButtonPressed == "Path Padrão"
        elif ButtonPressed == 'Path Padrão':
            self.Button_PathPadrao.setText("PADRÃO")
            self.Button_PathEspecifico.setText("específico")
            self.TextBox_Path.setText("")
            self.Button_PathEspecifico.setStyleSheet("background-color: black; color: white;")
            self.Button_PathPadrao.setStyleSheet("background-color: green; color: white;")

        #%% ButtonPressed == "Path Específico"
        elif ButtonPressed == 'Path Específico':
            if not self.filedialogIsOpen:
                root = tkinter.Tk()
                root.withdraw() # Used to hide tkinter window
                self.filedialogIsOpen = True
                Path = filedialog.askdirectory() # Select folder path
                self.filedialogIsOpen = False
                try:
                    if len(Path)>0:
                        self.Button_PathPadrao.setText("padrão")
                        self.Button_PathEspecifico.setText("ESPECÍFICO")
                        self.TextBox_Path.setText(Path)
                        self.Button_PathEspecifico.setStyleSheet("background-color: green; color: white;")
                        self.Button_PathPadrao.setStyleSheet("background-color: black; color: white;")
                except:pass

        #%% ButtonPressed == "Change UserPath"
        elif ButtonPressed == "Change UserPath":
            def UpdateUserPath(UserName, PathNovo):
                currdir = os.getcwd()
                conn = sqlite3.connect(currdir+'\\Users.db')
                cursor = conn.cursor()
                cursor.execute("UPDATE Usuários SET Path = (?) Where Usuário = (?)", [(PathNovo), UserName])
                conn.commit()
                conn.close()
            if not self.filedialogIsOpen:
                root = tkinter.Tk()
                root.withdraw() # Used to hide tkinter window
                self.filedialogIsOpen = True
                Path = filedialog.askdirectory() # Select folder path
                self.filedialogIsOpen = False
                try:
                    if len(Path) > 0 and not self.DBManager.UserPath == Path:
                        self.DBManager.Disconnect()
                        UpdateUserPath(self.User, Path)
                        copyfile(self.DBManager.UserPath+"/"+self.User+'.db', Path+"/"+self.User+'.db')
                        os.remove(self.DBManager.UserPath+"/"+self.User+'.db')
                        self.DBManager.UserPath = Path
                        self.DBManager.ConnectToDB(Path+"\\"+self.User+".db")
                except Exception as e: print(e)

        #%% ButtonPressed == "ChangeProfilePhoto"
        elif ButtonPressed == "ChangeProfilePhoto":
            root = tkinter.Tk()
            root.withdraw() # Used to hide tkinter window
            currdir = os.getcwd() # Get current directory
            if not self.filedialogIsOpen:
                self.filedialogIsOpen = True
                DB_Path = filedialog.askopenfilename(parent=root, title='Selecione a sua foto', filetypes=[("Arquivos JPG", ".jpg")]) # Select folder path
                self.setCursor(Qt.WaitCursor)
                self.filedialogIsOpen = False
                if len(DB_Path)>0:
                    try:
                        copyfile(DB_Path,currdir+"\\images\\ProfilePhoto_"+self.User+'.jpg')
                        im = Image.open(".\images\ProfilePhoto_"+self.User+'.jpg')
                        img_width, img_height = im.size
                        crop_width = min(im.size)
                        crop_height = min(im.size)
                        im.crop(((img_width - crop_width) // 2,
                                 (img_height - crop_height) // 2,
                                 (img_width + crop_width) // 2,
                                 (img_height + crop_height) // 2)).save(".\images\ProfilePhoto_"+self.User+'.jpg', quality=95)
                    except: pass
                self.unsetCursor()
                self.CreatePage('9')

        #%% ButtonPressed == "ExportToExcel"
        elif ButtonPressed == "ExportToExcel":
            if not self.filedialogIsOpen:
                MessageBox_Msg1 = QMessageBox.about(self,'Aviso','Para continuar a exportação,\nSALVE seu trabalho e FECHE todas\nas janelas de EXCEL abertas.')
                root = tkinter.Tk()
                root.withdraw() # Used to hide tkinter window
                self.filedialogIsOpen = True
                DB_Path = filedialog.askdirectory() # Select folder path
                self.setCursor(Qt.WaitCursor)
                self.filedialogIsOpen = False
                try:
                    if len(DB_Path)>0:
                        agora = datetime.now().strftime("%Y %m %d")
                        FileName = self.User+"_database_"+agora+'.xlsx'
                        workbook = xlsxwriter.Workbook(DB_Path+"/"+FileName)
                        conn = sqlite3.connect(self.User+'.db')
                        tables, cursor = self.DBManager.GetDataDB("SELECT name FROM sqlite_master WHERE type='table';")
                        tables = list(map(lambda x: x[0], tables))
                        for table in tables:
                            try: worksheet = workbook.add_worksheet(name=table[0:31])
                            except Exception as e:print(e)

                            CONTENT, cursor = self.DBManager.GetDataDB('SELECT * FROM '+table)
                            headers = list(map(lambda x: x[0], cursor.description))
                            row1 = []
                            if len(CONTENT)>0: row1 = CONTENT[0]
                            else: worksheet.write(0, 0, "Nada declarado aqui")
                            for column_number, item in enumerate(row1):
                                worksheet.write(0, column_number, headers[column_number])

                            for row_number, row in enumerate(CONTENT):
                                for column_number, item in enumerate(row):
                                    try: worksheet.write(row_number+1, column_number, item)
                                    except Exception as e:print(e)
                        workbook.close()

                        excel = win32.gencache.EnsureDispatch('Excel.Application')
                        wb = excel.Workbooks.Open(DB_Path+"/"+FileName)
                        for table in tables:
                            ws = wb.Worksheets(table[0:31])
                            ws.Columns.AutoFit()
                        wb.Save()
                        excel.Application.Quit()
                except Exception as e:
                    workbook.close()
                finally: self.unsetCursor()

        #%% ButtonPressed == "ExportToDB"
        elif ButtonPressed == "ExportToDB":
            root = tkinter.Tk()
            root.withdraw() # use to hide tkinter window
            currdir = os.getcwd() # Get current directory
            if not self.filedialogIsOpen:
                self.filedialogIsOpen = True
                DB_Path = filedialog.askdirectory(parent=root, initialdir=currdir, title='Selecione a pasta para salvar seu backup') # Select folder path
                self.filedialogIsOpen = False
                if len(DB_Path)>0:
                    try:
                        copyfile(currdir+"\\"+self.User+'.db', DB_Path+"\\"+self.User+' Database '+str(datetime.now().strftime("%Y %m %d %H %M %S"))+'.db')
                        MessageBox_Msg1 = QMessageBox.about(self,'Informação','Exportação bem sucedida.')
                    except: MessageBox_Msg1 = QMessageBox.about(self,'Aviso','Exportação mal sucedida.')

        #%% ButtonPressed == "ImportFromDB"
        elif ButtonPressed == "ImportFromDB":
            def TrocaUserName(NomeVelho, NomeNovo):
                self.DBManager.ModifyDB("UPDATE Usuários SET Usuário = (?) Where Usuário = (?)", [(NomeNovo), NomeVelho])

            def TrocaUserEmail(EmailVelho, EmailNovo):
                self.DBManager.ModifyDB("UPDATE Usuários SET Email = (?) Where Email = (?)", [(EmailNovo), EmailVelho])

            def UpdateUserPath(UserName, PathNovo):
                currdir = os.getcwd()
                conn = sqlite3.connect(currdir+'\\Users.db')
                cursor = conn.cursor()
                cursor.execute("UPDATE Usuários SET Path = (?) Where Usuário = (?)", [(PathNovo), UserName])
                conn.commit()
                conn.close()

            root = tkinter.Tk()
            root.withdraw() # use to hide tkinter window
            currdir = os.getcwd() # Get current directory
            if not self.filedialogIsOpen:
                self.filedialogIsOpen = True
                file_path = filedialog.askopenfilename(filetypes=[("DB files", "*.db")])
                self.filedialogIsOpen = False
                if len(file_path)>0:
                    try:
                        conn = sqlite3.connect(file_path)
                        cursor = conn.cursor()
                        corretoras = list(cursor.execute("SELECT * FROM Corretoras"))
                        TableNames = list(cursor.execute("SELECT name FROM sqlite_master WHERE type='table';"))
                        TableNames = list(map(lambda x: x[0],TableNames))
                        if not ("TiposDeAtivo" in TableNames and
                                "SubtiposDeAtivo" in TableNames and
                                "Setores" in TableNames and
                                "Bolsa" in TableNames and
                                "Cotações" in TableNames and
                                "MoedasFiat" in TableNames and
                                "TributaçãoDTBolsa" in TableNames and
                                "TributaçãoSTBolsa" in TableNames and
                                "TributaçãoDTCripto" in TableNames and
                                "TributaçãoSTCripto" in TableNames): raise ValueError("Faltam dados,")

                        retorno = list(cursor.execute("SELECT * FROM Perfil"))
                        UserName = str(retorno[0][0])
                        UserEmail = str(retorno[0][2])
                        UserPath = os.path.split(file_path)[0]

                        conn.close()

                        # Verificar se usuário já existe
                        self.DBManager.CriarDB('Users.db') # Create Usuários.db IF NOT EXISTS
                        self.DBManager.ConnectToDB('Users.db') # Connect to Users.db
                        self.DBManager.CreateTableInCurrentDB('Usuários', [('Usuário', 'TEXT'), ('Email', 'TEXT'), ('Path', 'TEXT')]) # Create Table Usuários IF NOT EXISTS

                        retorno, cursor = self.DBManager.GetDataDB("SELECT * FROM Usuários")
                        retorno = list(filter(lambda x: x[0]==UserName or x[1]==UserEmail, retorno))

                        if len(retorno) > 0: # Usuário já está aqui dentro do programa.
                            # Escolher com qual db quer ficar.
                            MessageBox_Msg1 = QMessageBox()
                            MessageBox_Msg1.setWindowTitle("Escolha um")
                            if retorno[0][0] == UserName and retorno[0][1] == UserEmail: MessageBox_Msg1.setText("Já existe um usuário com esse nome e email na base de dados.\nDeseja substitui-lo?")
                            if retorno[0][0] == UserName: MessageBox_Msg1.setText("Já existe um usuário com esse nome na base de dados.\nDeseja substitui-lo?")
                            if retorno[0][1] == UserEmail: MessageBox_Msg1.setText("Já existe um usuário com esse email na base de dados.\nDeseja substitui-lo?")
                            MessageBox_Msg1.setIcon(QMessageBox.Warning)
                            MessageBox_Msg1.setStandardButtons(QMessageBox.Yes|QMessageBox.No|QMessageBox.Cancel)
                            MessageBox_Msg1.setDefaultButton(QMessageBox.No)
                            MessageBox_Msg1.setWindowIcon(QIcon(".\images\logoARCA.png"))
                            returnValue = MessageBox_Msg1.exec()

                            if returnValue == QMessageBox.No: raise ValueError("Usuário já existe.")
                            elif returnValue == QMessageBox.Cancel: raise ValueError("Operação cancelada.")
                            elif returnValue == QMessageBox.Yes:
                                Flag_TrocaUserName = False
                                Flag_TrocaUserEmail = False
                                # Se diferente, escolher qual UserName manter
                                if not retorno[0][0] == UserName:
                                    MessageBox_Msg1 = QMessageBox()
                                    MessageBox_Msg1.setWindowTitle("Escolha um")
                                    MessageBox_Msg1.setText("Os nomes de usuário são diferentes.\nPressione 'Yes' para escolher o nome "+retorno[0][0]+" ou\npressione 'No' para escolher '"+UserName+"'.\n'Cancel' cancela a importação.")
                                    MessageBox_Msg1.setIcon(QMessageBox.Warning)
                                    MessageBox_Msg1.setStandardButtons(QMessageBox.Yes|QMessageBox.No|QMessageBox.Cancel)
                                    MessageBox_Msg1.setDefaultButton(QMessageBox.No)
                                    MessageBox_Msg1.setWindowIcon(QIcon(".\images\logoARCA.png"))
                                    returnValue = MessageBox_Msg1.exec()

                                    if returnValue == QMessageBox.Cancel: raise ValueError("Operação cancelada.")
                                    elif returnValue == QMessageBox.Yes: UserName = retorno[0][0]
                                    elif returnValue == QMessageBox.No: Flag_TrocaUserName = True

                                # Se diferente, escolher qual UserEmail manter
                                if not retorno[0][1] == UserEmail:
                                    MessageBox_Msg1 = QMessageBox()
                                    MessageBox_Msg1.setWindowTitle("Escolha um")
                                    MessageBox_Msg1.setText("Os emails de usuário são diferentes.\nPressione 'Yes' para escolher o email "+retorno[0][1]+" ou\npressione 'No' para escolher '"+UserEmail+"'.\n'Cancel' cancela a importação.")
                                    MessageBox_Msg1.setIcon(QMessageBox.Warning)
                                    MessageBox_Msg1.setStandardButtons(QMessageBox.Yes|QMessageBox.No|QMessageBox.Cancel)
                                    MessageBox_Msg1.setDefaultButton(QMessageBox.No)
                                    MessageBox_Msg1.setWindowIcon(QIcon(".\images\logoARCA.png"))
                                    returnValue = MessageBox_Msg1.exec()

                                    if returnValue == QMessageBox.Cancel: raise ValueError("Operação cancelada.")
                                    elif returnValue == QMessageBox.Yes: pass
                                    elif returnValue == QMessageBox.No: Flag_TrocaUserEmail = True

                                if Flag_TrocaUserName: TrocaUserName(retorno[0][0], UserName)
                                if Flag_TrocaUserEmail: TrocaUserEmail(retorno[0][1], UserEmail)
                                UpdateUserPath(UserName, UserPath)
                                try:os.remove(os.path.split(file_path)[0]+"\\"+UserName+".db")
                                except:pass
                                os.rename(file_path, os.path.split(file_path)[0]+"\\"+UserName+".db")

                        else: # Usuário nunca esteve nessa versão do programa
                            # Adicionar Usuário e Email
                            self.DBManager.AddRowInCurrentDB('Usuários', [('Usuário',UserName), ('Email',UserEmail), ('Path', UserPath)])
                        # Disconnect from Users.db
                        self.DBManager.Disconnect()

                        MessageBox_Msg1 = QMessageBox.about(self,'Informação','Importação bem sucedida.\nBem-vindo de volta '+UserName+'.\nAcesse suas informações entrando\npela página inicial.')
                        self.CreatePage("0")
                    except Exception as msg:
                        if "Usuário já existe." in repr(msg):
                            self.DBManager.Disconnect()
                            MessageBox_Msg1 = QMessageBox.about(self,'Informação','Importação cancelada.\n'+str(msg))
                        elif "Operação cancelada." in repr(msg):
                            self.DBManager.Disconnect()
                            MessageBox_Msg1 = QMessageBox.about(self,'Informação','Importação cancelada.')
                        else:
                            MessageBox_Msg1 = QMessageBox.about(self,'Aviso','Importação mal sucedida.\n'+str(msg)+'\nArquivo fora do formato padrão\nOU\nusuário já existe.')

        #%% ButtonPressed == "ImportFromExcel"
        elif ButtonPressed == "ImportFromExcel":
            def TrocaUserName(NomeVelho, NomeNovo):
                def GetRowId(UserName):
                    retorno, cursor = self.DBManager.GetDataDB("SELECT rowid, * FROM Usuários")
                    retorno = list(filter(lambda x: str(x).find(UserName)>-1, retorno))
                    retorno = retorno[0][0]
                    return str(retorno)
                self.DBManager.ModifyDB("UPDATE Usuários SET Usuário = (?) Where rowid = "+GetRowId(NomeVelho), [(NomeNovo)])

            def TrocaUserEmail(EmailVelho, EmailNovo):
                def GetRowId(UserEmail):
                    retorno, cursor = self.DBManager.GetDataDB("SELECT rowid, * FROM Usuários")
                    retorno = list(filter(lambda x: str(x).find(UserEmail)>-1, retorno))
                    retorno = retorno[0][0]
                    return str(retorno)
                self.DBManager.ModifyDB("UPDATE Usuários SET Email = (?) Where rowid = "+GetRowId(EmailVelho), [(EmailNovo)])

            MessageBox_Msg1 = QMessageBox.about(self,'Informação','Funcionalidade ainda não implementada.\nUse a importação por .db')
            # root = tkinter.Tk()
            # root.withdraw() # use to hide tkinter window
            # currdir = os.getcwd() # Get current directory
            # if not self.filedialogIsOpen:
            #     self.filedialogIsOpen = True
            #     file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
            #     self.filedialogIsOpen = False
            #     if len(file_path)>0:
            #         try:
            #             workbook = load_workbook(filename=file_path)
            #             TableNames = workbook.sheetnames
            #             if not ("Bolsa" in TableNames and
            #                     "Cotações" in TableNames and
            #                     "Corretoras" in TableNames): raise ValueError("Faltam dados,")

            #             sheet = workbook["Perfil"]
            #             UserName = sheet["A2"].value
            #             UserPswd = sheet["B2"].value
            #             UserEmail = sheet["C2"].value
            #             UserCoin = sheet["E2"].value
            #             UserPath = os.path.split(file_path)[0]

            #             # Verificar se usuário já existe
            #             self.DBManager.CriarDB('Users.db') # Create Usuários.db IF NOT EXISTS
            #             self.DBManager.ConnectToDB('Users.db') # Connect to Users.db
            #             self.DBManager.CreateTableInCurrentDB('Usuários', [('Usuário', 'TEXT'), ('Email', 'TEXT'), ('Path', 'TEXT')]) # Create Table Usuários IF NOT EXISTS

            #             retorno, cursor = self.DBManager.GetDataDB("SELECT * FROM Usuários")
            #             aux = list(filter(lambda x: x[0]==UserName or x[1]==UserEmail, retorno))

            #             # if len(aux) > 0: # Usuário já está aqui dentro do programa.
            #             #     self.DBManager.Disconnect()
            #             #     # Escolher com qual db quer ficar.
            #             #     MessageBox_Msg1 = QMessageBox()
            #             #     MessageBox_Msg1.setWindowTitle("Escolha um")
            #             #     if retorno[0][0] == UserName and retorno[0][1] == UserEmail: MessageBox_Msg1.setText("Já existe um usuário com esse nome e email na base de dados.\nDeseja substitui-lo?")
            #             #     if retorno[0][0] == UserName: MessageBox_Msg1.setText("Já existe um usuário com esse nome na base de dados.\nDeseja substitui-lo?")
            #             #     if retorno[0][1] == UserEmail: MessageBox_Msg1.setText("Já existe um usuário com esse email na base de dados.\nDeseja substitui-lo?")
            #             #     MessageBox_Msg1.setIcon(QMessageBox.Warning)
            #             #     MessageBox_Msg1.setStandardButtons(QMessageBox.Yes|QMessageBox.No|QMessageBox.Cancel)
            #             #     MessageBox_Msg1.setDefaultButton(QMessageBox.No)
            #             #     MessageBox_Msg1.setWindowIcon(QIcon(".\images\logoARCA.png"))
            #             #     returnValue = MessageBox_Msg1.exec()

            #             #     if returnValue == QMessageBox.No: raise ValueError("Usuário já existe.")
            #             #     elif returnValue == QMessageBox.Cancel: raise ValueError("Operação cancelada.")
            #             #     elif returnValue == QMessageBox.Yes:
            #             #         Flag_TrocaUserName = False
            #             #         Flag_TrocaUserEmail = False

            #             #         # Se diferente, escolher qual UserName manter
            #             #         if not retorno[0][0] == UserName:
            #             #             MessageBox_Msg1 = QMessageBox()
            #             #             MessageBox_Msg1.setWindowTitle("Escolha um")
            #             #             MessageBox_Msg1.setText("Os nomes de usuário são diferentes.\nPressione 'Yes' para escolher o nome "+retorno[0][0]+" ou\npressione 'No' para escolher '"+UserName+"'.\n'Cancel' cancela a importação.")
            #             #             MessageBox_Msg1.setIcon(QMessageBox.Warning)
            #             #             MessageBox_Msg1.setStandardButtons(QMessageBox.Yes|QMessageBox.No|QMessageBox.Cancel)
            #             #             MessageBox_Msg1.setDefaultButton(QMessageBox.No)
            #             #             MessageBox_Msg1.setWindowIcon(QIcon(".\images\logoARCA.png"))
            #             #             returnValue = MessageBox_Msg1.exec()

            #             #             if returnValue == QMessageBox.Cancel: raise ValueError("Operação cancelada.")
            #             #             elif returnValue == QMessageBox.Yes: UserName = retorno[0][0]
            #             #             elif returnValue == QMessageBox.No: Flag_TrocaUserName = True

            #             #         # Se diferente, escolher qual UserEmail manter
            #             #         if not retorno[0][1] == UserEmail:
            #             #             MessageBox_Msg1 = QMessageBox()
            #             #             MessageBox_Msg1.setWindowTitle("Escolha um")
            #             #             MessageBox_Msg1.setText("Os emails de usuário são diferentes.\nPressione 'Yes' para escolher o email "+retorno[0][1]+" ou\npressione 'No' para escolher '"+UserEmail+"'.\n'Cancel' cancela a importação.")
            #             #             MessageBox_Msg1.setIcon(QMessageBox.Warning)
            #             #             MessageBox_Msg1.setStandardButtons(QMessageBox.Yes|QMessageBox.No|QMessageBox.Cancel)
            #             #             MessageBox_Msg1.setDefaultButton(QMessageBox.No)
            #             #             MessageBox_Msg1.setWindowIcon(QIcon(".\images\logoARCA.png"))
            #             #             returnValue = MessageBox_Msg1.exec()

            #             #             if returnValue == QMessageBox.Cancel: raise ValueError("Operação cancelada.")
            #             #             elif returnValue == QMessageBox.Yes: pass
            #             #             elif returnValue == QMessageBox.No: Flag_TrocaUserEmail = True

            #             #         # Import excel

            #             #         # sheet = workbook["Corretoras"]
            #             #         # corretoras = []
            #             #         # for i in range(2, sheet.max_row+1):
            #             #         #     corretoras.append(sheet["A"+str(i)].value)

            #             #         # for corretora in corretoras:
            #             #         #     if "¨Bolsa" in corretora:
            #             #         #         # Importar ¨DepSaq
            #             #         #         # Importar ¨Bruto
            #             #         #         # Recalcular ¨DT e importar
            #             #         #         # Recalcular ¨ST e importar

            #             #         #     elif "¨Cripto" in corretora:
            #             #         #         # Importar ¨Bruto
            #             #         #         # Recalcular ¨Refinado e importar

            #             #         # # Recalcular folhas Refinado, DT e ST

            #             #         # # Importar:
            #             #         #     # 1) Perfil
            #             #         #     # 2) Bolsa
            #             #         #     # 3) Corretoras

            #             #         # if Flag_TrocaUserName: TrocaUserName(retorno[0][0], UserName)
            #             #         # if Flag_TrocaUserEmail: TrocaUserEmail(retorno[0][1], UserEmail)
            #             #         # UpdateUserPath(UserName, UserPath)
            #             #         # try:os.remove(os.path.split(file_path)[0]+"\\"+UserName+".db")
            #             #         # except:pass
            #             #         # os.rename(file_path, os.path.split(file_path)[0]+"\\"+UserName+".db")
            #             #     else: # Usuário nunca esteve nessa versão do programa
            #             self.DBManager.Disconnect()

            #             con = sqlite3.connect(os.path.split(file_path)[0]+"\\"+UserName+".db")
            #             wb = pd.read_excel(file_path,sheet_name = None)

            #             for sheet in wb:
            #                 try:wb[sheet].to_sql(sheet,con,index=False)
            #                 except Exception as e:
            #                     if not ("near " in str(e) and ": syntax error" in str(e)):
            #                         con.close()
            #                         # print('erro na folha ',sheet,": ", e)
            #                         raise ValueError("")
            #                     elif "already exists." in str(e):
            #                         con.close()
            #                         raise ValueError("Usuário já existe")
            #                     # else: print('erro na folha ',sheet,": ", e)
            #             con.commit()
            #             con.close()

            #             self.DBManager.ConnectToDB('Users.db')
            #             self.DBManager.AddRowInCurrentDB('Usuários', [('Usuário',UserName), ('Email',UserEmail), ('Path', UserPath)])
            #             self.DBManager.Disconnect()

            #             MessageBox_Msg1 = QMessageBox.about(self,'Informação','Importação bem sucedida.\nBem-vindo de volta '+UserName+'.\nAcesse suas informações entrando\npela página inicial.')
            #             self.CreatePage("0")
            #         except Exception as msg:
            #             if "Usuário já existe." in repr(msg):
            #                 self.DBManager.Disconnect()
            #                 MessageBox_Msg1 = QMessageBox.about(self,'Informação','Importação cancelada.\n'+str(msg))
            #             elif "Operação cancelada." in repr(msg):
            #                 self.DBManager.Disconnect()
            #                 MessageBox_Msg1 = QMessageBox.about(self,'Informação','Importação cancelada.')
            #             else:
            #                 MessageBox_Msg1 = QMessageBox.about(self,'Aviso','Importação mal sucedida.\n'+str(msg)+'\nArquivo fora do formato padrão\nOU\nusuário já existe.')

        #%% ButtonPressed == "Abrir Calendário"
        elif ButtonPressed == "Abrir Calendário":
            if not self.OutraJanelaAberta:
                self.Window_Calendario = Calendario(self)
                self.Window_Calendario.Abrir()
            elif self.CalendarWindowOpen:
                self.Window_Calendario.close()
                self.Window_Calendario.show()
            elif self.ClockWindowOpen:
                self.Window_Relogio.close()
                self.Window_Relogio.show()

        #%% ButtonPressed == "Abrir Relógio"
        elif ButtonPressed == "Abrir Relógio":
            if not self.OutraJanelaAberta:
                self.Window_Relogio = Relogio(self)
                self.Window_Relogio.Abrir()
            elif self.CalendarWindowOpen:
                self.Window_Calendario.close()
                self.Window_Calendario.show()
            elif self.ClockWindowOpen:
                self.Window_Relogio.close()
                self.Window_Relogio.show()

    #%% Keyboard
    def keyPressEvent(self, event):
        key = event.key()
        if key == Qt.Key_Escape:
            if self.PageIDAux in ['9', '45']:
                if self.PageID in ['28', '29', '34a', '34b', '35a', '35b', '36a', '36b']: self.CreatePage('23')
                elif self.PageID in ['41', '43']: self.CreatePage('40')
                else: self.CreatePage(self.PageID)
            elif self.PageIDAux in ['14', '15', '17a', '17b', 'ChooseCurrency']:
                self.CreatePage('9')
            elif self.PageID in ['2', '3', '4', '5', '7']:
                self.CreatePage('0')
            elif self.PageID in ['6']:
                self.CreatePage('2')
            elif self.PageID in ['8']:
                self.CreatePage('4')
            elif self.PageID in ['10', '11', '13', '12', '51']:
                self.CreatePage('1')
            elif self.PageID in ['46', '47', '48', '49', '50']:
                self.CreatePage('11')
            elif self.PageID in ['18', '19', 'Bancos']:
                self.CreatePage('10')
            elif self.PageID in['BancosAdd','BancosRename']:
                self.CreatePage('Bancos')
            elif self.PageID in ['20', '22', '23', '24', '25']:
                self.CreatePage('18')
            elif self.PageID in ['26', '27', '28', '29', '30']:
                self.CreatePage('23')
            elif self.PageID in ['31a', '32a', '33a', '34a', '35a']:
                self.CreatePage('26')
            elif self.PageID in ['36a']:
                self.CreatePage('34a')
            elif self.PageID in ['36b']:
                self.CreatePage('35a')
            elif self.PageID in ['31b', '32b', '33b', '34b', '35b']:
                self.CreatePage('27')
            elif self.PageID in ['37', '39', '40']:
                self.CreatePage('19')
            elif self.PageID in ['41', '43', '44']:
                self.CreatePage('40')
            elif self.PageID in ['42']:
                self.CreatePage('41')
            elif self.PageID in ['34c', '35c']:
                self.CreatePage('42')
        elif key == Qt.Key_H: self.OnButtonPressed('ShowHideValues')
        elif key == Qt.Key_P: self.CreatePage('9')
        elif key == Qt.Key_C: self.CreatePage('45')
        elif key == Qt.Key_M: self.CreatePage('1')
        elif key == Qt.Key_N: self.CreatePage('10')
        elif key == Qt.Key_T: self.CreatePage('51')
        elif key == Qt.Key_A: self.CreatePage('11')
        elif key == Qt.Key_D: self.CreatePage('12')

    #%% Resizing HMI
    def eventFilter(self, obj, event): # Função pra redimensionar alguns elementos do IHM dinamicamente
        if event.type() == QEvent.Resize:
            try:
                #%% Resizing HMI Page 9
                if self.PageIDAux == '9':
                    self.Button_ProfilePhoto.setFixedSize(int(self.frameGeometry().width()/8),int(self.frameGeometry().width()/8))
                    self.Button_ProfilePhoto.setIconSize(QSize(int(self.frameGeometry().width()/8*.95),int(self.frameGeometry().width()/8*.95)))

                    self.Button_ExportDBToDB.setFixedSize(int(self.frameGeometry().width()*15/20/2),int(self.frameGeometry().height()-160))
                    self.Button_ExportDBToDB.setIconSize(QSize(int(self.frameGeometry().width()*15/20/2*.98),int(self.frameGeometry().height())))

                    self.Button_ExportDBToExcel.setFixedSize(int(self.frameGeometry().width()*15/20/2),int(self.frameGeometry().height()-160))
                    self.Button_ExportDBToExcel.setIconSize(QSize(int(self.frameGeometry().width()*15/20/2*.98),int(self.frameGeometry().height())))

                #%% Resizing HMI Page 45
                elif self.PageIDAux == '45':
                    HHeader = ['Ativo',
                               'Última atualização',
                               'Cotação em '+self.DBManager.GetUserCoinCurrency()]
                    SizeCol0, SizeCol1, SizeCol2 = self.GetSizeOfTableColumns(self.data_45_1, HHeader)
                    self.SetTableWidth_45_1(SizeCol0, SizeCol1, SizeCol2)
                    self.Table_CotaçõesManuais.setFixedHeight(int(self.frameGeometry().height()*6/10))

                #%% Resizing HMI Page 0
                elif self.PageID == '0': # Home Page Not Logged In
                    self.Button_Mercado.setFixedSize(int(self.frameGeometry().width()*3/4*.9),int(self.frameGeometry().height()/4*.95))
                    self.Movie_Mercado.setScaledSize(QSize().scaled(int(self.frameGeometry().width()*3/4*.90),int(self.frameGeometry().height()*3/4*.95), Qt.KeepAspectRatio))
                    self.TextBox_NomeDeUsuario.setFixedWidth(int(self.frameGeometry().width()/3*0.9))
                    self.TextBox_Senha.setFixedWidth(int(self.frameGeometry().width()/3*0.9))
                    self.Button_LogIn.setFixedSize(int(self.frameGeometry().width()/3*0.9),int(self.frameGeometry().height()/10))
                    self.Button_LogIn.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_MercadoInvisivel.setFixedSize(int(self.frameGeometry().width()*3/4*.90),int(self.frameGeometry().height()*3/4*.95))

                #%% Resizing HMI Page 1
                elif self.PageID == '1':
                    self.Movie_Negociações.setScaledSize(QSize().scaled(int(self.frameGeometry().width()/3*.80),int(self.frameGeometry().height()/5*.95-20), Qt.KeepAspectRatio))
                    self.Button_Negociações.setFixedSize(int(self.frameGeometry().width()/3*.96),int(self.frameGeometry().height()/5))
                    self.Button_NegociaçõesInvisivel.setFixedSize(int(self.frameGeometry().width()/3*.98),int(self.frameGeometry().height()*4/5-160))

                    self.Movie_Tributacao.setScaledSize(QSize().scaled(int(self.frameGeometry().width()/3*.80),int(self.frameGeometry().height()/5*.95-20), Qt.KeepAspectRatio))
                    self.Button_Tributacao.setFixedSize(int(self.frameGeometry().width()/3*.96),int(self.frameGeometry().height()/5))
                    self.Button_TributacaoInvisivel.setFixedSize(int(self.frameGeometry().width()/3*.98),int(self.frameGeometry().height()*4/5-160))

                    self.Movie_ARCA.setScaledSize(QSize().scaled(int(self.frameGeometry().width()/3*.80),int(self.frameGeometry().height()/5*.95-20), Qt.KeepAspectRatio))
                    self.Button_ARCA.setFixedSize(int(self.frameGeometry().width()/3*.96),int(self.frameGeometry().height()/5))
                    self.Button_ARCAInvisivel.setFixedSize(int(self.frameGeometry().width()/3*.80),int(self.frameGeometry().height()/5*.95-20))

                    self.Movie_Desempenho.setScaledSize(QSize().scaled(int(self.frameGeometry().width()/3*.80),int(self.frameGeometry().height()/5*.95-20), Qt.KeepAspectRatio))
                    self.Button_Desempenho.setFixedSize(int(self.frameGeometry().width()/3*.96),int(self.frameGeometry().height()/5))
                    self.Button_DesempenhoInvisivel.setFixedSize(int(self.frameGeometry().width()/3*.80),int(self.frameGeometry().height()/5*.95-20))

                    self.Movie_Mercado.setScaledSize(QSize().scaled(int(self.frameGeometry().width()/3*.98),int(self.frameGeometry().height()*4/5-110), Qt.KeepAspectRatio))
                    self.Button_Mercado.setFixedSize(int(self.frameGeometry().width()/3*.96),int(self.frameGeometry().height()/5))
                    self.Button_MercadoInvisivel.setFixedSize(int(self.frameGeometry().width()/3*.98),int(self.frameGeometry().height()*4/5-110))

                #%% Resizing HMI Page 3
                elif self.PageID == '3':
                    self.Button_ExportDBToExcel.setFixedSize(int(self.frameGeometry().width()*2/5),int(self.frameGeometry().height()-160))
                    self.Button_ExportDBToDB.setFixedSize(int(self.frameGeometry().width()*2/5),int(self.frameGeometry().height()-160))

                #%% Resizing HMI Page 4
                elif self.PageID == '4':
                    minimo = 24
                    maximo = 36
                    a = (int(self.screen_width*0.5))/((1/minimo)-(1/maximo))
                    b = self.screen_width-(a/minimo)
                    razao = (self.frameGeometry().width()-b)/a
                    self.Label_NomeDeUsuario.setText('Nome de usuário:'+str((' ')*int(self.frameGeometry().width()*razao)))

                    minimo = 19
                    maximo = 21
                    a = (int(self.screen_width*0.5))/((1/minimo)-(1/maximo))
                    b = self.screen_width-(a/minimo)
                    razao = (self.frameGeometry().width()-b)/a
                    self.Label_Email.setText('Email:'+str((' ')*int(self.frameGeometry().width()*razao)))

                    minimo = 19
                    maximo = 21
                    a = (int(self.screen_width*0.5))/((1/minimo)-(1/maximo))
                    b = self.screen_width-(a/minimo)
                    razao = (self.frameGeometry().width()-b)/a
                    self.Label_Senha1.setText('Senha:'+str((' ')*int(self.frameGeometry().width()*razao)))

                    minimo = 24
                    maximo = 39
                    a = (int(self.screen_width*0.5))/((1/minimo)-(1/maximo))
                    b = self.screen_width-(a/minimo)
                    razao = (self.frameGeometry().width()-b)/a
                    self.Label_Senha2.setText('Confirme a senha:'+str((' ')*int(self.frameGeometry().width()*razao)))

                    self.TextBox_NomeDeUsuario.setFixedWidth(int(self.frameGeometry().width()/3*0.9))
                    self.TextBox_Email.setFixedWidth(int(self.frameGeometry().width()/3*0.9))

                    self.TextBox_Senha1.setFixedWidth(int(self.frameGeometry().width()/3*0.9))
                    self.TextBox_Senha2.setFixedWidth(int(self.frameGeometry().width()/3*0.9))

                    self.Button_ConfirmacaoDeEmail.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_ConfirmacaoDeEmail.setFixedWidth(int(self.frameGeometry().width()/3*0.9))

                #%% Resizing HMI Page 10
                elif self.PageID == '10':
                    self.Button_TradesCriptomoedasInvisivel.setFixedSize(int(self.frameGeometry().width()/3*.98),int(self.frameGeometry().height()*4/5-160))
                    self.Button_TradesCriptomoedas.setFixedSize(int(self.frameGeometry().width()/3*.96),int(self.frameGeometry().height()/5))
                    self.Movie_TradesCriptomoedas.setScaledSize(QSize().scaled(int(self.frameGeometry().width()/3*.98),int(self.frameGeometry().height()*4/5-160), Qt.KeepAspectRatio))
                    self.Button_TradesBolsaInvisivel.setFixedSize(int(self.frameGeometry().width()/3*.98),int(self.frameGeometry().height()*4/5-160))
                    self.Button_TradesBolsa.setFixedSize(int(self.frameGeometry().width()/3*.96),int(self.frameGeometry().height()/5))
                    self.Movie_TradesBolsa.setScaledSize(QSize().scaled(int(self.frameGeometry().width()/3*.98),int(self.frameGeometry().height()*4/5-160), Qt.KeepAspectRatio))
                    self.Button_BancosInvisivel.setFixedSize(int(self.frameGeometry().width()/3*.98),int(self.frameGeometry().height()*4/5-160))
                    self.Button_Bancos.setFixedSize(int(self.frameGeometry().width()/3*.96),int(self.frameGeometry().height()/5))
                    self.Movie_Bancos.setScaledSize(QSize().scaled(int(self.frameGeometry().width()/3*.98),int(self.frameGeometry().height()*4/5-160), Qt.KeepAspectRatio))

                #%% Resizing HMI Page Bancos
                elif self.PageID == 'Bancos':
                    self.Button_AtuatizarCC.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_AtuatizarCC.setFixedWidth(int(self.frameGeometry().width()/3*0.9))

                #%% Resizing HMI Page BancosAdd
                elif self.PageID == 'BancosAdd':
                    self.Button_AdicionarBanco.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_AdicionarBanco.setFixedWidth(int(self.frameGeometry().width()/3*0.9))

                #%% Resizing HMI Page BancosRename
                elif self.PageID == 'BancosRename':
                    self.Button_RenomearBanco.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_RenomearBanco.setFixedWidth(int(self.frameGeometry().width()/3*0.9))

                #%% Resizing HMI Page 11
                elif self.PageID == '11':
                    if self.frameGeometry().width()/self.frameGeometry().height() < 1.9577: proporcao = 1/6
                    elif self.frameGeometry().width()/self.frameGeometry().height() < 2.6202: proporcao = 1/8
                    else: proporcao = 1/10

                    self.Button_CarteiraInvisivel.setFixedSize(int(self.frameGeometry().width()*proporcao),int(self.frameGeometry().width()*proporcao))
                    self.Button_AcoesENegociosInvisivel.setFixedSize(int(self.frameGeometry().width()*proporcao),int(self.frameGeometry().width()*proporcao))
                    self.Button_RealEstateInvisivel.setFixedSize(int(self.frameGeometry().width()*proporcao),int(self.frameGeometry().width()*proporcao))
                    self.Button_CaixaInvisivel.setFixedSize(int(self.frameGeometry().width()*proporcao),int(self.frameGeometry().width()*proporcao))
                    self.Button_AcoesInternacionaisInvisivel.setFixedSize(int(self.frameGeometry().width()*proporcao),int(self.frameGeometry().width()*proporcao))
                    self.Button_EspecificaInvisivel.setFixedSize(int(self.frameGeometry().width()*proporcao),int(self.frameGeometry().width()*proporcao))
                    self.GraphWidget_Carteira.setFixedSize(int(self.frameGeometry().width()*proporcao),int(self.frameGeometry().width()*proporcao))
                    self.GraphWidget_AcoesENegocios.setFixedSize(int(self.frameGeometry().width()*proporcao),int(self.frameGeometry().width()*proporcao))
                    self.GraphWidget_RealEstate.setFixedSize(int(self.frameGeometry().width()*proporcao),int(self.frameGeometry().width()*proporcao))
                    self.GraphWidget_Caixa.setFixedSize(int(self.frameGeometry().width()*proporcao),int(self.frameGeometry().width()*proporcao))
                    self.GraphWidget_AtivosInternacionais.setFixedSize(int(self.frameGeometry().width()*proporcao),int(self.frameGeometry().width()*proporcao))
                    self.GraphWidget_Especifica.setFixedSize(int(self.frameGeometry().width()*proporcao),int(self.frameGeometry().width()*proporcao))

                    self.TextBox_MontanteNasCorretoras.setFixedWidth(int(self.frameGeometry().width()/6))
                    self.TextBox_MontanteAplicado.setFixedWidth(int(self.frameGeometry().width()/6))

                    self.TextBox_GastoMensal.setFixedWidth(int(self.frameGeometry().width()/10))
                    self.TextBox_MesesDeReserva.setFixedWidth(int(self.frameGeometry().width()/20))
                    self.TextBox_Aplicar.setFixedWidth(int(self.frameGeometry().width()/10))

                    self.Table_ContasCorrente.setFixedHeight(int(self.frameGeometry().height()*3/10))

                #%% Resizing HMI Page 12
                elif self.PageID == '12':
                    self.Movie_Grafico2.setScaledSize(QSize().scaled(int(self.frameGeometry().width()),int(self.frameGeometry().height()/5*.95), Qt.KeepAspectRatio))
                    self.Movie_Grafico1.setScaledSize(QSize().scaled(int(self.frameGeometry().width()),int(self.frameGeometry().height()/5*.95), Qt.KeepAspectRatio))

                #%% Resizing HMI Page 5 and 13
                elif self.PageID in ['5','13']:
                    self.Movie_Mercado1.setScaledSize(QSize().scaled(int(self.frameGeometry().width()/3*.98),int(self.frameGeometry().height()*4/5-110), Qt.KeepAspectRatio))
                    self.Button_Mercado1.setFixedSize(int(self.frameGeometry().width()/3*.96),int(self.frameGeometry().height()/5))
                    self.Button_MercadoInvisivel1.setFixedSize(int(self.frameGeometry().width()/3*.98),int(self.frameGeometry().height()*4/5-110))
                    self.Movie_Mercado2.setScaledSize(QSize().scaled(int(self.frameGeometry().width()/3*.98),int(self.frameGeometry().height()*4/5-110), Qt.KeepAspectRatio))
                    self.Button_Mercado2.setFixedSize(int(self.frameGeometry().width()/3*.96),int(self.frameGeometry().height()/5))
                    self.Button_MercadoInvisivel2.setFixedSize(int(self.frameGeometry().width()/3*.98),int(self.frameGeometry().height()*4/5-110))
                    self.Movie_Mercado3.setScaledSize(QSize().scaled(int(self.frameGeometry().width()/3*.98),int(self.frameGeometry().height()*4/5-110), Qt.KeepAspectRatio))
                    self.Button_Mercado3.setFixedSize(int(self.frameGeometry().width()/3*.96),int(self.frameGeometry().height()/5))
                    self.Button_MercadoInvisivel3.setFixedSize(int(self.frameGeometry().width()/3*.98),int(self.frameGeometry().height()*4/5-110))
                    if self.frameGeometry().width() < self.screen_width*0.7:
                        self.Button_Mercado1.setFont(self.font20)
                        self.Button_Mercado2.setFont(self.font20)
                        self.Button_Mercado3.setFont(self.font20)
                    else:
                        self.Button_Mercado1.setFont(self.font24)
                        self.Button_Mercado2.setFont(self.font24)
                        self.Button_Mercado3.setFont(self.font24)

                #%% Resizing HMI Page 15
                elif self.PageID == '15':
                    minimo = 23
                    maximo = 43
                    a = (int(self.screen_width*0.5))/((1/minimo)-(1/maximo))
                    b = self.screen_width-(a/minimo)
                    razao = (self.frameGeometry().width()-b)/a
                    self.Label_SenhaAtual.setText('Insira a senha atual:'+str((' ')*int(self.frameGeometry().width()*razao)))

                    minimo = 33
                    maximo = 47
                    a = (int(self.screen_width*0.5))/((1/minimo)-(1/maximo))
                    b = self.screen_width-(a/minimo)
                    razao = (self.frameGeometry().width()-b)/a
                    self.Label_Senha1.setText('Crie uma nova senha:'+str((' ')*int(self.frameGeometry().width()*razao)))

                    minimo = 21
                    maximo = 38
                    a = (int(self.screen_width*0.5))/((1/minimo)-(1/maximo))
                    b = self.screen_width-(a/minimo)
                    razao = (self.frameGeometry().width()-b)/a
                    self.Label_Senha2.setText('Confirme a senha:'+str((' ')*int(self.frameGeometry().width()*razao)))

                #%% Resizing HMI Page 18
                elif self.PageID == '18':
                    self.Movie_AlterarOPBolsa.setScaledSize(QSize().scaled(int(self.frameGeometry().width()*12/20/2),int(self.frameGeometry().height()*7/10*.95-30), Qt.KeepAspectRatio))
                    self.Button_AlterarOPBolsaInvisivel.setFixedSize(int(self.frameGeometry().width()*12/20/2),int(self.frameGeometry().height()*7/10*.95))
                    self.Button_AlterarOPBolsa.setFixedSize(int(self.frameGeometry().width()*12/20/2),int(self.frameGeometry().height()*1/10*.95))
                    self.Movie_DepositosESaquesBolsa.setScaledSize(QSize().scaled(int(self.frameGeometry().width()*12/20/2),int(self.frameGeometry().height()*3/10*.95-30), Qt.KeepAspectRatio))
                    self.Button_DepositosESaquesBolsaInvisivel.setFixedSize(int(self.frameGeometry().width()*12/20/2),int(self.frameGeometry().height()*3/10*.95))
                    self.Button_DepositosESaquesBolsa.setFixedSize(int(self.frameGeometry().width()*12/20/2),int(self.frameGeometry().height()*1/10*.95))
                    self.Movie_ContaCorrenteBolsa.setScaledSize(QSize().scaled(int(self.frameGeometry().width()*12/20/2),int(self.frameGeometry().height()*3/10*.95-30), Qt.KeepAspectRatio))
                    self.Button_ContaCorrenteBolsaInvisivel.setFixedSize(int(self.frameGeometry().width()*12/20/2),int(self.frameGeometry().height()*3/10*.95))
                    self.Button_ContaCorrenteBolsa.setFixedSize(int(self.frameGeometry().width()*12/20/2),int(self.frameGeometry().height()*1/10*.95))

                #%% Resizing HMI Page 19
                elif self.PageID == '19':
                    self.Movie_AlterarOPCriptomoedas.setScaledSize(QSize().scaled(int(self.frameGeometry().width()*12/20),int(self.frameGeometry().height()*7/10*.95-60), Qt.KeepAspectRatio))
                    self.Button_AlterarOPCriptomoedasInvisivel.setFixedSize(int(self.frameGeometry().width()*12/20),int(self.frameGeometry().height()*7/10*.95))
                    self.Button_AlterarOPCriptomoedas.setFixedSize(int(self.frameGeometry().width()*12/20),int(self.frameGeometry().height()*1/10*.95))

                #%% Resizing HMI Page 20
                elif self.PageID == '20':
                    self.TextBox_NovaCorretora.setFixedWidth(int(self.frameGeometry().width()/3*0.9))
                    self.Button_AdicionarCorretora.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_AdicionarCorretora.setFixedWidth(int(self.frameGeometry().width()/3*0.9))

                #%% Resizing HMI Page 21
                elif self.PageID == '21':pass

                #%% Resizing HMI Page 22
                elif self.PageID == '22':
                    self.TextBox_NovoNomeCorretora.setFixedWidth(int(self.frameGeometry().width()/3*0.9))
                    self.Button_RenomearCorretora.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_RenomearCorretora.setFixedWidth(int(self.frameGeometry().width()/3*0.9))

                #%% Resizing HMI Page 23
                elif self.PageID == '23':
                    HHeader = ['Data',
                               'Tipo',
                               'Qqt',
                               'Preço',
                               'Corretagem',
                               'Taxa B3%',
                               'Taxa B3',
                               'Obs']
                    SizeCol0,SizeCol1,SizeCol2,SizeCol3,SizeCol4,SizeCol5,SizeCol6,SizeCol7 = self.GetSizeOfTableColumns(self.HMI_Trades.HMI_Trades_Bolsa.data_23_1, HHeader)

                    self.Table_Operacoes_Realizadas_com_AtivoHeader.setSizeAdjustPolicy(QAbstractScrollArea.AdjustToContents)
                    self.Table_Operacoes_Realizadas_com_AtivoHeader.setFixedHeight(int(self.frameGeometry().height()*1/20))
                    self.Table_Operacoes_Realizadas_com_AtivoHeader.adjustSize()
                    self.Table_Operacoes_Realizadas_com_AtivoHeader.setColumnWidth(0, SizeCol0)
                    self.Table_Operacoes_Realizadas_com_AtivoHeader.setColumnWidth(1, SizeCol1)
                    self.Table_Operacoes_Realizadas_com_AtivoHeader.setColumnWidth(2, SizeCol2)
                    self.Table_Operacoes_Realizadas_com_AtivoHeader.setColumnWidth(3, SizeCol3)
                    self.Table_Operacoes_Realizadas_com_AtivoHeader.setColumnWidth(4, SizeCol4)
                    self.Table_Operacoes_Realizadas_com_AtivoHeader.setColumnWidth(5, SizeCol5)
                    self.Table_Operacoes_Realizadas_com_AtivoHeader.setColumnWidth(6, SizeCol6)

                    self.Table_Operacoes_Realizadas_com_Ativo.setSizeAdjustPolicy(QAbstractScrollArea.AdjustToContents)
                    self.Table_Operacoes_Realizadas_com_Ativo.setFixedHeight(int(self.frameGeometry().height()*6/20*1.4))
                    self.Table_Operacoes_Realizadas_com_Ativo.resizeColumnsToContents()
                    self.Table_Operacoes_Realizadas_com_Ativo.adjustSize()
                    self.Table_Operacoes_Realizadas_com_Ativo.setColumnWidth(0, SizeCol0)
                    self.Table_Operacoes_Realizadas_com_Ativo.setColumnWidth(1, SizeCol1)
                    self.Table_Operacoes_Realizadas_com_Ativo.setColumnWidth(2, SizeCol2)
                    self.Table_Operacoes_Realizadas_com_Ativo.setColumnWidth(3, SizeCol3)
                    self.Table_Operacoes_Realizadas_com_Ativo.setColumnWidth(4, SizeCol4)
                    self.Table_Operacoes_Realizadas_com_Ativo.setColumnWidth(5, SizeCol5)
                    self.Table_Operacoes_Realizadas_com_Ativo.setColumnWidth(6, SizeCol6)
                    self.Table_Operacoes_Realizadas_com_Ativo.setColumnWidth(7, SizeCol7)

                #%% Resizing HMI Page 24
                elif self.PageID == '24':
                    self.Button_Voltar.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_Voltar.setFixedWidth(int(self.frameGeometry().width()/3*0.9))
                    self.Button_Registrar.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_Registrar.setFixedWidth(int(self.frameGeometry().width()/3*0.9))
                    self.Table_DepositosESaques.setFixedHeight(int(self.frameGeometry().height()*5/20*1.4))

                #%% Resizing HMI Page 25
                elif self.PageID == '25':
                    self.Button_Voltar.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_Voltar.setFixedWidth(int(self.frameGeometry().width()/3*0.9))
                    self.Button_Atualizar.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_Atualizar.setFixedWidth(int(self.frameGeometry().width()/3*0.9))

                #%% Resizing HMI Page 26
                elif self.PageID == '26':
                    self.TextBox_NomeNovoAtivo.setFixedWidth(int(self.frameGeometry().width()/8*0.9))
                    self.Button_Seguinte.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_Seguinte.setFixedWidth(int(self.frameGeometry().width()/3*0.9))

                #%% Resizing HMI Page 27
                elif self.PageID == '27':
                    self.TextBox_NovoNomeAtivo.setFixedWidth(int(self.frameGeometry().width()/8*0.9))
                    self.Button_Seguinte.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_Seguinte.setFixedWidth(int(self.frameGeometry().width()/3*0.9))

                #%% Resizing HMI Page 28
                elif self.PageID == '28':
                    self.TextBox_Qqt.setFixedWidth(int(self.frameGeometry().width()/7*0.9))
                    self.TextBox_Preco.setFixedWidth(int(self.frameGeometry().width()/7*0.9))
                    self.TextBox_Corretagem.setFixedWidth(int(self.frameGeometry().width()/7*0.9))
                    self.TextBox_TaxaB3.setFixedWidth(int(self.frameGeometry().width()/7*0.9))
                    self.TextBox_Obs.setFixedWidth(int(self.frameGeometry().width()*4/7*0.9))
                    self.Button_LimparCampos.setFixedSize(int(self.frameGeometry().width()*2/10*0.9),int(self.frameGeometry().width()*1/10*0.9))
                    self.Button_RegistrarOperacao.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_RegistrarOperacao.setFixedSize(int(self.frameGeometry().width()*2/10*0.9),int(self.frameGeometry().width()*1/10*0.9))

                #%% Resizing HMI Page 29
                elif self.PageID == '29':
                    HHeader = ['Data',
                               'Tipo',
                               'Qqt',
                               'Preço',
                               'Corretagem',
                               'Taxa B3%',
                               'Taxa B3',
                               'Obs']
                    SizeCol0,SizeCol1,SizeCol2,SizeCol3,SizeCol4,SizeCol5,SizeCol6,SizeCol7 = self.GetSizeOfTableColumns(self.HMI_Trades.HMI_Trades_Bolsa.data_23_1, HHeader)

                    self.Table_Operacoes_SelecionadasHeader.setSizeAdjustPolicy(QAbstractScrollArea.AdjustToContents)
                    self.Table_Operacoes_SelecionadasHeader.setFixedHeight(int(self.frameGeometry().height()*1/20))
                    self.Table_Operacoes_SelecionadasHeader.adjustSize()
                    self.Table_Operacoes_SelecionadasHeader.setColumnWidth(0, SizeCol0)
                    self.Table_Operacoes_SelecionadasHeader.setColumnWidth(1, SizeCol1)
                    self.Table_Operacoes_SelecionadasHeader.setColumnWidth(2, SizeCol2)
                    self.Table_Operacoes_SelecionadasHeader.setColumnWidth(3, SizeCol3)
                    self.Table_Operacoes_SelecionadasHeader.setColumnWidth(4, SizeCol4)
                    self.Table_Operacoes_SelecionadasHeader.setColumnWidth(5, SizeCol5)
                    self.Table_Operacoes_SelecionadasHeader.setColumnWidth(6, SizeCol6)
                    self.Table_Operacoes_SelecionadasHeader.setColumnWidth(7, SizeCol7)

                    self.Table_Operacoes_Selecionadas.setSizeAdjustPolicy(QAbstractScrollArea.AdjustToContents)
                    self.Table_Operacoes_Selecionadas.setFixedHeight(int(self.frameGeometry().height()*1/20))
                    self.Table_Operacoes_Selecionadas.resizeColumnsToContents()
                    self.Table_Operacoes_Selecionadas.adjustSize()
                    self.Table_Operacoes_Selecionadas.setColumnWidth(0, SizeCol0)
                    self.Table_Operacoes_Selecionadas.setColumnWidth(1, SizeCol1)
                    self.Table_Operacoes_Selecionadas.setColumnWidth(2, SizeCol2)
                    self.Table_Operacoes_Selecionadas.setColumnWidth(3, SizeCol3)
                    self.Table_Operacoes_Selecionadas.setColumnWidth(4, SizeCol4)
                    self.Table_Operacoes_Selecionadas.setColumnWidth(5, SizeCol5)
                    self.Table_Operacoes_Selecionadas.setColumnWidth(6, SizeCol6)
                    self.Table_Operacoes_Selecionadas.setColumnWidth(7, SizeCol7)

                    self.Button_LimparCampos.setFixedSize(int(self.frameGeometry().width()*2/10*0.9),int(self.frameGeometry().width()*1/10*0.9))
                    self.Button_RegistrarOperacao.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_RegistrarOperacao.setFixedSize(int(self.frameGeometry().width()*2/10*0.9),int(self.frameGeometry().width()*1/10*0.9))
                    self.TextBox_Qqt.setFixedWidth(int(self.frameGeometry().width()/7*0.9))
                    self.TextBox_Preco.setFixedWidth(int(self.frameGeometry().width()/7*0.9))
                    self.TextBox_Corretagem.setFixedWidth(int(self.frameGeometry().width()/7*0.9))
                    self.TextBox_TaxaB3.setFixedWidth(int(self.frameGeometry().width()/7*0.9))
                    self.TextBox_Obs.setFixedWidth(int(self.frameGeometry().width()*4/7*0.9))

                #%% Resizing HMI Page 30
                elif self.PageID == '30':
                    HHeader = ['Data',
                               'Tipo',
                               'Qqt',
                               'Preço',
                               'Corretagem',
                               'Taxa B3%',
                               'Taxa B3',
                               'Obs']
                    SizeCol0,SizeCol1,SizeCol2,SizeCol3,SizeCol4,SizeCol5,SizeCol6,SizeCol7 = self.GetSizeOfTableColumns(self.HMI_Trades.HMI_Trades_Bolsa.data_23_1, HHeader)

                    self.Table_Operacoes_Realizadas_com_AtivoHeader.setSizeAdjustPolicy(QAbstractScrollArea.AdjustToContents)
                    self.Table_Operacoes_Realizadas_com_AtivoHeader.setFixedHeight(int(self.frameGeometry().height()*1/20))
                    self.Table_Operacoes_Realizadas_com_AtivoHeader.adjustSize()
                    self.Table_Operacoes_Realizadas_com_AtivoHeader.setColumnWidth(0, SizeCol0)
                    self.Table_Operacoes_Realizadas_com_AtivoHeader.setColumnWidth(1, SizeCol1)
                    self.Table_Operacoes_Realizadas_com_AtivoHeader.setColumnWidth(2, SizeCol2)
                    self.Table_Operacoes_Realizadas_com_AtivoHeader.setColumnWidth(3, SizeCol3)
                    self.Table_Operacoes_Realizadas_com_AtivoHeader.setColumnWidth(4, SizeCol4)
                    self.Table_Operacoes_Realizadas_com_AtivoHeader.setColumnWidth(5, SizeCol5)
                    self.Table_Operacoes_Realizadas_com_AtivoHeader.setColumnWidth(6, SizeCol6)
                    self.Table_Operacoes_Realizadas_com_AtivoHeader.setColumnWidth(7, SizeCol7)

                    self.Table_Operacoes_Realizadas_com_Ativo.setSizeAdjustPolicy(QAbstractScrollArea.AdjustToContents)
                    self.Table_Operacoes_Realizadas_com_Ativo.setFixedHeight(int(self.frameGeometry().height()*6/20*1.4))
                    self.Table_Operacoes_Realizadas_com_Ativo.resizeColumnsToContents()
                    self.Table_Operacoes_Realizadas_com_Ativo.adjustSize()
                    self.Table_Operacoes_Realizadas_com_Ativo.setColumnWidth(0, SizeCol0)
                    self.Table_Operacoes_Realizadas_com_Ativo.setColumnWidth(1, SizeCol1)
                    self.Table_Operacoes_Realizadas_com_Ativo.setColumnWidth(2, SizeCol2)
                    self.Table_Operacoes_Realizadas_com_Ativo.setColumnWidth(3, SizeCol3)
                    self.Table_Operacoes_Realizadas_com_Ativo.setColumnWidth(4, SizeCol4)
                    self.Table_Operacoes_Realizadas_com_Ativo.setColumnWidth(5, SizeCol5)
                    self.Table_Operacoes_Realizadas_com_Ativo.setColumnWidth(6, SizeCol6)
                    self.Table_Operacoes_Realizadas_com_Ativo.setColumnWidth(7, SizeCol7)

                    self.Button_Cancelar.setFixedSize(int(self.frameGeometry().width()*2/10*0.9),int(self.frameGeometry().width()*1/10*0.9))
                    self.Button_DeletarOperacoes.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_DeletarOperacoes.setFixedSize(int(self.frameGeometry().width()*2/10*0.9),int(self.frameGeometry().width()*1/10*0.9))

                #%% Resizing HMI Page 31a
                elif self.PageID == '31a':
                    self.TextBox_NovoTipoDeAtivo.setFixedWidth(int(self.frameGeometry().width()/3*0.9))
                    self.Button_Seguinte.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_Seguinte.setFixedWidth(int(self.frameGeometry().width()/3*0.9))

                #%% Resizing HMI Page 31b
                elif self.PageID == '31b':
                    self.TextBox_NovoTipoDeAtivo.setFixedWidth(int(self.frameGeometry().width()/3*0.9))
                    self.Button_Seguinte.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_Seguinte.setFixedWidth(int(self.frameGeometry().width()/3*0.9))

                #%% Resizing HMI Page 32a
                elif self.PageID == '32a':
                    self.TextBox_NovoSubtipoDeAtivo.setFixedWidth(int(self.frameGeometry().width()/3*0.9))
                    self.Button_Seguinte.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_Seguinte.setFixedWidth(int(self.frameGeometry().width()/3*0.9))

                #%% Resizing HMI Page 32b
                elif self.PageID == '32b':
                    self.TextBox_NovoSubtipoDeAtivo.setFixedWidth(int(self.frameGeometry().width()/3*0.9))
                    self.Button_Seguinte.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_Seguinte.setFixedWidth(int(self.frameGeometry().width()/3*0.9))

                #%% Resizing HMI Page 33a
                elif self.PageID == '33a':
                    self.TextBox_NovoTipoSetor.setFixedWidth(int(self.frameGeometry().width()/3*0.9))
                    self.Button_Seguinte.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_Seguinte.setFixedWidth(int(self.frameGeometry().width()/3*0.9))

                #%% Resizing HMI Page 33b
                elif self.PageID == '33b':
                    self.TextBox_NovoSetor.setFixedWidth(int(self.frameGeometry().width()/3*0.9))
                    self.Button_Seguinte.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_Seguinte.setFixedWidth(int(self.frameGeometry().width()/3*0.9))

                #%% Resizing HMI Page 34a
                elif self.PageID == '34a':
                    self.Button_Voltar.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_Voltar.setFixedWidth(int(self.frameGeometry().width()/3*0.9))
                    self.Button_AtivarModoManual.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_AtivarModoManual.setFixedWidth(int(self.frameGeometry().width()/3*0.9))
                    self.Button_Seguinte.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_Seguinte.setFixedWidth(int(self.frameGeometry().width()/3*0.9))
                    self.Button_Info1.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_Info1.setFixedWidth(int(self.frameGeometry().width()/3*0.9))

                #%% Resizing HMI Page 34b
                elif self.PageID == '34b':
                    self.Button_Voltar.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_Voltar.setFixedWidth(int(self.frameGeometry().width()/3*0.9))
                    self.Button_AtivarModoManual.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_AtivarModoManual.setFixedWidth(int(self.frameGeometry().width()/3*0.9))
                    self.Button_Seguinte.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_Seguinte.setFixedWidth(int(self.frameGeometry().width()/3*0.9))
                    self.Button_Info1.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_Info1.setFixedWidth(int(self.frameGeometry().width()/3*0.9))

                #%% Resizing HMI Page 35a
                elif self.PageID == '35a':
                    self.Button_Voltar.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_Voltar.setFixedWidth(int(self.frameGeometry().width()/3*0.9))
                    self.Button_AtivarModoManual.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_AtivarModoManual.setFixedWidth(int(self.frameGeometry().width()/3*0.9))
                    self.Button_Info1.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))

                #%% Resizing HMI Page 35b
                elif self.PageID == '35b':
                    self.Button_Voltar.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_Voltar.setFixedWidth(int(self.frameGeometry().width()/3*0.9))
                    self.Button_AtivarModoManual.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_AtivarModoManual.setFixedWidth(int(self.frameGeometry().width()/3*0.9))
                    self.Button_Info1.setFixedWidth(int(self.frameGeometry().width()/3*0.9))

                #%% Resizing HMI Page 36a
                elif self.PageID == '36a':
                    self.TextBox_Modo.setFixedWidth(int(self.frameGeometry().width()/7*0.9))
                    self.TextBox_Qqt.setFixedWidth(int(self.frameGeometry().width()/7*0.9))
                    self.TextBox_Preco.setFixedWidth(int(self.frameGeometry().width()/7*0.9))
                    self.TextBox_Corretagem.setFixedWidth(int(self.frameGeometry().width()/7*0.9))
                    self.TextBox_TaxaB3.setFixedWidth(int(self.frameGeometry().width()/7*0.9))
                    self.TextBox_Obs.setFixedWidth(int(self.frameGeometry().width()*4/7*0.9))
                    self.Button_LimparCampos.setFixedSize(int(self.frameGeometry().width()*2/10*0.9),int(self.frameGeometry().width()*1/10*0.9))
                    self.Button_RegistrarOperacao.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_RegistrarOperacao.setFixedSize(int(self.frameGeometry().width()*2/10*0.9),int(self.frameGeometry().width()*1/10*0.9))

                #%% Resizing HMI Page 40
                elif self.PageID == '40':
                    HHeader = ['Data',
                               'Par',
                               'Tipo',
                               'Preço',
                               'Qqt',
                               'Taxa',
                               'Moeda da taxa',
                               'Conversão',
                               'Obs']
                    SizeCol0,SizeCol1,SizeCol2,SizeCol3,SizeCol4,SizeCol5,SizeCol6,SizeCol7,SizeCol8 = self.GetSizeOfTableColumns(self.HMI_Trades.HMI_Trades_Criptomoedas.data_40_1, HHeader)

                    self.Table_Operacoes_RealizadasHeader.setSizeAdjustPolicy(QAbstractScrollArea.AdjustToContents)
                    self.Table_Operacoes_RealizadasHeader.setFixedHeight(int(self.frameGeometry().height()*1/20))
                    self.Table_Operacoes_RealizadasHeader.adjustSize()
                    self.Table_Operacoes_RealizadasHeader.setColumnWidth(0, SizeCol0)
                    self.Table_Operacoes_RealizadasHeader.setColumnWidth(1, SizeCol1)
                    self.Table_Operacoes_RealizadasHeader.setColumnWidth(2, SizeCol2)
                    self.Table_Operacoes_RealizadasHeader.setColumnWidth(3, SizeCol3)
                    self.Table_Operacoes_RealizadasHeader.setColumnWidth(4, SizeCol4)
                    self.Table_Operacoes_RealizadasHeader.setColumnWidth(5, SizeCol5)
                    self.Table_Operacoes_RealizadasHeader.setColumnWidth(6, SizeCol6)
                    self.Table_Operacoes_RealizadasHeader.setColumnWidth(7, SizeCol7)

                    self.Table_Operacoes_Realizadas.setSizeAdjustPolicy(QAbstractScrollArea.AdjustToContents)
                    self.Table_Operacoes_Realizadas.setFixedHeight(int(self.frameGeometry().height()*6/20*1.4))
                    self.Table_Operacoes_Realizadas.resizeColumnsToContents()
                    self.Table_Operacoes_Realizadas.adjustSize()
                    self.Table_Operacoes_Realizadas.setColumnWidth(0, SizeCol0)
                    self.Table_Operacoes_Realizadas.setColumnWidth(1, SizeCol1)
                    self.Table_Operacoes_Realizadas.setColumnWidth(2, SizeCol2)
                    self.Table_Operacoes_Realizadas.setColumnWidth(3, SizeCol3)
                    self.Table_Operacoes_Realizadas.setColumnWidth(4, SizeCol4)
                    self.Table_Operacoes_Realizadas.setColumnWidth(5, SizeCol5)
                    self.Table_Operacoes_Realizadas.setColumnWidth(6, SizeCol6)
                    self.Table_Operacoes_Realizadas.setColumnWidth(7, SizeCol7)
                    self.Table_Operacoes_Realizadas.setColumnWidth(8, SizeCol8)

                #%% Resizing HMI Page 41
                elif self.PageID == '41':
                    self.TextBox_Par_Esquerda.setFixedWidth(int(self.frameGeometry().width()/7*0.9))
                    self.TextBox_Par_Direita.setFixedWidth(int(self.frameGeometry().width()/7*0.9))
                    self.TextBox_Qqt.setFixedWidth(int(self.frameGeometry().width()/7*0.9))
                    self.TextBox_Preco.setFixedWidth(int(self.frameGeometry().width()/7*0.9))
                    self.TextBox_Taxa.setFixedWidth(int(self.frameGeometry().width()/7*0.9))
                    self.TextBox_MoedaDaTaxa.setFixedWidth(int(self.frameGeometry().width()/7*0.9))
                    self.TextBox_Conversao.setFixedWidth(int(self.frameGeometry().width()/7*0.9))
                    self.TextBox_Obs.setFixedWidth(int(self.frameGeometry().width()*4/7*0.9))
                    self.Button_LimparCampos.setFixedSize(int(self.frameGeometry().width()*2/10*0.9),int(self.frameGeometry().width()*1/10*0.9))
                    self.Button_RegistrarOperacao.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_RegistrarOperacao.setFixedSize(int(self.frameGeometry().width()*2/10*0.9),int(self.frameGeometry().width()*1/10*0.9))

                #%% Resizing HMI Page 43
                elif self.PageID == '43':
                    self.HMI_Trades.HMI_Trades_Criptomoedas.CreateTable_43_1() # Table_Operacoes_SelecionadasHeader
                    self.HMI_Trades.HMI_Trades_Criptomoedas.CreateTable_43_2() # Table_Operacoes_Selecionadas (Mostra apenas o item selecionado)

                    self.TextBox_Par_Esquerda.setFixedWidth(int(self.frameGeometry().width()/7*0.9))
                    self.TextBox_Par_Direita.setFixedWidth(int(self.frameGeometry().width()/7*0.9))
                    self.TextBox_Qqt.setFixedWidth(int(self.frameGeometry().width()/7*0.9))
                    self.TextBox_Preco.setFixedWidth(int(self.frameGeometry().width()/7*0.9))
                    self.TextBox_Taxa.setFixedWidth(int(self.frameGeometry().width()/7*0.9))
                    self.TextBox_MoedaDaTaxa.setFixedWidth(int(self.frameGeometry().width()/7*0.9))
                    self.TextBox_Conversao.setFixedWidth(int(self.frameGeometry().width()/7*0.9))
                    self.TextBox_Obs.setFixedWidth(int(self.frameGeometry().width()*4/7*0.9))
                    self.Button_LimparCampos.setFixedSize(int(self.frameGeometry().width()*2/10*0.9),int(self.frameGeometry().width()*1/10*0.9))
                    self.Button_RegistrarOperacao.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_RegistrarOperacao.setFixedSize(int(self.frameGeometry().width()*2/10*0.9),int(self.frameGeometry().width()*1/10*0.9))

                #%% Resizing HMI Page 44
                elif self.PageID == '44':
                    self.HMI_Trades.HMI_Trades_Criptomoedas.CreateTable_44_1() # Table_Operacoes_SelecionadasHeader
                    self.HMI_Trades.HMI_Trades_Criptomoedas.CreateTable_44_2() # Table_Operacoes_Selecionadas (Mostra apenas o item selecionado)

                    self.Button_Cancelar.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_Cancelar.setFixedSize(int(self.frameGeometry().width()*2/10*0.9),int(self.frameGeometry().width()*1/10*0.9))
                    self.Button_DeletarOperacoes.setIconSize(QSize(int(self.frameGeometry().height()/10*0.9),int(self.frameGeometry().height()/10*0.9)))
                    self.Button_DeletarOperacoes.setFixedSize(int(self.frameGeometry().width()*2/10*0.9),int(self.frameGeometry().width()*1/10*0.9))

                #%% Resizing HMI Page 51
                elif self.PageID == '51':
                    self.Table_Tributacao.setFixedHeight(int(self.frameGeometry().height()*6/20*1.4))
                    if not self.DTeST_51 == "DTeST":
                        HHeader = ['Mês/Ano',
                                   'Resultado líquido',
                                   'Imposto devido',
                                   'Alíquota',
                                   'Lucro mínimo taxável',
                                   'Resultado bruto',
                                   'Prejuízo acumulado',
                                   'Resultado final']

                        if int(self.frameGeometry().width()*6/20*1.4) < 510:
                            self.Table_Tributacao_Header.setFont(self.font10)
                            self.Table_Tributacao.setFont(self.font10)
                        else:
                            self.Table_Tributacao_Header.setFont(self.font14)
                            self.Table_Tributacao.setFont(self.font14)
                    else:
                        HHeader = ['Mês/Ano',
                                   'Resultado líquido',
                                   'Imposto devido',
                                   'Resultado bruto',
                                   'Prejuízo acumulado',
                                   'Resultado final']
                    if not self.DTeST_51 == "DTeST":
                        SizeCol0,SizeCol1,SizeCol2,SizeCol3,SizeCol4,SizeCol5,SizeCol6,SizeCol7 = self.GetSizeOfTableColumns(self.HMI_Tributacao.data_51_1, HHeader)

                        self.Table_Tributacao_Header.adjustSize()
                        self.Table_Tributacao_Header.setColumnWidth(0, SizeCol0)
                        self.Table_Tributacao_Header.setColumnWidth(1, SizeCol1)
                        self.Table_Tributacao_Header.setColumnWidth(2, SizeCol2)
                        self.Table_Tributacao_Header.setColumnWidth(3, SizeCol3)
                        self.Table_Tributacao_Header.setColumnWidth(4, SizeCol4)
                        self.Table_Tributacao_Header.setColumnWidth(5, SizeCol5)
                        self.Table_Tributacao_Header.setColumnWidth(6, SizeCol6)
                        self.Table_Tributacao_Header.setColumnWidth(7, SizeCol7)

                        self.Table_Tributacao.adjustSize()
                        self.Table_Tributacao.setColumnWidth(0, SizeCol0)
                        self.Table_Tributacao.setColumnWidth(1, SizeCol1)
                        self.Table_Tributacao.setColumnWidth(2, SizeCol2)
                        self.Table_Tributacao.setColumnWidth(3, SizeCol3)
                        self.Table_Tributacao.setColumnWidth(4, SizeCol4)
                        self.Table_Tributacao.setColumnWidth(5, SizeCol5)
                        self.Table_Tributacao.setColumnWidth(6, SizeCol6)
                        self.Table_Tributacao.setColumnWidth(7, SizeCol7)
                    else:
                        SizeCol0,SizeCol1,SizeCol2,SizeCol3,SizeCol4,SizeCol5 = self.GetSizeOfTableColumns(self.HMI_Tributacao.data_51_1, HHeader)

                        self.Table_Tributacao_Header.adjustSize()
                        self.Table_Tributacao_Header.setColumnWidth(0, SizeCol0)
                        self.Table_Tributacao_Header.setColumnWidth(1, SizeCol1)
                        self.Table_Tributacao_Header.setColumnWidth(2, SizeCol2)
                        self.Table_Tributacao_Header.setColumnWidth(3, SizeCol3)
                        self.Table_Tributacao_Header.setColumnWidth(4, SizeCol4)
                        self.Table_Tributacao_Header.setColumnWidth(5, SizeCol5)

                        self.Table_Tributacao.adjustSize()
                        self.Table_Tributacao.setColumnWidth(0, SizeCol0)
                        self.Table_Tributacao.setColumnWidth(1, SizeCol1)
                        self.Table_Tributacao.setColumnWidth(2, SizeCol2)
                        self.Table_Tributacao.setColumnWidth(3, SizeCol3)
                        self.Table_Tributacao.setColumnWidth(4, SizeCol4)
                        self.Table_Tributacao.setColumnWidth(5, SizeCol5)

            except:pass
        return False

#%% APP starter
class graphical_thread():
    def __init__(self,):
        self.mIHM = None
        self.app = None

    def run(self,):
        self.app = QApplication(sys.argv)
        self.app.setStyle("fusion")
        try:
            self.mIHM = IHMmain()
            self.app.exec_()
            if not self.DBManager.FLAG or self.AtualizarCotacoes_running:
                self.mIHM.DBManager.Disconnect()
            else:
                while self.DBManager.FLAG or self.AtualizarCotacoes_running:
                    pass
                self.mIHM.DBManager.Disconnect()
        except Exception as error:
            self.mIHM.DBManager.Disconnect()
            self.app.quit()
            sys.exit()
            QCoreApplication.quit()
            print('Erro crítico: \n' + repr(error))

mgraphical_thread = graphical_thread() # Debug
mgraphical_thread.run() # Debug