# -*- coding: utf-8 -*-
"""
Created on Fri Oct 15 23:05:53 2021

@author: caiop
"""

import pandas as pd
import numpy as np
import time

from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *

from datetime import datetime
import dateutil

import tkinter
from tkinter import filedialog
import xlsxwriter
from openpyxl import load_workbook
import win32com.client as win32


class HMI_Tributacao(object):
    def __init__(self, HMI):
        self.HMI = HMI
        QToolTip.setFont(QFont('Arial', 12))

    def CreatePage51(self):
        black = 'rgb(0, 0, 0)'
        Background_1 = QLabel()
        Background_1.setStyleSheet("background-color: "+black)
        Background_2 = QLabel()
        Background_2.setStyleSheet("background-color: "+black)
        Background_3 = QLabel()
        Background_3.setStyleSheet("background-color: "+black)

        self.HMI.Label_Titulo = QLabel('Tributação')
        self.HMI.Label_Titulo.setStyleSheet('color: white')
        self.HMI.Label_Titulo.setFont(self.HMI.font30)
        self.HMI.Label_Titulo.setAlignment(Qt.AlignCenter)

        self.HMI.Label_Msg2 = QLabel('Resultado geral líquido:')
        self.HMI.Label_Msg2.setStyleSheet('color: white')
        self.HMI.Label_Msg2.setFont(self.HMI.font16)
        self.HMI.Label_Msg2.setAlignment(Qt.AlignRight)

        self.HMI.TextBox_Msg2 = QLineEdit()
        self.HMI.TextBox_Msg2.setText(str(self.HMI.DBManager.GetResultadoGeralLiquido()))
        if self.HMI.ShowValues:
            self.HMI.TextBox_Msg2.setEchoMode(QLineEdit.Normal)
        else:
            self.HMI.TextBox_Msg2.setEchoMode(QLineEdit.Password)
        if float(self.HMI.TextBox_Msg2.text()) >= 0: self.HMI.TextBox_Msg2.setStyleSheet('QLineEdit {background-color: black; color: green;}')
        else: self.HMI.TextBox_Msg2.setStyleSheet('QLineEdit {background-color: black; color: red;}')
        self.HMI.TextBox_Msg2.setFont(self.HMI.font16)
        self.HMI.TextBox_Msg2.setEnabled(False)
        self.HMI.TextBox_Msg2.setAlignment(Qt.AlignLeft)

        self.HMI.Button_DTeST = QPushButton()
        self.HMI.Button_DTeST.pressed.connect(lambda: self.OnButtonPressed('ChangeDTeST'))
        self.HMI.Button_DTeST.setFixedSize(70*3+5,35*3+5)
        self.HMI.Button_DTeST.setIconSize(QSize(70*3, 35*3))
        self.HMI.Button_DTeST.setStyleSheet('QPushButton {background-color: rgba(0, 0, 0, 0); color: white; border: 1px solid rgba(0, 0, 0, 0); border: none; outline: none;} QToolTip {background-color: black; color: white;border: black solid 1px}')
        self.HMI.Button_DTeST.setCursor(QCursor(Qt.PointingHandCursor))
        self.HMI.Button_DTeST.setIcon(QIcon("./images/Visualizar Day Trades e Swing Trades.png"))
        self.HMI.Button_DTeST.setToolTip("Alterar apresentação (DT e ST)")
        self.HMI.DTeST_51 = "DTeST"

        self.HMI.Button_CriptoeB3 = QPushButton()
        self.HMI.Button_CriptoeB3.pressed.connect(lambda: self.OnButtonPressed('ChangeCriptoeB3'))
        self.HMI.Button_CriptoeB3.setFixedSize(70*3+5,35*3+5)
        self.HMI.Button_CriptoeB3.setIconSize(QSize(70*3, 35*3))
        self.HMI.Button_CriptoeB3.setStyleSheet('QPushButton {background-color: rgba(0, 0, 0, 0); color: white; border: 1px solid rgba(0, 0, 0, 0); border: none; outline: none;} QToolTip {background-color: black; color: white;border: black solid 1px}')
        self.HMI.Button_CriptoeB3.setCursor(QCursor(Qt.PointingHandCursor))
        self.HMI.Button_CriptoeB3.setIcon(QIcon("./images/Visualizar cripto e B3.png"))
        self.HMI.Button_CriptoeB3.setToolTip("Alterar apresentação (cripto e B3)")
        self.HMI.CriptoeB3_51 = "CriptoeB3"

        self.HMI.Label_Msg3 = QLabel('Resultado líquido (último mês):')
        self.HMI.Label_Msg3.setStyleSheet('color: white')
        self.HMI.Label_Msg3.setFont(self.HMI.font12)
        self.HMI.Label_Msg3.setAlignment(Qt.AlignRight)

        self.HMI.TextBox_Msg3 = QLineEdit()
        self.HMI.TextBox_Msg3.setText(str(self.HMI.DBManager.GetResultadoLiquidoUltimoMes()))
        if self.HMI.ShowValues:
            self.HMI.TextBox_Msg3.setEchoMode(QLineEdit.Normal)
        else:
            self.HMI.TextBox_Msg3.setEchoMode(QLineEdit.Password)
        if float(self.HMI.TextBox_Msg3.text()) >= 0: self.HMI.TextBox_Msg3.setStyleSheet('QLineEdit {background-color: black; color: green;}')
        else: self.HMI.TextBox_Msg3.setStyleSheet('QLineEdit {background-color: black; color: red;}')
        self.HMI.TextBox_Msg3.setFont(self.HMI.font12)
        self.HMI.TextBox_Msg3.setEnabled(False)
        self.HMI.TextBox_Msg3.setAlignment(Qt.AlignLeft)

        self.HMI.Label_Msg4 = QLabel('Imposto de DT devido (último mês):')
        self.HMI.Label_Msg4.setStyleSheet('color: white')
        self.HMI.Label_Msg4.setFont(self.HMI.font12)
        self.HMI.Label_Msg4.setAlignment(Qt.AlignRight)

        self.HMI.TextBox_Msg4 = QLineEdit()
        self.HMI.TextBox_Msg4.setText(str(self.HMI.DBManager.GetImpostoDevidoDeDTUltimoMes()))
        if self.HMI.ShowValues:
            self.HMI.TextBox_Msg4.setEchoMode(QLineEdit.Normal)
        else:
            self.HMI.TextBox_Msg4.setEchoMode(QLineEdit.Password)
        self.HMI.TextBox_Msg4.setStyleSheet('QLineEdit {background-color: black; color: green;}')
        self.HMI.TextBox_Msg4.setFont(self.HMI.font12)
        self.HMI.TextBox_Msg4.setEnabled(False)
        self.HMI.TextBox_Msg4.setAlignment(Qt.AlignLeft)

        self.HMI.Button_RegistrarMudancas = QPushButton()
        self.HMI.Button_RegistrarMudancas.pressed.connect(lambda: self.OnButtonPressed('ExportIR'))
        self.HMI.Button_RegistrarMudancas.setFixedSize(40,40)
        self.HMI.Button_RegistrarMudancas.setIconSize(QSize(35, 35))
        self.HMI.Button_RegistrarMudancas.setIcon(QIcon("./images/ExportIR.png"))
        self.HMI.Button_RegistrarMudancas.setCursor(QCursor(Qt.PointingHandCursor))
        self.HMI.Button_RegistrarMudancas.setToolTip("Exportar Informe de Rendimentos")
        self.HMI.Button_RegistrarMudancas.setStyleSheet('QPushButton{background-color: gray} QToolTip {background-color: black; color: white;border: black solid 1px}')

        # self.HMI.Button_Imprimir = QPushButton()
        # self.HMI.Button_Imprimir.pressed.connect(lambda: self.OnButtonPressed('Gerar Excel de Tributação'))
        # self.HMI.Button_Imprimir.setFixedSize(40,40)
        # self.HMI.Button_Imprimir.setIconSize(QSize(35, 35))
        # self.HMI.Button_Imprimir.setIcon(QIcon("./images/printer.png"))
        # self.HMI.Button_Imprimir.setCursor(QCursor(Qt.PointingHandCursor))
        # self.HMI.Button_Imprimir.setStyleSheet('QPushButton {background-color: black; border: none; outline: none;} QToolTip {background-color: black; color: white;border: black solid 1px}')
        # self.HMI.Button_Imprimir.setToolTip("Gerar Excel de Tributação")

        self.HMI.Table_Tributacao_Header = QTableWidget()
        self.HMI.Table_Tributacao = QTableWidget()

        self.CreateTable_51_1()
        self.CreateTable_51_2()

        self.HMI.F_GLayout[self.HMI.GCount-1].addWidget(Background_1, 0, 0, 1, 10)
        self.HMI.F_GLayout[self.HMI.GCount-1].addWidget(Background_2, 1, 0, 3, 10)
        self.HMI.F_GLayout[self.HMI.GCount-1].addWidget(Background_3, 4, 0, 8, 10)
        self.HMI.F_GLayout[self.HMI.GCount-1].addWidget(self.HMI.Label_Titulo, 0, 0, 1, 10, Qt.AlignCenter)
        self.HMI.F_GLayout[self.HMI.GCount-1].addWidget(self.HMI.Label_Msg2, 1, 1, 1, 1, Qt.AlignCenter | Qt.AlignRight)
        self.HMI.F_GLayout[self.HMI.GCount-1].addWidget(self.HMI.TextBox_Msg2, 1, 2, 1, 1, Qt.AlignCenter | Qt.AlignLeft)
        self.HMI.F_GLayout[self.HMI.GCount-1].addWidget(self.HMI.Label_Msg3, 2, 1, 1, 1, Qt.AlignCenter | Qt.AlignRight)
        self.HMI.F_GLayout[self.HMI.GCount-1].addWidget(self.HMI.TextBox_Msg3, 2, 2, 1, 1, Qt.AlignCenter | Qt.AlignLeft)
        self.HMI.F_GLayout[self.HMI.GCount-1].addWidget(self.HMI.Label_Msg4, 3, 1, 1, 1, Qt.AlignCenter | Qt.AlignRight)
        self.HMI.F_GLayout[self.HMI.GCount-1].addWidget(self.HMI.TextBox_Msg4, 3, 2, 1, 1, Qt.AlignCenter | Qt.AlignLeft)
        self.HMI.F_GLayout[self.HMI.GCount-1].addWidget(self.HMI.Button_DTeST, 1, 5, 3, 1, Qt.AlignCenter)
        self.HMI.F_GLayout[self.HMI.GCount-1].addWidget(self.HMI.Button_CriptoeB3, 1, 6, 3, 1, Qt.AlignCenter)
        self.HMI.F_GLayout[self.HMI.GCount-1].addWidget(self.HMI.Button_RegistrarMudancas, 4, 0, 1, 10, Qt.AlignRight | Qt.AlignBottom)
        # self.HMI.F_GLayout[self.HMI.GCount-1].addWidget(self.HMI.Button_Imprimir, 5, 0, 1, 10, Qt.AlignRight | Qt.AlignTop)
        self.HMI.F_GLayout[self.HMI.GCount-1].addWidget(self.HMI.Table_Tributacao_Header, 5, 0, 1, 10, Qt.AlignHCenter | Qt.AlignBottom)
        self.HMI.F_GLayout[self.HMI.GCount-1].addWidget(self.HMI.Table_Tributacao, 6, 0, 6, 10, Qt.AlignHCenter | Qt.AlignTop)

    def CreateTable_51_1(self):
        if not self.HMI.DTeST_51 == "DTeST":
            HHeader = ['Mês/Ano',
                       'Resultado líquido',
                       'Imposto devido',
                       'Alíquota',
                       'Lucro mínimo taxável',
                       'Resultado bruto',
                       'Prejuízo acumulado',
                       'Resultado final']
            fakedata = [('','','','','','','','')]
        else:
            HHeader = ['Mês/Ano',
                       'Resultado líquido',
                       'Imposto devido',
                       'Resultado bruto',
                       'Prejuízo acumulado',
                       'Resultado final']
            fakedata = [('','','','','','')]

        self.HMI.Table_Tributacao_Header.setRowCount(1)
        self.HMI.Table_Tributacao_Header.setColumnCount(len(HHeader))

        self.HMI.Table_Tributacao_Header.setShowGrid(False)
        self.HMI.Table_Tributacao_Header.setFont(self.HMI.font16)
        self.HMI.Table_Tributacao_Header.setStyleSheet('''QTableWidget {background-color: rgb(0, 0, 0);
                                                                        color: white;
                                                                        border: 1px solid rgba(0, 0, 0, 0);}
                                                            QTableView {border-top: 2px solid white;
                                                                        border-right: 2px dashed white;
                                                                        border-left: 2px dashed white;}
                                                            QTableView::item {border-right: 1px dashed white}''') # gridline-color: rgba(255, 255, 255, 0)

        self.HMI.Table_Tributacao_Header.verticalHeader().hide()
        self.HMI.Table_Tributacao_Header.horizontalHeader().hide()

        for col in range(len(HHeader)):
            celula = QTableWidgetItem(str(HHeader[col]))
            celula.setFlags(Qt.ItemFlags(Qt.ItemIsSelectable))
            celula.setTextAlignment(Qt.AlignHCenter | Qt.AlignBottom)
            self.HMI.Table_Tributacao_Header.setItem(0, col, celula)

        self.HMI.Table_Tributacao_Header.setSizeAdjustPolicy(QAbstractScrollArea.AdjustToContents)
        self.HMI.Table_Tributacao_Header.setFixedHeight(int(self.HMI.frameGeometry().height()*1/20))
        self.HMI.Table_Tributacao_Header.resizeColumnsToContents()

        if not self.HMI.DTeST_51 == "DTeST":
            SizeCol0,SizeCol1,SizeCol2,SizeCol3,SizeCol4,SizeCol5,SizeCol6,SizeCol7 = self.HMI.GetSizeOfTableColumns(fakedata, HHeader)
            self.HMI.Table_Tributacao_Header.adjustSize()
            self.HMI.Table_Tributacao_Header.setColumnWidth(0, SizeCol0)
            self.HMI.Table_Tributacao_Header.setColumnWidth(1, SizeCol1)
            self.HMI.Table_Tributacao_Header.setColumnWidth(2, SizeCol2)
            self.HMI.Table_Tributacao_Header.setColumnWidth(3, SizeCol3)
            self.HMI.Table_Tributacao_Header.setColumnWidth(4, SizeCol4)
            self.HMI.Table_Tributacao_Header.setColumnWidth(5, SizeCol5)
            self.HMI.Table_Tributacao_Header.setColumnWidth(6, SizeCol6)
            self.HMI.Table_Tributacao_Header.setColumnWidth(7, SizeCol7)
        else:
            SizeCol0,SizeCol1,SizeCol2,SizeCol3,SizeCol4,SizeCol5 = self.HMI.GetSizeOfTableColumns(fakedata, HHeader)
            self.HMI.Table_Tributacao_Header.adjustSize()
            self.HMI.Table_Tributacao_Header.setColumnWidth(0, SizeCol0)
            self.HMI.Table_Tributacao_Header.setColumnWidth(1, SizeCol1)
            self.HMI.Table_Tributacao_Header.setColumnWidth(2, SizeCol2)
            self.HMI.Table_Tributacao_Header.setColumnWidth(3, SizeCol3)
            self.HMI.Table_Tributacao_Header.setColumnWidth(4, SizeCol4)
            self.HMI.Table_Tributacao_Header.setColumnWidth(5, SizeCol5)

    def CreateTable_51_2(self):
        self.HMI.Table_Tributacao.setSizeAdjustPolicy(QAbstractScrollArea.AdjustToContents)
        self.HMI.Table_Tributacao.setFixedHeight(int(self.HMI.frameGeometry().height()*6/20*1.4))
        self.HMI.Table_Tributacao.setShowGrid(False)
        self.HMI.Table_Tributacao.setStyleSheet('''QTableWidget {background-color: rgb(0, 0, 0);
                                                                 color: white;
                                                                 border: 1px solid rgba(0, 0, 0, 0);}
                                                   QTableView {border-bottom: 2px dashed white;
                                                               border-right: 1px solid white;
                                                               border-left: 1px solid white;}
                                                   QTableView::item {border-bottom: 1px dashed white;border-right: 1px dashed white}''') # gridline-color: rgba(255, 255, 255, 0)
        self.HMI.Table_Tributacao.verticalHeader().hide()
        self.HMI.Table_Tributacao.horizontalHeader().hide()

        #Create Thread
        if not self.HMI.DTeST_51 == "DTeST":
            HHeader = ['Mês/Ano',
                       'Resultado líquido',
                       'Imposto devido',
                       'Alíquota',
                       'Lucro mínimo taxável',
                       'Resultado bruto',
                       'Prejuízo acumulado',
                       'Resultado final']
            if int(self.HMI.frameGeometry().width()*6/20*1.4) < 510:
                self.HMI.Table_Tributacao_Header.setFont(self.HMI.font10)
                self.HMI.Table_Tributacao.setFont(self.HMI.font10)
            else:
                self.HMI.Table_Tributacao_Header.setFont(self.HMI.font14)
                self.HMI.Table_Tributacao.setFont(self.HMI.font14)
        else:
            HHeader = ['Mês/Ano',
                       'Resultado líquido',
                       'Imposto devido',
                       'Resultado bruto',
                       'Prejuízo acumulado',
                       'Resultado final']
            self.HMI.Table_Tributacao_Header.setFont(self.HMI.font14)
            self.HMI.Table_Tributacao.setFont(self.HMI.font14)
        self.data_51_1 = self.HMI.DBManager.GetAllTributacoes()
        self.data_51_1 = self.data_51_1[::-1]
        self.HMI.Table_Tributacao.setRowCount(len(self.data_51_1))
        self.HMI.Table_Tributacao.setColumnCount(len(HHeader))
        for i, item in enumerate(self.data_51_1):
            for j in range(len(item)):
                it = self.HMI.Table_Tributacao.item(i, j)
                if self.HMI.ShowValues:
                    if j == 0: text = datetime.strptime(item[j], '%Y-%m-%d %H:%M:%S').strftime('%Y-%m')
                    elif '%' in str(item[j]): text = str(round(item[j],2))+'%'
                    else: text = str(round(abs(item[j]),2))
                    it = QTableWidgetItem(text)
                    it.setFlags(Qt.ItemFlags(Qt.ItemIsSelectable)) # Desabilita selecionar o item
                    it.setFlags(it.flags() & ~Qt.ItemIsEditable) # Desabilita a edição do item
                else:
                    # Filtrar colunas que não serão mostradas
                    if not self.HMI.DTeST_51 == "DTeST":
                        if j in [1,2,5,6,7]: text = '***'
                        elif j == 0: text = datetime.strptime(item[j], '%Y-%m-%d %H:%M:%S').strftime('%Y-%m')
                        else: text = str(item[j])
                    else:
                        if not j == 0: text = '***'
                        else: text = datetime.strptime(item[j], '%Y-%m-%d %H:%M:%S').strftime('%Y-%m')
                    it = QTableWidgetItem(text)
                    # Desabilitar todos os itens
                    it.setFlags(Qt.ItemFlags(Qt.ItemIsSelectable)) # Desabilita selecionar o item
                    it.setFlags(it.flags() & ~Qt.ItemIsEditable) # Desabilita a edição do item
                if j == 0: pass
                elif not '%' in str(item[j]) and float(item[j]) < 0: it.setForeground(QBrush(QColor('orange')))
                else: it.setForeground(QBrush(QColor(0, 255, 0)))

                it.setTextAlignment(Qt.AlignCenter)
                self.HMI.Table_Tributacao.setItem(i, j, it)


        if not self.HMI.DTeST_51 == "DTeST":
            SizeCol0,SizeCol1,SizeCol2,SizeCol3,SizeCol4,SizeCol5,SizeCol6,SizeCol7 = self.HMI.GetSizeOfTableColumns(self.data_51_1, HHeader)

            self.HMI.Table_Tributacao_Header.adjustSize()
            self.HMI.Table_Tributacao_Header.setColumnWidth(0, SizeCol0)
            self.HMI.Table_Tributacao_Header.setColumnWidth(1, SizeCol1)
            self.HMI.Table_Tributacao_Header.setColumnWidth(2, SizeCol2)
            self.HMI.Table_Tributacao_Header.setColumnWidth(3, SizeCol3)
            self.HMI.Table_Tributacao_Header.setColumnWidth(4, SizeCol4)
            self.HMI.Table_Tributacao_Header.setColumnWidth(5, SizeCol5)
            self.HMI.Table_Tributacao_Header.setColumnWidth(6, SizeCol6)
            self.HMI.Table_Tributacao_Header.setColumnWidth(7, SizeCol7)

            self.HMI.Table_Tributacao.adjustSize()
            self.HMI.Table_Tributacao.setColumnWidth(0, SizeCol0)
            self.HMI.Table_Tributacao.setColumnWidth(1, SizeCol1)
            self.HMI.Table_Tributacao.setColumnWidth(2, SizeCol2)
            self.HMI.Table_Tributacao.setColumnWidth(3, SizeCol3)
            self.HMI.Table_Tributacao.setColumnWidth(4, SizeCol4)
            self.HMI.Table_Tributacao.setColumnWidth(5, SizeCol5)
            self.HMI.Table_Tributacao.setColumnWidth(6, SizeCol6)
            self.HMI.Table_Tributacao.setColumnWidth(7, SizeCol7)
        else:
            SizeCol0,SizeCol1,SizeCol2,SizeCol3,SizeCol4,SizeCol5 = self.HMI.GetSizeOfTableColumns(self.data_51_1, HHeader)

            self.HMI.Table_Tributacao_Header.adjustSize()
            self.HMI.Table_Tributacao_Header.setColumnWidth(0, SizeCol0)
            self.HMI.Table_Tributacao_Header.setColumnWidth(1, SizeCol1)
            self.HMI.Table_Tributacao_Header.setColumnWidth(2, SizeCol2)
            self.HMI.Table_Tributacao_Header.setColumnWidth(3, SizeCol3)
            self.HMI.Table_Tributacao_Header.setColumnWidth(4, SizeCol4)
            self.HMI.Table_Tributacao_Header.setColumnWidth(5, SizeCol5)

            self.HMI.Table_Tributacao.adjustSize()
            self.HMI.Table_Tributacao.setColumnWidth(0, SizeCol0)
            self.HMI.Table_Tributacao.setColumnWidth(1, SizeCol1)
            self.HMI.Table_Tributacao.setColumnWidth(2, SizeCol2)
            self.HMI.Table_Tributacao.setColumnWidth(3, SizeCol3)
            self.HMI.Table_Tributacao.setColumnWidth(4, SizeCol4)
            self.HMI.Table_Tributacao.setColumnWidth(5, SizeCol5)

    def CalcularTributacao(self):
        self.HMI.DBManager.LW.Label_Titulo.setText("Calculando tributação.\nPor favor aguarde...\n")
        def GenerateListOfMonthsYears(DataInicial, DataFinal):
            Resposta = [DataInicial]
            UmMes = dateutil.relativedelta.relativedelta(months=1)
            while not max(Resposta) == DataFinal:
                Resposta.append(max(Resposta)+UmMes)
            return Resposta
        def GetFirstDate(TabelasTributacao):
            Datas = []
            for TabelaTibutacao_Nome, TabelasOPs_Nome in TabelasTributacao:
                for TabelaOPs_Nome in TabelasOPs_Nome:
                    # Ler tabela
                    TabelaOPs, cursor = self.HMI.DBManager.GetDataDB("SELECT * FROM "+TabelaOPs_Nome)
                    # Selecionar apenas as datas
                    TabelaOPs = [t[0] for t in TabelaOPs]
                    # Adicionar no vetor geral
                    for item in TabelaOPs:
                        Datas.append(item)
            if len(Datas) > 0:
                return datetime.strptime(min(Datas), '%Y-%m-%d %H:%M:%S').replace(day=1, hour=0, minute=0, second=0, microsecond=0) # Retorna a menor data em datetime
            return datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        self.HMI.DBManager.LW.Label_Titulo.setText("Calculando tributação.\nPor favor aguarde...")
        cursor = self.HMI.DBManager.conn.cursor()
        # Definir tabelas de tributação
        retorno, cursor = self.HMI.DBManager.GetDataDB("SELECT name FROM sqlite_master WHERE type='table';")
        retorno = list(map(lambda x: x[0],retorno))
        T1 = list(filter(lambda x: x.find('OPs¨')>-1 and x.find("Bolsa¨DT")>-1, retorno))
        T2 = list(filter(lambda x: x.find('OPs¨')>-1 and x.find("Bolsa¨ST")>-1, retorno))
        T3 = []#list(filter(lambda x: x.find('Operações¨')>-1 and x.find("Cripto¨DT")>-1, retorno))
        T4 = []#list(filter(lambda x: x.find('Operações¨')>-1 and x.find("Cripto¨ST")>-1, retorno))
        TabelasTributacao = [('TributaçãoDTBolsa',T1),('TributaçãoSTBolsa',T2),('TributaçãoDTCripto',T3),('TributaçãoSTCripto',T4)]
        # Definir data inicial da tabela
        DataInicial = GetFirstDate(TabelasTributacao)
        # Definir data final da tabela
        DataFinal = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        MesesAnos = GenerateListOfMonthsYears(DataInicial, DataFinal)
        loading = 1
        self.HMI.DBManager.LW.pbar.setValue(loading)
        # Para cada tabela de tributação, faça
        for TabelaTibutacao_Nome, TabelasOPs_Nome in TabelasTributacao:
            loading += 1
            self.HMI.DBManager.LW.pbar.setValue(loading)
            # Para cada mês/ano, faça
            content = []
            for i, MesAno in enumerate(MesesAnos):
                loading += 23/len(MesesAnos)
                self.HMI.DBManager.LW.pbar.setValue(int(loading))
                self.HMI.DBManager.LW.Label_Titulo.setText("Calculando tributação.\nPor favor aguarde...\n\n"+str(TabelaTibutacao_Nome)+": "+str(MesAno.strftime("%Y/%m")))
                # counter = 0
                # while counter < 1000000/len(MesesAnos): # Tempo pra ver a animação da barra de carregamento ;)
                #     counter+=1
                #### Perfumaria
                if len(MesesAnos) < 20: time.sleep(1/60) # Perfumaria. Comentar linha se preferir performance
                if i == 0: # Primeira linha
                    # Definir o resultado bruto de cada linha
                    ResultadoBruto = 0
                    # Para cada Tabela de Operações de ST ou DT, faça
                    for TabelaOPs_Nome in TabelasOPs_Nome:
                        # Ler tabela
                        TabelaOPs, cursor = self.HMI.DBManager.GetDataDB("SELECT * FROM "+TabelaOPs_Nome)
                        if len(TabelaOPs)>0:
                            # Filtrar linhas que estão dentro do intervalo de datas
                            TabelaOPs = list(filter(lambda x: datetime.strptime(x[0], '%Y-%m-%d %H:%M:%S').replace(day=1, hour=0, minute=0, second=0, microsecond=0) == MesAno, TabelaOPs))
                            # Filtrar linhas que têm resultado diferente de zero
                            TabelaOPs = list(filter(lambda x: isinstance(x[12], float), TabelaOPs))
                            # Somar todos os resultados
                            TabelaOPs = [t[12] for t in TabelaOPs]
                            ResultadoBruto += sum(TabelaOPs)
                    # Definir 'Prejuízo acumulado'
                    PrejuizoAcumulado = 0
                    # Definir 'Resultado final'
                    ResultadoFinal = 0
                    if not ResultadoBruto == 0:
                        ResultadoFinal = ResultadoBruto + PrejuizoAcumulado
                    # Definir 'Alíquota'
                    Aliquota = 20 if 'DT' in TabelaTibutacao_Nome else 15
                    # Definir 'Lucro mínimo taxável'
                    LucroMinimoTaxavel = 20000
                    # Definir 'Imposto devido'
                    if ResultadoFinal >= LucroMinimoTaxavel:
                        ImpostoDevido = Aliquota*ResultadoFinal/100
                    else:
                        ImpostoDevido = 0
                    # Definir 'Resultado líquido'
                    ResultadoLiquido = ResultadoFinal - ImpostoDevido
                    content.append([('MesAno', MesAno),
                                    ('Resultado', round(ResultadoBruto,2)),
                                    ('PrejuizoAcumulado', round(PrejuizoAcumulado,2)),
                                    ('ResultadoFinal', round(ResultadoFinal,2)),
                                    ('Alíquota', Aliquota),
                                    ('LucroMínimoTaxável', LucroMinimoTaxavel),
                                    ('ImpostoDevido', round(ImpostoDevido,2)),
                                    ('ResultadoLíquido', round(ResultadoLiquido,2))])
                else:
                    # Definir o resultado bruto de cada linha
                    ResultadoBruto = 0
                    # Para cada Tabela de Operações de ST ou DT, faça
                    for TabelaOPs_Nome in TabelasOPs_Nome:
                        # Ler tabela
                        TabelaOPs, cursor = self.HMI.DBManager.GetDataDB("SELECT * FROM "+TabelaOPs_Nome)
                        if len(TabelaOPs)>0:
                            # Filtrar linhas que estão dentro do intervalo de datas
                            TabelaOPs = list(filter(lambda x: datetime.strptime(x[0], '%Y-%m-%d %H:%M:%S').replace(day=1, hour=0, minute=0, second=0, microsecond=0) == MesAno, TabelaOPs))
                            # Filtrar linhas que têm resultado diferente de zero
                            TabelaOPs = list(filter(lambda x: isinstance(x[12], float), TabelaOPs))
                            # Somar todos os resultados
                            TabelaOPs = [t[12] for t in TabelaOPs]
                            ResultadoBruto += sum(TabelaOPs)
                    # Definir 'Prejuízo acumulado'
                    PrejuizoAcumulado = 0
                    if content[-1][1][1] == 0: # Se não houve resultado no mês anterior, carrega o prejuízo
                        PrejuizoAcumulado = content[-1][2][1]
                    elif content[-1][3][1] < 0:
                        PrejuizoAcumulado = content[-1][3][1]
                    # Definir 'Resultado final'
                    ResultadoFinal = 0
                    if not ResultadoBruto == 0:
                        ResultadoFinal = ResultadoBruto + PrejuizoAcumulado
                    # Definir 'Alíquota'
                    Aliquota = content[-1][4][1]
                    # Definir 'Lucro mínimo taxável'
                    LucroMinimoTaxavel = content[-1][5][1]
                    # Definir 'Imposto devido'
                    if ResultadoFinal >= LucroMinimoTaxavel:
                        ImpostoDevido = Aliquota*ResultadoFinal/100
                    else:
                        ImpostoDevido = 0
                    # Definir 'Resultado líquido'
                    ResultadoLiquido = ResultadoFinal - ImpostoDevido
                    content.append([('MesAno', MesAno),
                                ('Resultado', round(ResultadoBruto,2)),
                                ('PrejuizoAcumulado', round(PrejuizoAcumulado,2)),
                                ('ResultadoFinal', round(ResultadoFinal,2)),
                                ('Alíquota', Aliquota),
                                ('LucroMínimoTaxável', LucroMinimoTaxavel),
                                ('ImpostoDevido', round(ImpostoDevido,2)),
                                ('ResultadoLíquido', round(ResultadoLiquido,2))])
            # print(TabelaTibutacao_Nome, ":\n", content) # Debug
            self.HMI.DBManager.ModifyDB("DELETE FROM "+TabelaTibutacao_Nome)
            for row in content: # Overwrite Tables
                self.HMI.DBManager.AddRowInCurrentDB(TabelaTibutacao_Nome, row)
            self.HMI.DBManager.LW.pbar.setValue(100)
        self.HMI.FLAG = False

    def CalcularTributacao_SemThread(self):
        def GenerateListOfMonthsYears(DataInicial, DataFinal):
            Resposta = [DataInicial]
            UmMes = dateutil.relativedelta.relativedelta(months=1)
            while not max(Resposta) == DataFinal:
                Resposta.append(max(Resposta)+UmMes)
            return Resposta
        def GetFirstDate(TabelasTributacao):
            Datas = []
            for TabelaTibutacao_Nome, TabelasOPs_Nome in TabelasTributacao:
                for TabelaOPs_Nome in TabelasOPs_Nome:
                    # Ler tabela
                    TabelaOPs, cursor = self.HMI.DBManager.GetDataDB("SELECT * FROM "+TabelaOPs_Nome)
                    # Selecionar apenas as datas
                    TabelaOPs = [t[0] for t in TabelaOPs]
                    # Adicionar no vetor geral
                    for item in TabelaOPs:
                        Datas.append(item)
            if len(Datas) > 0:
                return datetime.strptime(min(Datas), '%Y-%m-%d %H:%M:%S').replace(day=1, hour=0, minute=0, second=0, microsecond=0) # Retorna a menor data em datetime
            return datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        cursor = self.HMI.DBManager.conn.cursor()
        # Definir tabelas de tributação
        retorno, cursor = self.HMI.DBManager.GetDataDB("SELECT name FROM sqlite_master WHERE type='table';")
        retorno = list(map(lambda x: x[0],retorno))
        T1 = list(filter(lambda x: x.find('OPs¨')>-1 and x.find("Bolsa¨DT")>-1, retorno))
        T2 = list(filter(lambda x: x.find('OPs¨')>-1 and x.find("Bolsa¨ST")>-1, retorno))
        T3 = []#list(filter(lambda x: x.find('Operações¨')>-1 and x.find("Cripto¨DT")>-1, retorno))
        T4 = []#list(filter(lambda x: x.find('Operações¨')>-1 and x.find("Cripto¨ST")>-1, retorno))
        TabelasTributacao = [('TributaçãoDTBolsa',T1),('TributaçãoSTBolsa',T2),('TributaçãoDTCripto',T3),('TributaçãoSTCripto',T4)]
        # Definir data inicial da tabela
        DataInicial = GetFirstDate(TabelasTributacao)
        # Definir data final da tabela
        DataFinal = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        MesesAnos = GenerateListOfMonthsYears(DataInicial, DataFinal)
        # Para cada tabela de tributação, faça
        for TabelaTibutacao_Nome, TabelasOPs_Nome in TabelasTributacao:
            # Para cada mês/ano, faça
            content = []
            for i, MesAno in enumerate(MesesAnos):
                if i == 0: # Primeira linha
                    # Definir o resultado bruto de cada linha
                    ResultadoBruto = 0
                    # Para cada Tabela de Operações de ST ou DT, faça
                    for TabelaOPs_Nome in TabelasOPs_Nome:
                        # Ler tabela
                        TabelaOPs, cursor = self.HMI.DBManager.GetDataDB("SELECT * FROM "+TabelaOPs_Nome)
                        if len(TabelaOPs)>0:
                            # Filtrar linhas que estão dentro do intervalo de datas
                            TabelaOPs = list(filter(lambda x: datetime.strptime(x[0], '%Y-%m-%d %H:%M:%S').replace(day=1, hour=0, minute=0, second=0, microsecond=0) == MesAno, TabelaOPs))
                            # Filtrar linhas que têm resultado diferente de zero
                            TabelaOPs = list(filter(lambda x: isinstance(x[12], float), TabelaOPs))
                            # Somar todos os resultados
                            TabelaOPs = [t[12] for t in TabelaOPs]
                            ResultadoBruto += sum(TabelaOPs)
                    # Definir 'Prejuízo acumulado'
                    PrejuizoAcumulado = 0
                    # Definir 'Resultado final'
                    ResultadoFinal = 0
                    if not ResultadoBruto == 0:
                        ResultadoFinal = ResultadoBruto + PrejuizoAcumulado
                    # Definir 'Alíquota'
                    Aliquota = 20 if 'DT' in TabelaTibutacao_Nome else 15
                    # Definir 'Lucro mínimo taxável'
                    LucroMinimoTaxavel = 20000
                    # Definir 'Imposto devido'
                    if ResultadoFinal >= LucroMinimoTaxavel:
                        ImpostoDevido = Aliquota*ResultadoFinal/100
                    else:
                        ImpostoDevido = 0
                    # Definir 'Resultado líquido'
                    ResultadoLiquido = ResultadoFinal - ImpostoDevido
                    content.append([('MesAno', MesAno),
                                    ('Resultado', round(ResultadoBruto,2)),
                                    ('PrejuizoAcumulado', round(PrejuizoAcumulado,2)),
                                    ('ResultadoFinal', round(ResultadoFinal,2)),
                                    ('Alíquota', Aliquota),
                                    ('LucroMínimoTaxável', LucroMinimoTaxavel),
                                    ('ImpostoDevido', round(ImpostoDevido,2)),
                                    ('ResultadoLíquido', round(ResultadoLiquido,2))])
                else:
                    # Definir o resultado bruto de cada linha
                    ResultadoBruto = 0
                    # Para cada Tabela de Operações de ST ou DT, faça
                    for TabelaOPs_Nome in TabelasOPs_Nome:
                        # Ler tabela
                        TabelaOPs, cursor = self.HMI.DBManager.GetDataDB("SELECT * FROM "+TabelaOPs_Nome)
                        if len(TabelaOPs)>0:
                            # Filtrar linhas que estão dentro do intervalo de datas
                            TabelaOPs = list(filter(lambda x: datetime.strptime(x[0], '%Y-%m-%d %H:%M:%S').replace(day=1, hour=0, minute=0, second=0, microsecond=0) == MesAno, TabelaOPs))
                            # Filtrar linhas que têm resultado diferente de zero
                            TabelaOPs = list(filter(lambda x: isinstance(x[12], float), TabelaOPs))
                            # Somar todos os resultados
                            TabelaOPs = [t[12] for t in TabelaOPs]
                            ResultadoBruto += sum(TabelaOPs)
                    # Definir 'Prejuízo acumulado'
                    PrejuizoAcumulado = 0
                    if content[-1][1][1] == 0: # Se não houve resultado no mês anterior, carrega o prejuízo
                        PrejuizoAcumulado = content[-1][2][1]
                    elif content[-1][3][1] < 0:
                        PrejuizoAcumulado = content[-1][3][1]
                    # Definir 'Resultado final'
                    ResultadoFinal = 0
                    if not ResultadoBruto == 0:
                        ResultadoFinal = ResultadoBruto + PrejuizoAcumulado
                    # Definir 'Alíquota'
                    Aliquota = content[-1][4][1]
                    # Definir 'Lucro mínimo taxável'
                    LucroMinimoTaxavel = content[-1][5][1]
                    # Definir 'Imposto devido'
                    if ResultadoFinal >= LucroMinimoTaxavel:
                        ImpostoDevido = Aliquota*ResultadoFinal/100
                    else:
                        ImpostoDevido = 0
                    # Definir 'Resultado líquido'
                    ResultadoLiquido = ResultadoFinal - ImpostoDevido
                    content.append([('MesAno', MesAno),
                                ('Resultado', round(ResultadoBruto,2)),
                                ('PrejuizoAcumulado', round(PrejuizoAcumulado,2)),
                                ('ResultadoFinal', round(ResultadoFinal,2)),
                                ('Alíquota', Aliquota),
                                ('LucroMínimoTaxável', LucroMinimoTaxavel),
                                ('ImpostoDevido', round(ImpostoDevido,2)),
                                ('ResultadoLíquido', round(ResultadoLiquido,2))])
            # print(TabelaTibutacao_Nome, ":\n", content) # Debug
            self.HMI.DBManager.ModifyDB("DELETE FROM "+TabelaTibutacao_Nome)
            for row in content: # Overwrite Tables
                self.HMI.DBManager.AddRowInCurrentDB(TabelaTibutacao_Nome, row)
        self.HMI.FLAG = False

    def OnButtonPressed(self, ButtonPressed):
        if ButtonPressed == 'ChangeDTeST':
            if self.HMI.DTeST == "": self.HMI.DTeST = "DTeST"
            elif self.HMI.DTeST == "DTeST": self.HMI.DTeST = "ST"
            elif self.HMI.DTeST == "ST": self.HMI.DTeST = "DT"
            elif self.HMI.DTeST == "DT": self.HMI.DTeST = ""

            if self.HMI.DTeST == "":
                self.HMI.Button_DTeST.setIcon(QIcon("./images/Visualizar Trades Brutos.png"))
                # Habilitar Botões de edição
                self.HMI.Button_AlterarOperacao.setEnabled(True)
                self.HMI.Button_DeletarOperacao.setEnabled(True)
            elif self.HMI.DTeST == "DTeST":
                self.HMI.Button_DTeST.setIcon(QIcon("./images/Visualizar Day Trades e Swing Trades.png"))
                # Desabilitar Botões de edição
                self.HMI.Button_AlterarOperacao.setEnabled(False)
                self.HMI.Button_DeletarOperacao.setEnabled(False)
            elif self.HMI.DTeST == "ST":
                self.HMI.Button_DTeST.setIcon(QIcon("./images/Visualizar Apenas ST.png"))
                # Desabilitar Botões de edição
                self.HMI.Button_AlterarOperacao.setEnabled(False)
                self.HMI.Button_DeletarOperacao.setEnabled(False)
            elif self.HMI.DTeST == "DT":
                self.HMI.Button_DTeST.setIcon(QIcon("./images/Visualizar Apenas DT.png"))
                # Desabilitar Botões de edição
                self.HMI.Button_AlterarOperacao.setEnabled(False)
                self.HMI.Button_DeletarOperacao.setEnabled(False)

            if self.HMI.PageID == '23':
                self.HMI.Table_Operacoes_Realizadas_com_Ativo.deleteLater()
                self.HMI.Table_Operacoes_Realizadas_com_Ativo = QTableWidget()
                self.HMI.HMI_Trades.HMI_Trades_Bolsa.CreateTable_23_2()
                self.HMI.F_GLayout[self.HMI.GCount-1].addWidget(self.HMI.Table_Operacoes_Realizadas_com_Ativo, 6, 0, 6, 10, Qt.AlignHCenter | Qt.AlignTop)

            if self.HMI.PageID == '40':
                self.HMI.Table_Operacoes_Realizadas.deleteLater()
                self.HMI.Table_Operacoes_Realizadas = QTableWidget()
                self.HMI.HMI_Trades.HMI_Trades_Criptomoedas.CreateTable_40_2()
                self.HMI.F_GLayout[self.HMI.GCount-1].addWidget(self.HMI.Table_Operacoes_Realizadas, 6, 0, 6, 10, Qt.AlignHCenter | Qt.AlignTop)

            elif self.HMI.PageID == '51':
                if self.HMI.DTeST_51 == "DTeST": self.HMI.DTeST_51 = "ST"
                elif self.HMI.DTeST_51 == "ST": self.HMI.DTeST_51 = "DT"
                elif self.HMI.DTeST_51 == "DT": self.HMI.DTeST_51 = "DTeST"

                if self.HMI.DTeST_51 == "DTeST":
                    self.HMI.Button_DTeST.setIcon(QIcon("./images/Visualizar Day Trades e Swing Trades.png"))
                elif self.HMI.DTeST_51 == "ST":
                    self.HMI.Button_DTeST.setIcon(QIcon("./images/Visualizar Apenas ST.png"))
                elif self.HMI.DTeST_51 == "DT":
                    self.HMI.Button_DTeST.setIcon(QIcon("./images/Visualizar Apenas DT.png"))
                self.HMI.Table_Tributacao_Header.deleteLater()
                self.HMI.Table_Tributacao_Header = QTableWidget()
                self.HMI.HMI_Tributacao.CreateTable_51_1()
                self.HMI.F_GLayout[self.HMI.GCount-1].addWidget(self.HMI.Table_Tributacao_Header, 5, 0, 1, 10, Qt.AlignHCenter | Qt.AlignBottom)
                self.HMI.Table_Tributacao.deleteLater()
                self.HMI.Table_Tributacao = QTableWidget()
                self.HMI.HMI_Tributacao.CreateTable_51_2()
                self.HMI.F_GLayout[self.HMI.GCount-1].addWidget(self.HMI.Table_Tributacao, 6, 0, 6, 10, Qt.AlignHCenter | Qt.AlignTop)

        elif ButtonPressed == 'ExportIR':
            def WriteHeader(worksheet, header):
                for column_idx, item in enumerate(header):
                    try:
                        worksheet.write(0, column_idx, item)
                    except Exception as e:
                        print('header: ',e)

            # 1. Ask IR year
            # 2. Ask filepath to be saved
            # 3. Get from DB last estoque of defined year for all shares, for all corretoras
            # 4. Save

            # 1. Ask IR year
            years = []
            for i, item in enumerate(self.data_51_1):
                for j in range(len(item)):
                    if j == 0:
                        years.append(datetime.strptime(item[j], '%Y-%m-%d %H:%M:%S').strftime('%Y'))
            years = sorted(set(years), reverse=True)
            year, answered = QInputDialog.getItem(self.HMI, 'Exportar IR', 'Ano de IR que deseja exportar:', years)

            if answered:
                # 2. Ask filepath to be saved
                if not self.HMI.filedialogIsOpen:
                    root = tkinter.Tk()
                    root.withdraw()  # Used to hide tkinter window
                    self.HMI.filedialogIsOpen = True
                    DB_Path = filedialog.askdirectory()  # Select folder path
                    self.HMI.setCursor(Qt.WaitCursor)
                    self.HMI.filedialogIsOpen = False
                    try:
                        if len(DB_Path) > 0:
                            # 3. Get from DB last estoque of defined year for all shares, for all corretoras
                            sheets, cursor = self.HMI.DBManager.GetDataDB("SELECT name FROM sqlite_master WHERE type='table';")
                            sheets = list(map(lambda x: x[0], sheets))
                            agora = datetime.now().strftime("%Y %m %d")
                            FileName = self.HMI.User + "_ImpostoDeRenda " + year + "_Gerado em " + agora + '.xlsx'
                            workbook = xlsxwriter.Workbook(DB_Path + "/" + FileName)
                            corretoras, corretorasCripto = self.HMI.DBManager.GetAllCorretoras_ComBolsaOuCripto()
                            corretoras.extend(corretorasCripto)
                            for corretora in corretoras:
                                row1=[]; Ativo = '...'; QuantidadeEmYear = 'NAN'; PreçoMedioEmYear = 'NAN'; CustoEmYear = 'NAN'; rows = []; pre_rows = [];

                                try: worksheet = workbook.add_worksheet(name=corretora.replace("¨Bolsa","").replace("¨Cripto","")[0:31]) # Create sheets
                                except Exception as e: print('creating worksheet ',worksheet,': ',e)

                                if "¨Bolsa" in corretora:
                                    SheetsToBeConsidered = list(filter(lambda x: corretora+"¨Bruto" in x, sheets))
                                    corretora = corretora.replace("¨Bolsa", "")
                                    CorretoraCoinCurrency = self.HMI.DBManager.GetCorretoraCoinCurrency(corretora, "Bolsa")
                                    header = ['Ativo', 'Quantidade em ' + year, 'Preço Médio (' + CorretoraCoinCurrency + ')']
                                    WriteHeader(worksheet, header)

                                    for Sheet in SheetsToBeConsidered:
                                        try:
                                            CONTENT, cursor = self.HMI.DBManager.GetDataDB('SELECT * FROM ' + Sheet + ' WHERE Data BETWEEN (?) AND (?);', [min(years) + '-01-01 00:00:00', year + '-12-31 23:59:59'])
                                            if len(CONTENT) == 0: break
                                            Ativo = Sheet.replace(corretora+"¨Bolsa¨Bruto","").replace("OPs¨","").replace("¨","")
                                            QuantidadeEmYear = CONTENT[-1][8]
                                            PreçoMedioEmYear = CONTENT[-1][11]
                                            rows.append([Ativo, QuantidadeEmYear, PreçoMedioEmYear])
                                        except Exception as e: print('writing rows in Bolsa ',[Ativo, QuantidadeEmYear, PreçoMedioEmYear],': ',e)

                                elif "¨Cripto" in corretora:
                                    corretora = corretora.replace("¨Cripto", "")
                                    sheet = "OPs¨" + corretora + "¨Cripto¨Refinado"
                                    CorretoraCoinCurrency = self.HMI.DBManager.GetCorretoraCoinCurrency(corretora, "Cripto")
                                    header = ['Ativo', 'Quantidade em ' + year, 'Preço Médio (' + CorretoraCoinCurrency + ')', 'Custo (' + CorretoraCoinCurrency + ')']
                                    WriteHeader(worksheet, header)

                                    CONTENT, cursor = self.HMI.DBManager.GetDataDB('SELECT * FROM ' + sheet + ' WHERE Data BETWEEN (?) AND (?);', [min(years) + '-01-01 00:00:00', year + '-12-31 23:59:59'])
                                    if len(CONTENT) == 0: break
                                    Headers = list(map(lambda x: x[0], cursor.description))[13:]
                                    for idx, Header in enumerate(Headers):
                                        if "EstoqueDe¨" in Header and CorretoraCoinCurrency in Header:
                                            try:
                                                Ativo = CorretoraCoinCurrency
                                                QuantidadeEmYear = CONTENT[-1][idx + 13] + CONTENT[-1][idx + 14] + CONTENT[-1][idx + 15]
                                                PreçoMedioEmYear = CONTENT[-1][idx + 17]
                                                CustoEmYear = CONTENT[-1][idx + 16]
                                                row1.append([Ativo, QuantidadeEmYear, PreçoMedioEmYear, CustoEmYear])
                                            except Exception as e: print('writing row1 in Cripto ',[Ativo, QuantidadeEmYear, PreçoMedioEmYear, CustoEmYear],': ',e)
                                        elif "EstoqueDe¨" in Header:
                                            try:
                                                Ativo = Header.replace("EstoqueDe¨","")
                                                QuantidadeEmYear = CONTENT[-1][idx + 13] + CONTENT[-1][idx + 14] + CONTENT[-1][idx + 15]
                                                PreçoMedioEmYear = CONTENT[-1][idx + 17]
                                                CustoEmYear = CONTENT[-1][idx + 16]
                                                pre_rows.append([Ativo, QuantidadeEmYear, PreçoMedioEmYear, CustoEmYear])
                                            except Exception as e: print('writing pre_rows in Cripto ',[Ativo, QuantidadeEmYear, PreçoMedioEmYear, CustoEmYear],': ',e)

                                    pre_rows = sorted(pre_rows, key=lambda x: x[3])
                                    if len(row1)>0: pre_rows.extend(row1)
                                    pre_rows = pre_rows[::-1]
                                    for idx in list(range(min(len(pre_rows),6))):
                                        try: rows.append(pre_rows[idx])
                                        except Exception as e: print('writing rows in Cripto ',[Ativo, QuantidadeEmYear, PreçoMedioEmYear, CustoEmYear],': ',e)
                                    if len(pre_rows) > 6:
                                        custos = sum(list(map(lambda x: x[3], pre_rows[6:])))
                                        rows.append(['Outros', "", "", custos])

                                for row_idx, row in enumerate(rows):
                                    for column_idx, item in enumerate(row):
                                        try:
                                            worksheet.write(row_idx+1, column_idx, item)
                                        except Exception as e:
                                            print(e)

                            workbook.close()

                            excel = win32.gencache.EnsureDispatch('Excel.Application')
                            wb = excel.Workbooks.Open(DB_Path + "/" + FileName)
                            for corretora in corretoras:
                                ws = wb.Worksheets(corretora.replace("¨Bolsa","").replace("¨Cripto","")[0:31])
                                ws.Columns.AutoFit()
                            wb.Save()
                            wb.Close(True)
                            # excel.Application.Quit() # Bugado
                    except Exception as e:
                        print(e)
                        try: workbook.close()
                        except: pass
                        try: wb.Close(True)
                        except: pass
                    finally:
                        self.HMI.unsetCursor()

        elif ButtonPressed == 'ChangeCriptoeB3':
            if self.HMI.PageID == '51':
                if self.HMI.CriptoeB3_51 == "CriptoeB3": self.HMI.CriptoeB3_51 = "B3"
                elif self.HMI.CriptoeB3_51 == "B3": self.HMI.CriptoeB3_51 = "Cripto"
                elif self.HMI.CriptoeB3_51 == "Cripto": self.HMI.CriptoeB3_51 = "CriptoeB3"

                if self.HMI.CriptoeB3_51 == "CriptoeB3":
                    self.HMI.Button_CriptoeB3.setIcon(QIcon("./images/Visualizar cripto e B3.png"))
                elif self.HMI.CriptoeB3_51 == "Cripto":
                    self.HMI.Button_CriptoeB3.setIcon(QIcon("./images/Visualizar Apenas cripto.png"))
                elif self.HMI.CriptoeB3_51 == "B3":
                    self.HMI.Button_CriptoeB3.setIcon(QIcon("./images/Visualizar Apenas B3.png"))
                self.HMI.Table_Tributacao_Header.deleteLater()
                self.HMI.Table_Tributacao_Header = QTableWidget()
                self.HMI.HMI_Tributacao.CreateTable_51_1()
                self.HMI.F_GLayout[self.HMI.GCount-1].addWidget(self.HMI.Table_Tributacao_Header, 5, 0, 1, 10, Qt.AlignHCenter | Qt.AlignBottom)
                self.HMI.Table_Tributacao.deleteLater()
                self.HMI.Table_Tributacao = QTableWidget()
                self.HMI.HMI_Tributacao.CreateTable_51_2()
                self.HMI.F_GLayout[self.HMI.GCount-1].addWidget(self.HMI.Table_Tributacao, 6, 0, 6, 10, Qt.AlignHCenter | Qt.AlignTop)

        # elif ButtonPressed == "Gerar Excel de Tributação":
        #     # MessageBox_Msg1 = QMessageBox.about(self,'Informação','Funcionalidade ainda não implementada.')
        #     if not self.filedialogIsOpen:
        #         MessageBox_Msg1 = QMessageBox.about(self,'Aviso','Para continuar a exportação,\nSALVE seu trabalho e FECHE todas\nas janelas de EXCEL abertas.')
        #         root = tkinter.Tk()
        #         root.withdraw() # Used to hide tkinter window
        #         self.filedialogIsOpen = True
        #         DB_Path = filedialog.askdirectory() # Select folder path
        #         self.setCursor(Qt.WaitCursor)
        #         self.filedialogIsOpen = False
        #         try:
        #             if len(DB_Path)>0:
        #                 agora = datetime.now().strftime("%Y %m %d")
        #                 FileName = self.User+"_database_"+agora+'.xlsx'
        #                 workbook = xlsxwriter.Workbook(DB_Path+"/"+FileName)
        #                 tables = ["TributaçãoDTBolsa", "TributaçãoSTBolsa", "TributaçãoDTCripto", "TributaçãoSTCripto"]
        #                 conn = sqlite3.connect(self.User+'.db')
        #                 cursor = conn.cursor()
        #                 for TableName in tables:
        #                     workbook.add_worksheet(name=TableName)
        #                     CONTENT, cursor = self.HMI.DBManager.GetDataDB('SELECT * FROM '+TableName)
        #                     headers = list(map(lambda x: x[0], cursor.description))
        #                     row1 = []
        #                     if len(CONTENT)>0: row1 = CONTENT[0]
        #                     for column_number, item in enumerate(row1):
        #                         worksheet.write(0, column_number, headers[column_number])

        #                     for row_number, row in enumerate(CONTENT):
        #                         for column_number, item in enumerate(row):
        #                             try: worksheet.write(row_number+1, column_number, item)
        #                             except:pass
        #                 workbook.close()

        #                 excel = win32.gencache.EnsureDispatch('Excel.Application')
        #                 wb = excel.Workbooks.Open(DB_Path+"/"+FileName)
        #                 for table in tables:
        #                     ws = wb.Worksheets(table[0:31])
        #                     ws.Columns.AutoFit()
        #                 wb.Save()
        #                 excel.Application.Quit()
        #         except:
        #             workbook.close()
        #         finally: self.unsetCursor()